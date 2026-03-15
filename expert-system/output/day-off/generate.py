#!/usr/bin/env python3
"""Generate unified day-off.xlsx workbook with test plan + all test suites."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

OUTPUT = "/home/v/Dev/ttt-expert-v1/expert-system/output/day-off/day-off.xlsx"

# ── Styles ──────────────────────────────────────────────────────────────────
ARIAL = "Arial"
HEADER_FONT = Font(name=ARIAL, bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
SUBTITLE_FONT = Font(name=ARIAL, bold=True, size=12)
BODY_FONT = Font(name=ARIAL, size=10)
LINK_FONT = Font(name=ARIAL, size=10, color="0563C1", underline="single")
BACK_LINK_FONT = Font(name=ARIAL, size=9, color="0563C1", underline="single")
ROW_EVEN = PatternFill("solid", fgColor="D6E4F0")
ROW_ODD = PatternFill("solid", fgColor="FFFFFF")
RISK_CRIT = PatternFill("solid", fgColor="FF6B6B")
RISK_HIGH = PatternFill("solid", fgColor="FFA07A")
RISK_MED = PatternFill("solid", fgColor="FFD700")
RISK_LOW = PatternFill("solid", fgColor="90EE90")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TAB_GREEN = "00B050"
TAB_BLUE = "4472C4"


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def style_data_row(ws, row, max_col, idx):
    fill = ROW_EVEN if idx % 2 == 0 else ROW_ODD
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Test Suite Definitions ──────────────────────────────────────────────────
# Each suite: (tab_name, title, cases_list)
# Each case: (id, title, preconditions, steps, expected, priority, type, req_ref, component, notes)

SUITES = []

# ══════════════════════════════════════════════════════════════════════════════
# TS-DO-Lifecycle — Request Lifecycle (Create/Approve/Reject/Delete/Edit)
# ══════════════════════════════════════════════════════════════════════════════
SUITES.append(("TS-DO-Lifecycle", "Day-Off Request Lifecycle", [
    ("TC-DO-001", "Create day-off transfer request for future holiday",
     "Logged in as employee. Production calendar has future public holiday (e.g. 2026-05-01 May Day). No existing day-off request for that date.",
     "1. Navigate to /vacation/my → Days off tab\n2. Click edit (pencil) icon on future holiday row with duration=0\n3. TransferDaysoffModal opens\n4. Select a valid working day as personal date (e.g. 2026-05-11)\n5. Click OK",
     "New day-off request created with status NEW. Row shows arrow format: '01.05.2026 → 11.05.2026'. Cancel (X) button appears. Approver auto-assigned (employee's manager or self for CPO).",
     "High", "Functional", "REQ-day-off", "frontend-day-off-module, day-off-service", ""),

    ("TC-DO-002", "Create day-off — CPO self-approval on creation",
     "Logged in as user with PROJECT role (CPO). Employee is their own manager.",
     "1. Create day-off transfer request via UI\n2. Check request status immediately",
     "Request created with status APPROVED (not NEW). Self-approval applied automatically. Manager becomes optional approver with ASKED status.",
     "High", "Functional", "", "day-off-service", "CPO self-approve pattern: approver=self, manager→optional"),

    ("TC-DO-003", "Create day-off — optional approvers assigned",
     "Logged in as employee with configured optional approvers.",
     "1. Create day-off request\n2. GET /v1/employee-dayOff/{id} to check response\n3. Inspect optionalApprovers array",
     "Optional approvers present with status ASKED. Each has individual approve/reject capability (non-blocking FYI-style).",
     "Medium", "Functional", "", "day-off-service", "employee_dayoff_approval table populated"),

    ("TC-DO-004", "Approve NEW request by primary approver",
     "Logged in as approver. Existing NEW day-off request for another employee.",
     "1. Navigate to /vacation/request → Days off rescheduling → Approval tab\n2. Click approve (checkmark) on NEW request\n3. Confirm approval",
     "Status changes to APPROVED. Two ledger entries created (credit for holiday + debit for personal day). Vacation days recalculated. Month norms updated.",
     "High", "Functional", "", "day-off-service", "Critical: verify 2 ledger entries + vacation recalc"),

    ("TC-DO-005", "Approve REJECTED request (re-approve)",
     "Existing REJECTED day-off request.",
     "1. PUT /v1/employee-dayOff/approve/{id}\n2. Check status and ledger",
     "Status changes from REJECTED to APPROVED. New ledger entries created. Previous orphaned ledger entries (if any from prior approve) remain — compounding.",
     "High", "Functional", "", "day-off-service", "BUG-DO-6 context: orphaned entries accumulate"),

    ("TC-DO-006", "Reject NEW request",
     "Logged in as approver. Existing NEW day-off request.",
     "1. Click reject icon on NEW request in Approval tab\n2. Confirm rejection",
     "Status changes to REJECTED. No ledger entries created (none existed). Employee notified.",
     "High", "Functional", "", "day-off-service", "Clean reject — no ledger to revert"),

    ("TC-DO-007", "Reject APPROVED request (personalDate >= report period start)",
     "Existing APPROVED request with personalDate in future (>= current report period start).",
     "1. PUT /v1/employee-dayOff/reject/{id}\n2. Check status and ledger entries",
     "Status changes to REJECTED. BUG: Ledger entries from approval NOT reverted — orphaned credit/debit entries remain in employee_dayoff table.",
     "High", "Bug verification", "BUG-DO-6", "day-off-service", "Known bug: reject does not undo ledger"),

    ("TC-DO-008", "Reject APPROVED request (personalDate < report period start) — blocked",
     "Existing APPROVED request with personalDate before current report period start.",
     "1. PUT /v1/employee-dayOff/reject/{id}\n2. Check response",
     "Rejection blocked — personalDate is in a closed period. Error returned.",
     "High", "Negative", "", "day-off-service", "Guard: personalDate >= report period start"),

    ("TC-DO-009", "Delete NEW request by owner",
     "Logged in as request owner. Existing NEW day-off request.",
     "1. Click cancel (X) button on NEW request row\n2. Confirm deletion",
     "Status set to DELETED. Row disappears from default view. No ledger entries affected (none existed).",
     "High", "Functional", "", "frontend-day-off-module", "Cancel postponement = DELETE endpoint"),

    ("TC-DO-010", "Delete APPROVED request (personalDate in future)",
     "Existing APPROVED request with personalDate >= period start (future).",
     "1. DELETE /v1/employee-dayOff/{id}\n2. Check status",
     "Status set to DELETED. Ledger entries from approval remain orphaned (not cleaned up on delete).",
     "Medium", "Functional", "", "day-off-service", "Delete allowed when personalDate >= period start"),

    ("TC-DO-011", "Delete APPROVED request (personalDate before period) — blocked",
     "Existing APPROVED request with personalDate < current period start.",
     "1. DELETE /v1/employee-dayOff/{id}\n2. Check response",
     "Deletion blocked — cannot delete APPROVED request in closed period.",
     "High", "Negative", "", "day-off-service", "Guard: personalDate >= period start OR status != APPROVED"),

    ("TC-DO-012", "Change primary approver",
     "Existing request (any status). Logged in as current approver.",
     "1. PUT /v1/employee-dayOff/change-approver/{id}/{newLogin}\n2. Check request approver field",
     "New approver assigned. Old approver moved to optional approvers list. Old approver removed from optional list if already present.",
     "Medium", "Functional", "", "day-off-service", ""),

    ("TC-DO-013", "Change approver — old becomes optional, new removed from optional",
     "Request has optional approver 'userB'. Primary approver is 'userA'.",
     "1. PUT /v1/employee-dayOff/change-approver/{id}/userB\n2. Inspect optionalApprovers",
     "userB becomes primary approver. userA becomes optional. userB removed from optional list (no duplicate).",
     "Medium", "Functional", "", "day-off-service", "Bidirectional swap logic"),

    ("TC-DO-014", "Optional approver votes APPROVED",
     "Request has optional approver with status ASKED.",
     "1. As optional approver, PATCH approval status to APPROVED\n2. Check employee_dayoff_approval table",
     "Optional approval status changes to APPROVED. Main request status unchanged (non-blocking). ApproveBar shows updated vote count.",
     "Medium", "Functional", "", "day-off-service", "FYI-style: does not affect main approval"),

    ("TC-DO-015", "Optional approver votes REJECTED (non-blocking)",
     "Request has optional approver with status ASKED.",
     "1. As optional approver, PATCH approval status to REJECTED\n2. Check main request status",
     "Optional approval status = REJECTED. Main request status unchanged — optional rejection does not block approval.",
     "Medium", "Functional", "", "day-off-service", "Non-blocking: only primary approver controls lifecycle"),

    ("TC-DO-016", "Edit request — change personalDate",
     "Existing NEW request owned by current user.",
     "1. PATCH /v1/employee-dayOff/{id} with new personalDate\n2. Check updated fields",
     "personalDate updated. Optional approvals reset to ASKED. Only personalDate is changeable via PATCH.",
     "High", "Functional", "", "day-off-service", "Edit resets optional approvals"),

    ("TC-DO-017", "Edit request — attempt to change publicDate (immutable)",
     "Existing request.",
     "1. PATCH /v1/employee-dayOff/{id} with different publicDate/originalDate\n2. Check response",
     "publicDate/originalDate unchanged — only personalDate is editable. Other fields silently ignored or rejected.",
     "Medium", "Negative", "", "day-off-service", "original_date is immutable"),

    ("TC-DO-018", "System rejection — period change triggers bulk reject",
     "Multiple NEW requests exist with last_approved_date = target date. Approve period changed.",
     "1. Admin changes approve period for office\n2. PeriodChangedEventHandler fires\n3. Check affected requests",
     "All NEW requests with matching last_approved_date → status REJECTED. NOTIFY_DAYOFF_AUTODELETE_TO_EMPLOYEE sent. APPROVED requests NOT affected.",
     "High", "Integration", "", "day-off-service, accounting-backend", "Path C: only NEW status affected"),

    ("TC-DO-019", "Create request on past holiday",
     "Production calendar has holiday in the past (e.g., 2026-02-23).",
     "1. POST /v1/employee-dayOff with publicDate = 2026-02-23\n2. Check response",
     "Request created. Past holidays are valid source dates. TransferDaysoffModal minDate = yesterday (not originalDate) for past holidays.",
     "Medium", "Boundary", "", "day-off-service", "Past holidays: minDate shifts to yesterday"),

    ("TC-DO-020", "Full lifecycle: create → approve → verify ledger → vacation balance",
     "Clean state employee with known vacation balance.",
     "1. Create day-off request (POST)\n2. Approve (PUT /approve/{id})\n3. GET employee vacation days\n4. Query employee_dayoff table for 2 new entries\n5. Query vacation balance",
     "Ledger: entry 1 = credit (holiday, duration from calendar/norm), entry 2 = debit (personal day, duration from request). Vacation balance: net zero change (credit +1, debit -1 cancel out).",
     "High", "Integration", "", "day-off-service, vacation-service", "Confirmed: vacation balance unaffected by day-offs"),
]))

# ══════════════════════════════════════════════════════════════════════════════
# TS-DO-Ledger — Credit/Debit Ledger Mechanics
# ══════════════════════════════════════════════════════════════════════════════
SUITES.append(("TS-DO-Ledger", "Day-Off Ledger Mechanics", [
    ("TC-DO-021", "Approve writes 2 ledger entries (credit + debit)",
     "NEW day-off request exists. employee_dayoff table has no entries for this request's dates.",
     "1. PUT /v1/employee-dayOff/approve/{id}\n2. SELECT * FROM employee_dayoff WHERE employee = {empId} AND (personal_date = {personalDate} OR personal_date = {lastApprovedDate})",
     "Two new rows in employee_dayoff: (1) lastApprovedDate slot — credit reversal with duration from calendar/norm, (2) personalDate slot — debit with duration+reason from request.",
     "High", "Data integrity", "", "day-off-service", "Upsert pattern: onDuplicateKeyUpdate"),

    ("TC-DO-022", "Credit entry — duration from existing ledger",
     "Employee already has a ledger entry for the lastApprovedDate.",
     "1. Approve request\n2. Check credit entry duration source",
     "Credit entry duration = existing ledger entry's duration (not calendar). Lookup priority: existing ledger → calendar → reportingNorm fallback.",
     "Medium", "Data integrity", "", "day-off-service", "Three-tier duration source resolution"),

    ("TC-DO-023", "Credit entry — duration from calendar (no prior ledger)",
     "No existing ledger entry for lastApprovedDate. Calendar has entry for that date.",
     "1. Approve request\n2. Check credit entry duration",
     "Credit entry duration = calendar day's duration for that date.",
     "Medium", "Data integrity", "", "day-off-service", ""),

    ("TC-DO-024", "Credit entry — duration from reportingNorm fallback",
     "No existing ledger AND no calendar entry for lastApprovedDate.",
     "1. Approve request\n2. Check credit entry duration",
     "Credit entry duration = office reportingNorm (default 8 for full-day offices).",
     "Low", "Boundary", "", "day-off-service", "Fallback tier — rare scenario"),

    ("TC-DO-025", "Ledger duration=0 — full day-off (debit)",
     "Request with duration=0 (standard full day-off).",
     "1. Approve request\n2. SELECT duration FROM employee_dayoff WHERE personal_date = {personalDate}",
     "Debit entry: duration=0. Represents taking full day off from work.",
     "Medium", "Data integrity", "", "day-off-service", "Most common: 2454 of 5334 total entries"),

    ("TC-DO-026", "Ledger duration=8 — full working day credit",
     "Ledger entry for worked public holiday.",
     "1. Query: SELECT * FROM employee_dayoff WHERE duration = 8 LIMIT 5\n2. Verify credit entries correlate with public holidays",
     "duration=8 entries represent full working days on public holidays (credit). Most common: 2853 of 5334 entries.",
     "Medium", "Data integrity", "", "day-off-service", "Credit entries outnumber debits (399 surplus)"),

    ("TC-DO-027", "Ledger duration=7 — half-day / short day",
     "Calendar day with duration=7 (pre-holiday shortened day).",
     "1. Create request for half-day calendar entry\n2. Approve\n3. Check ledger duration",
     "Ledger entry created with duration=7. Only 27 of 5334 entries are half-day — rare scenario.",
     "Low", "Boundary", "", "day-off-service", "Half-day: 0.5% of all entries"),

    ("TC-DO-028", "Reject does NOT revert ledger entries (BUG-DO-6)",
     "APPROVED request with 2 ledger entries from prior approval.",
     "1. PUT /v1/employee-dayOff/reject/{id}\n2. Verify status = REJECTED\n3. SELECT * FROM employee_dayoff WHERE employee = {empId}",
     "BUG: Status changed to REJECTED but ledger entries remain intact. Credit/debit records orphaned — not cleaned up on rejection.",
     "Critical", "Bug verification", "BUG-DO-6", "day-off-service", "Known bug: reject = status-only change"),

    ("TC-DO-029", "Approve/reject/re-approve cycle — compounding phantom entries",
     "Clean state request (no prior ledger).",
     "1. Approve (creates 2 entries)\n2. Reject (entries remain)\n3. Approve again (creates 2 MORE entries via upsert)\n4. Count ledger entries",
     "BUG: After approve→reject→approve, upsert may overwrite or compound entries. Each approve cycle writes without checking prior state. Risk of phantom balance drift.",
     "Critical", "Bug verification", "BUG-DO-6", "day-off-service", "Repeated cycles compound data inconsistency"),

    ("TC-DO-030", "Vacation day recalculation triggered by approval",
     "Known vacation balance before approval.",
     "1. Record vacation balance\n2. Approve day-off request\n3. RecalculateVacationDaysHandler fires\n4. Check new vacation balance",
     "Two CalendarDaysChanged events: diff=+1 for lastApprovedDate (credit), diff=-1 for personalDate (debit). Net vacation balance change = 0 (cancel out).",
     "High", "Integration", "", "day-off-service, vacation-service", "Net zero: confirmed in API testing"),

    ("TC-DO-031", "Month norm update triggered by approval",
     "Request with personalDate and lastApprovedDate in different months.",
     "1. Approve request\n2. UpdateMonthNormHandler fires\n3. Check norms for both months",
     "Working hours norm recalculated for both affected months. If same month, single recalculation.",
     "Medium", "Integration", "", "day-off-service, statistics-service", "Cross-month approval triggers 2 norm updates"),

    ("TC-DO-032", "Physical deletion of ledger on DELETED_FROM_CALENDAR (Path B)",
     "APPROVED request with ledger entries. Admin about to delete calendar holiday.",
     "1. Admin deletes calendar day entry\n2. CalendarDeletedApplicationEvent fires\n3. Check employee_dayoff and employee_dayoff_request tables",
     "Ledger entries physically deleted (not soft-delete). Request status = DELETED_FROM_CALENDAR. Vacation days recalculated.",
     "High", "Data integrity", "", "day-off-service, calendar-service", "Path B: physical delete unlike soft-delete elsewhere"),

    ("TC-DO-033", "Physical deletion of ALL year's ledger on office change (Path D)",
     "Employee has multiple day-off requests and ledger entries for 2026.",
     "1. Employee changes office (CompanyStaff sync or admin action)\n2. AutoDeleteHelper.update fires\n3. Check all 2026 ledger entries",
     "ALL ledger entries for the year physically deleted (max 100 requests processed). All requests set to DELETED_FROM_CALENDAR.",
     "High", "Data integrity", "", "day-off-service, companystaff-integration", "Path D: year-wide wipe on office change"),

    ("TC-DO-034", "Transaction isolation — non-atomic ledger + status (BUG-DO-15)",
     "Request being approved.",
     "1. Monitor: changeDayOffDaysAfterApprove (ledger) and changeDayOffStatus (status) calls\n2. Simulate failure between the two calls",
     "BUG: Ledger write and status update are NOT atomic. Ledger can update without status change (or vice versa). No compensating transaction.",
     "High", "Bug verification", "BUG-DO-15", "day-off-service", "Transaction isolation gap — separate non-transactional calls"),
]))

# ══════════════════════════════════════════════════════════════════════════════
# TS-DO-CalConflict — Calendar Conflict Resolution (4 Paths)
# ══════════════════════════════════════════════════════════════════════════════
SUITES.append(("TS-DO-CalConflict", "Calendar Conflict Resolution", [
    ("TC-DO-035", "Path A — Create holiday on date with approved day-off → MOVE",
     "Employee has APPROVED day-off with personal_date = 2026-04-16. No holiday exists on that date.",
     "1. Admin creates holiday 'Test Holiday' on 2026-04-16 in production calendar\n2. CalendarChangedApplicationEvent fires via RabbitMQ\n3. Check employee_dayoff table",
     "New ledger entry created with personal_date = 2026-04-15 (previous working day). Old entry on 2026-04-16 unchanged (orphaned). Request status unchanged (APPROVED).",
     "Critical", "Integration", "", "day-off-service, calendar-service", "Verified live in Session 15"),

    ("TC-DO-036", "Path A — PreviousWorkingDayCalculator logic",
     "Holiday created on Monday (day-off exists for employee).",
     "1. Create holiday on Monday\n2. Check moved ledger entry date",
     "Entry moved to Friday (previous working day). BUG: Calculator only checks Sat/Sun — does NOT query production calendar for additional holidays.",
     "High", "Functional", "BUG (architecture)", "day-off-service", "PreviousWorkingDayCalculator: weekend-only check"),

    ("TC-DO-037", "Path A — Orphaned old ledger entry (BUG-DO-8)",
     "Pre-existing ledger entry on the newly-created holiday date.",
     "1. Verify entry exists on target date\n2. Create holiday on that date\n3. Check both old and new entries",
     "BUG: Old entry remains unchanged on the conflicting date. New moved entry created alongside. Two entries coexist — old is orphaned.",
     "High", "Bug verification", "BUG-DO-8", "day-off-service", "Confirmed live: entry 5351 orphaned while 5491 created"),

    ("TC-DO-038", "Path A — Email notification sent (DAY_MOVED)",
     "Calendar change triggers day-off move.",
     "1. Create holiday on existing day-off date\n2. Check email logs",
     "NOTIFY_VACATION_CALENDAR_UPDATE_0H_DAY_MOVED email sent. Contains moved dates info. BUG: Hardcoded production URL in email body.",
     "Medium", "Functional", "BUG-DO-10", "email-service", "Hardcoded https://ttt.noveogroup.com in all notifications"),

    ("TC-DO-039", "Path A — Half-day (7h) calendar change notification",
     "Calendar day changed to duration=7 (pre-holiday shortened day) where employee has day-off.",
     "1. Admin changes calendar day to 7h\n2. Check notification template",
     "NOTIFY_VACATION_CALENDAR_UPDATE_7H_DAY_MOVED template used (different from 0h). Day-off moved to previous working day.",
     "Medium", "Functional", "", "day-off-service, email-service", "Different template for 7h vs 0h changes"),

    ("TC-DO-040", "Path B — Delete calendar day → DELETED_FROM_CALENDAR",
     "Calendar has holiday with associated employee day-off requests (NEW or APPROVED).",
     "1. Admin deletes calendar day entry\n2. CalendarDeletedApplicationEvent fires\n3. Check requests and ledger",
     "All requests with original_date = deleted date set to DELETED_FROM_CALENDAR (hardcoded SQL). Ledger entries physically deleted. Vacation days recalculated.",
     "Critical", "Integration", "", "day-off-service, calendar-service", "Path B: hardcoded status in SQL, bulk operation"),

    ("TC-DO-041", "Path B — Silent deletion (no email notification, BUG-DO-9)",
     "Calendar day deleted affecting employee day-offs.",
     "1. Delete calendar day\n2. Check email logs for notifications",
     "BUG: No notification email sent on calendar deletion. Asymmetric: Path A (create) sends email, Path B (delete) is silent.",
     "Medium", "Bug verification", "BUG-DO-9", "day-off-service", "Notification asymmetry confirmed live"),

    ("TC-DO-042", "Path B — Bulk operation with multiple employees",
     "Holiday date (e.g. June 12) with 19+ employee day-off requests.",
     "1. Delete the holiday from calendar\n2. Check all affected requests",
     "All matching requests bulk-updated to DELETED_FROM_CALENDAR. Historical: 82 records from 2025-06-12 'День России' removal. All ledger entries for that date deleted.",
     "High", "Integration", "", "day-off-service", "Mass conflict: 82 records in one batch"),

    ("TC-DO-043", "Path B — Entity state bug in updateAll()",
     "Bulk status update via Path B.",
     "1. Trigger calendar deletion\n2. Check Java entity state after SQL update",
     "BUG: updateAll() hardcodes DELETED_FROM_CALENDAR in SQL but Java entities passed to post-update events retain original status (NEW/APPROVED). Events process stale data.",
     "High", "Bug verification", "", "day-off-service", "Architecture: SQL bypasses JPA entity state"),

    ("TC-DO-044", "Path C — Period change → NEW requests REJECTED",
     "Office has NEW day-off requests with last_approved_date on specific date. Approve period about to change.",
     "1. Admin changes approve period for the office\n2. PeriodChangedEventHandler fires (periodType == APPROVE)\n3. Check affected requests",
     "NEW requests with matching last_approved_date set to REJECTED. NOTIFY_DAYOFF_AUTODELETE_TO_EMPLOYEE notification sent. APPROVED requests NOT affected.",
     "High", "Integration", "", "day-off-service, accounting-backend", "Path C: only NEW status targeted"),

    ("TC-DO-045", "Path D — Office change → year-wide DELETED_FROM_CALENDAR",
     "Employee changes office. Has 10 day-off requests for current year (mixed NEW/APPROVED).",
     "1. Employee office changes (via CompanyStaff sync)\n2. AutoDeleteHelper.update fires\n3. Check all current year requests and ledger",
     "All NEW/APPROVED requests (max 100) set to DELETED_FROM_CALENDAR. ALL ledger entries for the year physically deleted. Notification sent per request.",
     "High", "Integration", "", "day-off-service, companystaff-integration", "Path D: most destructive — year-wide wipe"),

    ("TC-DO-046", "Cascading conflict — reassigned date becomes holiday too",
     "Employee's day-off moved from Holiday A date to date X. Later, X becomes Holiday B.",
     "1. Create Holiday A → day-off moved to X\n2. Create Holiday B on date X\n3. Check final state",
     "System handles iteratively: day-off moved again to previous working day before X. Historical: employee 615 went through 3 iterations (Jun 12 → Jul 31 → Dec 31).",
     "High", "Integration", "", "day-off-service", "Cascading: not atomic, handled sequentially"),

    ("TC-DO-047", "Race condition — rapid create-then-delete calendar entries",
     "Day-off exists. Calendar day about to be rapidly created then deleted.",
     "1. Create calendar day (Path A fires → creates moved entry)\n2. Immediately delete calendar day (Path B fires → deletes entries)\n3. Check ledger consistency",
     "Risk: Paths A and B use separate RabbitMQ queues. Rapid sequence could cause Path A's moved entry to not be cleaned by Path B if messages process out of order.",
     "High", "Integration", "", "day-off-service, rabbitmq-messaging", "Race condition: separate MQ queues for create/delete"),

    ("TC-DO-048", "Half-day boundary — duration=7 does NOT trigger conflict",
     "Calendar day with duration=7 (half-day). Employee has day-off on that date.",
     "1. Check CalendarUpdateProcessorImpl.processDay logic\n2. Create calendar entry with duration=7\n3. Verify no conflict resolution triggered",
     "Half-day (7h) does NOT trigger conflict resolution for duration=0 day-offs. processDay only fires for full non-working (0h) changes. Multiple APPROVED requests exist on half-days in production data.",
     "Medium", "Boundary", "", "day-off-service", "Boundary: 7h ≠ 0h in conflict logic"),

    ("TC-DO-049", "Vacation recalculation after DELETED_FROM_CALENDAR (advanceVacation=false)",
     "Russian office employee. Holiday removed from calendar, day-off deleted.",
     "1. Delete calendar holiday\n2. RecalculateVacationDaysHandler fires\n3. Check vacation balance and type",
     "If insufficient accrued days for year → vacation converted to ADMINISTRATIVE type. If sufficient → delayed 10-min check including same-payment-month requests. Email notification (ID_85) on conversion.",
     "High", "Integration", "REQ-day-off §Calendar-Triggered", "day-off-service, vacation-service", "AV=false: conversion to administrative"),

    ("TC-DO-050", "Vacation recalculation after DELETED_FROM_CALENDAR (advanceVacation=true)",
     "Cyprus/Germany office employee. Holiday removed from calendar.",
     "1. Delete calendar holiday\n2. Check vacation balance",
     "Deducts from current year balance (can go negative). No automatic conversion to administrative type.",
     "Medium", "Integration", "REQ-day-off §Calendar-Triggered", "day-off-service, vacation-service", "AV=true: negative balance allowed"),
]))

# ══════════════════════════════════════════════════════════════════════════════
# TS-DO-Validation — Form and API Validation Rules
# ══════════════════════════════════════════════════════════════════════════════
SUITES.append(("TS-DO-Validation", "Day-Off Validation Rules", [
    ("TC-DO-051", "Create with valid publicDate from production calendar",
     "Production calendar has public holiday on target date. No existing request.",
     "1. POST /v1/employee-dayOff with publicDate = valid holiday\n2. Check response",
     "Request created successfully. EmployeeDayOffPublicDateExistsValidator passes: date exists in calendar AND no prior request.",
     "High", "Functional", "", "day-off-service", "3-step validator: no-existing + calendar-check"),

    ("TC-DO-052", "Create with non-existent publicDate — PUBLIC_DATE_NOT_FOUND_IN_CALENDAR",
     "No calendar entry or existing day-off for target date.",
     "1. POST /v1/employee-dayOff with publicDate = 2026-07-15 (random non-holiday)\n2. Check error response",
     "Validation error: PUBLIC_DATE_NOT_FOUND_IN_CALENDAR. Request not created.",
     "High", "Negative", "", "day-off-service", "Validator step 3: neither calendar nor existing day-offs"),

    ("TC-DO-053", "Create with already-requested publicDate — EMPLOYEE_DAY_OFF_PUBLIC_DATE_EXISTS",
     "Existing day-off request for the target publicDate.",
     "1. POST /v1/employee-dayOff with publicDate that has existing request\n2. Check error response",
     "BUG-DO-3: Instead of clean rejection, upsert creates new record shadowing original. Validator returns EMPLOYEE_DAY_OFF_PUBLIC_DATE_EXISTS but upsert bypasses it.",
     "High", "Bug verification", "BUG-DO-3", "day-off-service", "Upsert shadow: new ID replaces old"),

    ("TC-DO-054", "Duplicate personalDate — uniqueness validation",
     "Existing day-off request with personalDate = 2026-06-15.",
     "1. POST new request with personalDate = 2026-06-15\n2. Check response",
     "EmployeeDayOffPersonalDateExistsValidator rejects: personalDate already used by another day-off.",
     "High", "Negative", "", "day-off-service", "PersonalDate uniqueness enforced"),

    ("TC-DO-055", "Past personalDate accepted via API (BUG-DO-4)",
     "Current date is 2026-03-15.",
     "1. PATCH /v1/employee-dayOff/{id} with personalDate = 2026-02-01\n2. Check response",
     "BUG: Accepted without validation error. Past personal date stored. No backend validation for date currency.",
     "High", "Bug verification", "BUG-DO-4", "day-off-service", "Confirmed: past date accepted on 2026-03-13"),

    ("TC-DO-056", "Weekend personalDate accepted via API (BUG-DO-5)",
     "Existing day-off request.",
     "1. POST /v1/employee-dayOff with personalDate = 2026-06-13 (Saturday)\n2. Check response",
     "BUG: Request created with weekend personalDate. No server-side working-day validation. Frontend blocks weekends but API does not.",
     "High", "Bug verification", "BUG-DO-5", "day-off-service", "UI-only validation: API bypass possible"),

    ("TC-DO-057", "UI blocks weekend selection in TransferDaysoffModal",
     "TransferDaysoffModal open for day-off reschedule.",
     "1. Navigate calendar to a weekend\n2. Attempt to click Saturday/Sunday dates",
     "Weekend dates greyed out and unresponsive. renderDay disables Sat/Sun. Weekends only bypass-able via direct API call.",
     "Medium", "Functional", "", "frontend-day-off-module", "Client-side only: no server validation"),

    ("TC-DO-058", "Working weekend exception — selectable in UI datepicker",
     "Production calendar marks a Saturday as working day (make-up weekend).",
     "1. Open TransferDaysoffModal\n2. Navigate to the working Saturday\n3. Attempt to click it",
     "Working weekend date is selectable (re-enabled despite being Saturday). renderDay checks calendar entries for working-weekend exceptions.",
     "Medium", "Boundary", "", "frontend-day-off-module", "Exception: calendar working days override weekend rule"),

    ("TC-DO-059", "maxDate boundary — originalDate + 1 year end",
     "TransferDaysoffModal open for May Day 2026 holiday.",
     "1. Navigate datepicker to December 2027\n2. Try to select December 31, 2027\n3. Try to select January 1, 2028",
     "Dec 31, 2027 selectable (maxDate = end of year after originalDate's year). Jan 1, 2028 not selectable (beyond maxDate).",
     "Medium", "Boundary", "", "frontend-day-off-module", "maxDate = moment(originalDate).add(1,'year').endOf('year')"),

    ("TC-DO-060", "minDate for future holidays — equals originalDate",
     "TransferDaysoffModal open for future holiday (e.g., 2026-12-31).",
     "1. Try to select date before originalDate\n2. Check datepicker constraints",
     "Dates before originalDate greyed/disabled. minDate = originalDate for future holidays.",
     "Medium", "Boundary", "", "frontend-day-off-module", "isMinCurrentDay = false → minDate = originalDate"),

    ("TC-DO-061", "minDate for past holidays — equals yesterday",
     "TransferDaysoffModal open for past holiday (e.g., 2026-02-23, already passed).",
     "1. Check datepicker minimum selectable date\n2. Verify dates before yesterday are disabled",
     "minDate = yesterday. Allows selecting today and forward, but not older dates. Past holiday: isMinCurrentDay = true.",
     "Medium", "Boundary", "", "frontend-day-off-module", "Past holidays: minDate shifts to yesterday"),

    ("TC-DO-062", "Disabled dates — existing day-off personalDates blocked",
     "Employee has day-off personalDate = 2026-05-11 already.",
     "1. Open TransferDaysoffModal for different holiday\n2. Navigate to May 2026\n3. Check if May 11 is selectable",
     "May 11 disabled (greyed). renderDay checks existing day-off personalDates to prevent double-booking.",
     "Medium", "Functional", "", "frontend-day-off-module", "Prevents scheduling two day-offs on same date"),

    ("TC-DO-063", "Short-day conflict dates disabled",
     "Calendar has duration=7 entry on a date where employee already has personalDate.",
     "1. Open TransferDaysoffModal\n2. Check the conflicting short-day date",
     "Date disabled. renderDay checks for calendar days with duration=7 that conflict with existing personalDates.",
     "Low", "Boundary", "", "frontend-day-off-module", "Short-day conflicts: rare but handled"),

    ("TC-DO-064", "No duration/reason DTO validation (service-level only)",
     "Creating day-off request.",
     "1. POST /v1/employee-dayOff with duration = null, reason = null\n2. Check response",
     "No DTO-level validation for duration or reason fields. Validated at service level. Missing values may use defaults or cause NPE.",
     "Medium", "Negative", "", "day-off-service", "No @NotNull on duration/reason in CreateDTO"),

    ("TC-DO-065", "Null personalDate passes through validator",
     "Creating day-off request without personalDate.",
     "1. POST /v1/employee-dayOff with personalDate = null\n2. Check response",
     "EmployeeDayOffPersonalDateExistsValidator: null → valid (pass-through). Request may be created without personal date.",
     "Medium", "Boundary", "", "day-off-service", "Null-safe validator: null = pass"),
]))

# ══════════════════════════════════════════════════════════════════════════════
# TS-DO-Permissions — Access Control and Security
# ══════════════════════════════════════════════════════════════════════════════
SUITES.append(("TS-DO-Permissions", "Day-Off Permissions & Access Control", [
    ("TC-DO-066", "CPO (PROJECT role) self-approval on creation",
     "User has ROLE_PROJECT_MANAGER (CPO equivalent). User is their own manager in CompanyStaff.",
     "1. Create day-off request as CPO user\n2. Check request status and approver",
     "Request auto-approved (status APPROVED). Approver = self. Manager moved to optional approver with ASKED status.",
     "High", "Security", "", "day-off-service", "Self-approval: no external validation gate"),

    ("TC-DO-067", "Approver can APPROVE NEW request",
     "Logged in as primary approver. Request with status NEW.",
     "1. PUT /v1/employee-dayOff/approve/{id}\n2. Check response",
     "Approved successfully. Status → APPROVED. Ledger entries created.",
     "High", "Functional", "", "day-off-service", "Standard approval path"),

    ("TC-DO-068", "Approver can APPROVE REJECTED request",
     "Request with status REJECTED.",
     "1. PUT /v1/employee-dayOff/approve/{id}\n2. Check response",
     "Re-approval succeeds. Status → APPROVED. New ledger entries created (may compound if prior entries exist).",
     "High", "Functional", "", "day-off-service", "Re-approval from REJECTED allowed"),

    ("TC-DO-069", "Approver can REJECT NEW/APPROVED (personalDate >= period start)",
     "APPROVED request, personalDate in current or future period.",
     "1. PUT /v1/employee-dayOff/reject/{id}\n2. Check response",
     "Rejection succeeds. Status → REJECTED. Ledger NOT reverted (BUG-DO-6).",
     "High", "Functional", "BUG-DO-6", "day-off-service", "Rejection allowed when date in open period"),

    ("TC-DO-070", "Approver cannot REJECT if personalDate < period start",
     "APPROVED request with personalDate in closed period.",
     "1. PUT /v1/employee-dayOff/reject/{id}\n2. Check response",
     "Rejection blocked. personalDate is in a closed/past period. Permission denied.",
     "High", "Negative", "", "day-off-service", "Guard: personalDate >= report period start"),

    ("TC-DO-071", "Owner can DELETE (personalDate >= period start OR status != APPROVED)",
     "Own NEW request with personalDate in current period.",
     "1. DELETE /v1/employee-dayOff/{id}\n2. Check response",
     "Deletion succeeds. Status → DELETED. Non-APPROVED status OR future personalDate allows delete.",
     "High", "Functional", "", "day-off-service", "Owner delete guard: OR condition"),

    ("TC-DO-072", "Owner cannot DELETE APPROVED with personalDate < period start",
     "Own APPROVED request with personalDate in closed period.",
     "1. DELETE /v1/employee-dayOff/{id}\n2. Check response",
     "Deletion blocked. Both conditions fail: status IS APPROVED AND personalDate < period start.",
     "High", "Negative", "", "day-off-service", "Both conditions must fail to block"),

    ("TC-DO-073", "Owner can always EDIT (personalDate change)",
     "Own request, any status.",
     "1. PATCH /v1/employee-dayOff/{id} with new personalDate\n2. Check response",
     "Edit succeeds. personalDate updated. Optional approvals reset to ASKED.",
     "Medium", "Functional", "", "day-off-service", "Edit always allowed for owner"),

    ("TC-DO-074", "Non-approver cannot approve/reject",
     "Logged in as random employee (not approver for the request).",
     "1. PUT /v1/employee-dayOff/approve/{id} on someone else's request\n2. Check response",
     "Permission denied / 403. Only primary approver can approve/reject.",
     "High", "Security", "", "day-off-service", "Authorization check on approve/reject"),

    ("TC-DO-075", "Read-only / non-EMPLOYEE — no permissions",
     "User with ROLE_VIEW_ALL (read-only) or non-ROLE_EMPLOYEE.",
     "1. Attempt to create/approve/delete day-off requests\n2. Check response",
     "All mutations blocked. Read-only users cannot create, approve, reject, or delete day-offs.",
     "Medium", "Security", "", "day-off-service", "ROLE_EMPLOYEE required for mutations"),

    ("TC-DO-076", "NPE on findAll without type parameter (BUG-DO-1)",
     "Logged in as any user.",
     "1. GET /v1/employee-dayOff (no type query parameter)\n2. Check response",
     "BUG: NullPointerException at EmployeeDayOffSearchServiceImpl.java:134. ordinal() called on null EmployeeDayOffTypeFilter. Should return error or default.",
     "High", "Bug verification", "BUG-DO-1", "day-off-service", "NPE: type is required but not validated"),

    ("TC-DO-077", "NPE on list endpoint (BUG-DO-2)",
     "Logged in as any user.",
     "1. GET /v1/employee-dayOff/list\n2. Check response",
     "BUG: NullPointerException in Caffeine cache — null key passed to computeIfAbsent at InternalEmployeeService.java:160.",
     "High", "Bug verification", "BUG-DO-2", "day-off-service", "NPE in cache layer — null employee login"),

    ("TC-DO-078", "Approver can always EDIT_APPROVER (change approver)",
     "Current approver for a request.",
     "1. PUT /v1/employee-dayOff/change-approver/{id}/{newLogin}\n2. Check response",
     "Approver change succeeds regardless of request status. Always-available permission for current approver.",
     "Medium", "Functional", "", "day-off-service", "EDIT_APPROVER: no status guard"),
]))

# ══════════════════════════════════════════════════════════════════════════════
# TS-DO-Search — Search Types and Filtering
# ══════════════════════════════════════════════════════════════════════════════
SUITES.append(("TS-DO-Search", "Day-Off Search & Filtering", [
    ("TC-DO-079", "MY search type — 3-source merge (own + calendar + ledger)",
     "Logged in as employee with day-off requests and calendar entries.",
     "1. GET /v1/employee-dayOff?type=MY\n2. Inspect response entries",
     "Returns merged view: (1) own requests with status/dates, (2) calendar holidays without personalDate/status, (3) ledger entries. Date-matching heuristics merge related entries.",
     "High", "Functional", "", "day-off-service", "Complex merge: 3 sources with heuristic matching"),

    ("TC-DO-080", "MY search — calendar entries without status/personalDate",
     "Employee has calendar holidays not yet requested.",
     "1. GET /v1/employee-dayOff?type=MY&year=2026\n2. Filter entries without status",
     "Calendar-sourced entries appear without status or personalDate. These represent available holidays not yet requested for transfer.",
     "Medium", "Functional", "", "day-off-service", "UI filters out duration=8 working-weekend entries"),

    ("TC-DO-081", "ALL search type — admin view",
     "Logged in as admin.",
     "1. GET /v1/employee-dayOff?type=ALL\n2. Check response scope",
     "Returns all day-off requests across all employees. Used by admin interface.",
     "Medium", "Functional", "", "day-off-service", "Admin-only search type"),

    ("TC-DO-082", "APPROVER search — over-inclusion bug (BUG-DO-7)",
     "Logged in as user who is primary approver for 13 employees.",
     "1. GET /v1/employee-dayOff?type=APPROVER\n2. Count results vs actual primary approvals",
     "BUG: Returns 484 results including requests where user is department manager or optional approver, not just primary approver. Only 13-18 should be from actual primary approver role.",
     "High", "Bug verification", "BUG-DO-7", "day-off-service", "Over-includes: 484 vs 18 actual (26x)"),

    ("TC-DO-083", "OPTIONAL_APPROVER search type",
     "User is optional approver for some requests.",
     "1. GET /v1/employee-dayOff?type=OPTIONAL_APPROVER\n2. Check results",
     "Returns requests where user is in employee_dayoff_approval as optional approver. Non-blocking FYI-style approvals.",
     "Medium", "Functional", "", "day-off-service", ""),

    ("TC-DO-084", "MY_DEPARTMENT search type",
     "Logged in as department manager.",
     "1. GET /v1/employee-dayOff?type=MY_DEPARTMENT\n2. Check scope",
     "Returns day-off requests for employees in user's department. Status filter available (NEW, APPROVED, REJECTED).",
     "Medium", "Functional", "", "day-off-service", "Department-scoped view"),

    ("TC-DO-085", "ON_PAID search type — credit entries only",
     "Querying for paid holiday entries.",
     "1. GET /v1/employee-dayOff?type=ON_PAID\n2. Check results",
     "Returns credit-only entries (duration=8, worked holiday). No debit (day-off taken) entries.",
     "Medium", "Functional", "", "day-off-service", "Credit-only view for ledger"),

    ("TC-DO-086", "Year filter interaction with search types",
     "Day-off requests exist across 2025 and 2026.",
     "1. GET /v1/employee-dayOff?type=MY&year=2025\n2. GET /v1/employee-dayOff?type=MY&year=2026\n3. Compare results",
     "Results filtered by year. 2025 shows historical (including DELETED_FROM_CALENDAR). 2026 shows current year. Year selector in UI controls this filter.",
     "Medium", "Functional", "", "day-off-service", "Year filter: UI datepicker year selector"),

    ("TC-DO-087", "Status filter — NEW, APPROVED, REJECTED",
     "Manager view with MY_DEPARTMENT search type.",
     "1. GET /v1/employee-dayOff?type=MY_DEPARTMENT&statuses=NEW\n2. GET with statuses=APPROVED\n3. GET with statuses=REJECTED",
     "Results filtered by status. Status filter available on MY_DEPARTMENT tab. Values: Approved, NEW, Rejected.",
     "Medium", "Functional", "", "frontend-day-off-module", "Status filter on My department tab only"),

    ("TC-DO-088", "Pagination — 20 items per page",
     "More than 20 results for a search type.",
     "1. GET /v1/employee-dayOff?type=MY_DEPARTMENT&size=20&page=0\n2. GET page=1\n3. Check total count",
     "20 items per page. Response includes total count and page metadata. 3 pages observed on My department tab.",
     "Low", "Functional", "", "day-off-service", "Default page size = 20"),

    ("TC-DO-089", "DELEGATED and DELEGATED_TO_ME search types",
     "User has delegated/received day-off approvals.",
     "1. GET /v1/employee-dayOff?type=DELEGATED\n2. GET type=DELEGATED_TO_ME\n3. Check results",
     "DELEGATED: requests user redirected to another approver. DELEGATED_TO_ME: requests redirected to this user.",
     "Low", "Functional", "", "day-off-service", "Redirect-related search types"),

    ("TC-DO-090", "RELATED search type",
     "User has related employees (project members, etc.).",
     "1. GET /v1/employee-dayOff?type=RELATED\n2. Check scope",
     "Returns day-off requests for employees related to user (project members, direct reports). Scope depends on user's role.",
     "Low", "Functional", "", "day-off-service", "Related employees scope"),
]))

# ══════════════════════════════════════════════════════════════════════════════
# TS-DO-ManagerUI — Manager View (5 Sub-tabs)
# ══════════════════════════════════════════════════════════════════════════════
SUITES.append(("TS-DO-ManagerUI", "Day-Off Manager View", [
    ("TC-DO-091", "Manager view — 5 sub-tabs layout",
     "Logged in as manager. Navigate to /vacation/request → Days off rescheduling tab.",
     "1. Navigate to /vacation/request\n2. Click 'Days off rescheduling' tab\n3. Observe sub-tab layout",
     "5 sub-tabs displayed: Approval | Agreement | My department | My projects | Redirected. Badge counts on Approval and Agreement tabs.",
     "High", "Functional", "", "frontend-day-off-module", "RequestTypeWeekendTabsContainer"),

    ("TC-DO-092", "Approval tab — approve action",
     "NEW day-off request visible in Approval tab.",
     "1. Click checkmark (approve) icon on NEW request row\n2. Confirm approval",
     "Status changes to APPROVED. Row moves out of NEW filter. Badge count decrements. Ledger entries created on backend.",
     "High", "Functional", "", "frontend-day-off-module", "Inline action button"),

    ("TC-DO-093", "Approval tab — reject action",
     "NEW day-off request visible in Approval tab.",
     "1. Click X (reject) icon on NEW request row\n2. Confirm rejection",
     "Status changes to REJECTED. Row updates. Employee notified.",
     "High", "Functional", "", "frontend-day-off-module", ""),

    ("TC-DO-094", "Approval tab — redirect to different approver",
     "Request in Approval tab. User wants to delegate approval.",
     "1. Click redirect icon on request row\n2. WeekendRedirectFormContainer opens\n3. Select new approver from dropdown\n4. Confirm",
     "Old approver → optional. New approver assigned. Request reappears in new approver's Approval tab.",
     "Medium", "Functional", "", "frontend-day-off-module", "PUT /change-approver/{id}/{login}"),

    ("TC-DO-095", "Request details modal — full information display",
     "Any request in manager view.",
     "1. Click info/details icon on request row\n2. Observe WeekendDetailsModal content",
     "Shows: Employee (CS link), Manager (CS link), Reason, Initial date (YYYY-MM-DD), Requested date, Status. Approve/Reject/Redirect buttons if approver.",
     "Medium", "Functional", "", "frontend-day-off-module", "WeekendDetailsModal"),

    ("TC-DO-096", "Details modal — optional approvers sub-table",
     "Request has optional approvers with mixed statuses.",
     "1. Open details modal\n2. Scroll to 'Agreed by' section",
     "Sub-table shows optional approvers with individual statuses (Requested, Agreed, Rejected). 'Agreed by' column in main table is empty — only visible in modal.",
     "Medium", "Functional", "", "frontend-day-off-module", "ApproveBar with vote counts"),

    ("TC-DO-097", "Agreement tab — optional approve/reject",
     "User is optional approver for a request. Request appears in Agreement tab.",
     "1. Click Agreement tab\n2. Approve or reject the optional approval\n3. Check main request status",
     "Optional approval status updated (ASKED → APPROVED/REJECTED). Main request status unchanged. Non-blocking.",
     "Medium", "Functional", "", "frontend-day-off-module", "FYI-style: does not affect main status"),

    ("TC-DO-098", "My department tab — status filter",
     "Department manager with day-off requests in department.",
     "1. Click My department tab\n2. Use Status filter dropdown\n3. Select 'NEW'\n4. Check filtered results",
     "Table shows only NEW requests. Filter values: Approved, NEW, Rejected. Status filter unique to MY_DEPARTMENT tab.",
     "Medium", "Functional", "", "frontend-day-off-module", "Status filter on My department only"),

    ("TC-DO-099", "Badge counts on Approval and Agreement tabs",
     "Manager has pending approvals and agreements.",
     "1. Navigate to Days off rescheduling\n2. Observe badge numbers on Approval and Agreement tabs",
     "Approval tab badge = approverWeekendCount (NEW requests to approve). Agreement tab badge = optionalApproverWeekendCount. Counts update on action.",
     "Medium", "Functional", "", "frontend-day-off-module", "Redux: approverWeekendCount/optionalApproverWeekendCount"),

    ("TC-DO-100", "Table columns — 7-column layout",
     "Manager view with data.",
     "1. Check table column headers in any sub-tab",
     "Columns: Employee | Initial date (default sort desc) | Requested date | Manager | Approved by | Agreed by | Status | Actions. Agreed by column empty in table (data in modal only).",
     "Low", "Functional", "", "frontend-day-off-module", "7 columns + Actions"),

    ("TC-DO-101", "Overdue warning banner — broadcast to all users (BUG-DO-11)",
     "Overdue NEW requests exist (personalDate or lastApprovedDate < today).",
     "1. Log in as regular employee (no approver role)\n2. Check for banner on vacation/request page\n3. Click banner link",
     "BUG: Warning 'You have overdue day off rescheduling requests' shown to ALL users including non-approvers. Link goes to APPROVER tab → shows 0 items or 403 for non-managers.",
     "High", "Bug verification", "BUG-DO-11", "day-off-service, frontend-day-off-module", "Confirmed: regular employee sees irrelevant warning"),

    ("TC-DO-102", "Warning banner — correct behavior for actual approvers",
     "Overdue NEW requests exist. Logged in as actual approver for those requests.",
     "1. Check for banner\n2. Click link to APPROVER tab\n3. Verify overdue requests visible",
     "Banner correctly shows for approvers. APPROVER tab lists the overdue requests. Approver can take action (approve/reject).",
     "Medium", "Functional", "", "frontend-day-off-module", "Expected behavior for actual approvers"),
]))

# ══════════════════════════════════════════════════════════════════════════════
# TS-DO-EmployeeUI — Employee View (Days Off Tab)
# ══════════════════════════════════════════════════════════════════════════════
SUITES.append(("TS-DO-EmployeeUI", "Day-Off Employee View", [
    ("TC-DO-103", "Employee view — Days off tab layout",
     "Logged in as employee. Navigate to /vacation/my.",
     "1. Click 'Days off' tab\n2. Observe layout",
     "Title: 'My vacations and days off'. Year selector datepicker. 'Weekend regulation' info link → Confluence page. Table with 6 columns: Date of event | Duration | Reason | Approved by | Status | Actions.",
     "High", "Functional", "", "frontend-day-off-module", "WeekendTab component"),

    ("TC-DO-104", "Year selector — filter entries by year",
     "Employee has day-off data across 2025 and 2026.",
     "1. Select 2025 in year datepicker\n2. Select 2026\n3. Compare row counts",
     "Table filters by selected year. 2026 Russia: 17 holidays. 2026 Cyprus: 6 holidays. Historical years show past data.",
     "Medium", "Functional", "", "frontend-day-off-module", "Year-filtered table"),

    ("TC-DO-105", "Edit button visibility logic",
     "Employee view with mix of future and past holidays.",
     "1. Check which rows show edit (pencil) icon\n2. Compare with row data",
     "Edit button only on rows where: NOT a weekend AND duration=0 AND lastApprovedDate >= today. Past holidays and short-day (7h) rows have no edit. 5 of 17 rows had edit in session 4.",
     "High", "Functional", "", "frontend-day-off-module", "Conditional: future + duration=0 + not weekend"),

    ("TC-DO-106", "Cancel button (X) on NEW status rows",
     "Employee has NEW (pending) day-off request.",
     "1. Observe row with status NEW\n2. Check for red X cancel button\n3. Click X to cancel",
     "Red X button visible on NEW status rows. Click triggers DELETE /v1/employee-dayOff/{id}. Status → DELETED. Row removed from view.",
     "High", "Functional", "", "frontend-day-off-module", "Cancel postponement flow"),

    ("TC-DO-107", "TransferDaysoffModal — calendar picker with constraints",
     "Click edit on future holiday row.",
     "1. TransferDaysoffModal opens with title 'Reschedule an event'\n2. Read-only 'Day off date' shows original date\n3. Calendar month-view datepicker displayed\n4. Navigate months with arrows",
     "Modal shows read-only original date + interactive calendar. Weekends disabled (grey). Existing day-off dates disabled. Short-day conflicts disabled. Working weekends enabled (exception).",
     "High", "Functional", "", "frontend-day-off-module", "TransferDaysoffModal with renderDay logic"),

    ("TC-DO-108", "TransferDaysoffModal — OK button disabled until valid date selected",
     "TransferDaysoffModal open.",
     "1. Observe OK button state before selection\n2. Click valid date\n3. Observe OK button state after selection",
     "OK button starts disabled. Enables immediately after valid date click. Cancel always enabled.",
     "Medium", "Functional", "", "frontend-day-off-module", "Submit-gate: disabled until value selected"),

    ("TC-DO-109", "Date format — DD.MM.YYYY (weekday) display",
     "Employee view with day-off entries.",
     "1. Check 'Date of event' column format",
     "Format: 'DD.MM.YYYY (abbreviated weekday)', e.g. '01.05.2026 (fr)'. NEW status shows arrow format: '01.05.2026 → 11.05.2026'.",
     "Low", "Functional", "", "frontend-day-off-module", "useWeekendTableHeaders date display logic"),

    ("TC-DO-110", "NEW status shows originalDate → personalDate arrow",
     "NEW request with both originalDate and personalDate.",
     "1. Check Date of event column for NEW row",
     "Arrow format: 'lastApprovedDate → personalDate'. Other statuses show only lastApprovedDate.",
     "Medium", "Functional", "", "frontend-day-off-module", "Status-dependent display logic"),

    ("TC-DO-111", "Filtering hides working-weekend entries (duration=8)",
     "Calendar has working-weekend entries (Sat marked as working day).",
     "1. Check employee Days off tab\n2. Look for duration=8 entries without compensatory day-off",
     "WeekendTab filtering hides rows where checking date is normal working day with duration=8. Only day-offs and holidays visible, not standalone working weekends.",
     "Medium", "Functional", "", "frontend-day-off-module", "May confuse users expecting full calendar view"),

    ("TC-DO-112", "Navigation bug — Days off tab may redirect to /sick-leave/my",
     "On /vacation/my page.",
     "1. Click 'Days off' tab button\n2. Observe navigation target",
     "BUG: Sometimes navigates to /sick-leave/my instead of /vacation/my/daysoff. Direct URL navigation works correctly. Intermittent behavior.",
     "Medium", "Bug verification", "", "frontend-day-off-module", "Session 4 finding: intermittent redirect"),

    ("TC-DO-113", "Localization — reasons display in Russian in EN mode (BUG-DO-12)",
     "UI set to English language.",
     "1. Switch to EN locale via nav bar language dropdown\n2. Navigate to Days off tab\n3. Check Reason column text",
     "BUG: Reasons display in Russian even in EN mode (e.g., 'Новый год' instead of 'New Year'). Holiday names from production calendar not localized.",
     "Medium", "Bug verification", "BUG-DO-12", "frontend-day-off-module, calendar-service", "Localization gap: reason field from DB, not i18n"),

    ("TC-DO-114", "Duration column display semantics",
     "Employee view with various duration values.",
     "1. Check Duration column for different rows",
     "Duration = '0' for full day-off, '7' for pre-holiday shortened day, '8' for full working day. UI shows raw hours — no human-friendly label.",
     "Low", "Functional", "", "frontend-day-off-module", "Raw duration: 0/7/8 hours displayed"),

    ("TC-DO-115", "Hardcoded date '2024-03-10' in WeekendTableActions (BUG-DO-13)",
     "Employee view code.",
     "1. Review WeekendTableActions component\n2. Check isOnlyOneAction logic",
     "BUG: Hardcoded date '2024-03-10' in WeekendTableActions isOnlyOneAction — test/stub value never replaced. May affect action button visibility logic.",
     "Low", "Bug verification", "BUG-DO-13", "frontend-day-off-module", "Dead test value in production code"),
]))


# ══════════════════════════════════════════════════════════════════════════════
# Risk Assessment Data
# ══════════════════════════════════════════════════════════════════════════════
RISKS = [
    ("Ledger not reverted on reject (BUG-DO-6)", "Orphaned ledger entries accumulate on approve/reject cycles, causing phantom vacation balance drift", "High", "Critical", "Critical",
     "Test approve→reject→approve sequences. Verify ledger entry count after each operation. Monitor vacation balance for drift."),
    ("Calendar conflict orphaned entries (BUG-DO-8)", "Path A creates moved ledger entry but does not clean up old entry — duplicate ledger records", "High", "High", "Critical",
     "Test calendar create→check ledger for orphans. Verify old+new entries don't cause double day-off credit/debit."),
    ("NPE on API endpoints (BUG-DO-1, BUG-DO-2)", "findAll and list endpoints crash with NullPointerException on missing parameters", "High", "High", "High",
     "Test all GET endpoints without optional parameters. Verify graceful error handling or default values."),
    ("Past/weekend date validation gaps (BUG-DO-4, DO-5)", "API accepts past dates and weekend personalDate — no server-side validation", "High", "Medium", "High",
     "Test API with boundary dates: past, weekend, holidays. Compare UI vs API validation coverage."),
    ("Upsert shadow behavior (BUG-DO-3)", "Duplicate publicDate creates shadow record via upsert, confusing IDs and potentially losing data", "Medium", "High", "High",
     "Test duplicate creation attempts. Verify old record not lost. Check ID consistency."),
    ("Transaction isolation gap (BUG-DO-15)", "Ledger write and status update are non-atomic — partial failure possible", "Medium", "Critical", "High",
     "Test concurrent operations. Verify ledger and status consistency after failures."),
    ("Calendar delete silent (BUG-DO-9)", "Path B does not send notification email — asymmetric with Path A", "Medium", "Medium", "Medium",
     "Compare notification behavior across all 4 calendar conflict paths."),
    ("Overdue warning broadcast (BUG-DO-11)", "All users see overdue day-off warning, not just approvers", "High", "Medium", "Medium",
     "Test warning visibility across roles: employee, PM, DM, admin. Verify link leads to actionable page."),
    ("APPROVER search over-inclusion (BUG-DO-7)", "Returns 26x more results than actual primary approvals", "Medium", "Medium", "Medium",
     "Test APPROVER search type. Count results vs actual primary approver assignments."),
    ("Hardcoded production URL (BUG-DO-10)", "All notification emails contain https://ttt.noveogroup.com regardless of environment", "High", "Low", "Medium",
     "Check notification emails on test environments for correct URL."),
    ("Navigation bug — Days off tab misdirect", "Clicking Days off tab sometimes navigates to /sick-leave/my", "Medium", "Medium", "Medium",
     "Test tab navigation 10+ times. Check for intermittent redirects."),
    ("Localization (BUG-DO-12)", "Day-off reasons display in Russian even in English UI mode", "High", "Low", "Low",
     "Switch to EN locale. Verify reason column text language."),
    ("Hardcoded date in UI (BUG-DO-13)", "WeekendTableActions has hardcoded '2024-03-10' test value", "Low", "Low", "Low",
     "Review code impact. Test action button visibility around that date."),
]

# ══════════════════════════════════════════════════════════════════════════════
# Feature Matrix Data
# ══════════════════════════════════════════════════════════════════════════════
# (feature, functional, negative, boundary, data_integrity, bug_verification, security, integration)
FEATURE_MATRIX = [
    ("Request Lifecycle (CRUD)", 12, 2, 1, 0, 1, 0, 4),
    ("Ledger Mechanics", 2, 0, 3, 6, 3, 0, 0),
    ("Calendar Conflicts (4 Paths)", 2, 0, 1, 1, 3, 0, 9),
    ("Validation Rules", 3, 3, 5, 0, 3, 0, 1),
    ("Permissions & Access", 5, 2, 0, 0, 2, 3, 1),
    ("Search & Filtering", 10, 0, 0, 0, 1, 0, 1),
    ("Manager View (5 Tabs)", 9, 0, 0, 0, 1, 0, 2),
    ("Employee View", 8, 0, 0, 0, 3, 0, 2),
]


# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK GENERATION
# ══════════════════════════════════════════════════════════════════════════════
def build_workbook():
    wb = openpyxl.Workbook()

    # ── Plan Overview ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Plan Overview"
    ws.sheet_properties.tabColor = TAB_GREEN

    plan_rows = [
        ("Test Plan: Day-Off (Weekend Transfer) Module", ""),
        ("", ""),
        ("Document", "day-off.xlsx"),
        ("Module", "Day-Off / Weekend Transfer (Compensatory Day-Off Lifecycle)"),
        ("Version", "1.0"),
        ("Date", date.today().isoformat()),
        ("Branch", "release/2.1"),
        ("Environment", "timemachine (primary dev), qa-1 (secondary dev), stage (prod baseline)"),
        ("", ""),
        ("SCOPE & OBJECTIVES", ""),
        ("", ""),
        ("Scope", "Full lifecycle of compensatory day-off requests: create, approve, reject, delete, edit (reschedule). "
                  "Two-table architecture (employee_dayoff_request + employee_dayoff ledger). "
                  "4 calendar conflict resolution paths (calendar create/delete, period change, office change). "
                  "Credit/debit ledger mechanics, vacation day recalculation, month norm updates. "
                  "8 search types including 3-source MY merge. "
                  "Employee view (Days off tab), Manager view (5 sub-tabs), TransferDaysoffModal. "
                  "15 known bugs across API, UI, data integrity, and localization layers."),
        ("Out of scope", "Availability chart integration, notification email content/formatting (covered in email module), "
                         "CompanyStaff sync internals, production calendar CRUD (covered in admin module)."),
        ("", ""),
        ("Objectives", "1. Verify complete request lifecycle (create → approve → use → ledger)\n"
                       "2. Validate credit/debit ledger mechanics and vacation balance impact\n"
                       "3. Test all 4 calendar conflict resolution paths (A: move, B: delete, C: system reject, D: office change)\n"
                       "4. Verify permission model (approver, owner, CPO self-approve, read-only)\n"
                       "5. Confirm 15 known bugs (7 from API testing, 8 from UI/code analysis)\n"
                       "6. Test 8 search types with filtering and pagination\n"
                       "7. Validate UI flows: employee view, manager 5-tab view, TransferDaysoffModal constraints"),
        ("", ""),
        ("APPROACH", ""),
        ("", ""),
        ("Test types", "Functional (51 cases), Negative (7), Boundary (10), Data integrity (7), "
                       "Bug verification (17), Security (3), Integration (20)"),
        ("Total cases", f"{sum(len(s[2]) for s in SUITES)} test cases across {len(SUITES)} test suites"),
        ("", ""),
        ("TEST DATA STRATEGY", ""),
        ("", ""),
        ("Test data", "Timemachine env: 3,241 day-off requests (2,902 APPROVED, 17 NEW, 14 REJECTED). "
                      "5,334 ledger entries (2,853 credit, 2,454 debit, 27 half-day). "
                      "Key test users: pvaynmaster (7 roles, CPO self-approve), perekrest (admin+accountant), "
                      "asmirnov (regular employee, Cyprus office). "
                      "SQL queries for test data:\n"
                      "  SELECT * FROM employee_dayoff_request WHERE status = 'NEW' -- 17 pending requests\n"
                      "  SELECT * FROM employee_dayoff WHERE duration = 0 LIMIT 10 -- day-off debits\n"
                      "  SELECT e.login, COUNT(*) FROM employee_dayoff_request r JOIN employee e ON r.approver = e.id "
                      "WHERE r.status = 'NEW' GROUP BY e.login -- overdue by approver\n"
                      "  SELECT * FROM employee_dayoff_request WHERE status = 'DELETED_FROM_CALENDAR' -- 82 conflict records"),
        ("", ""),
        ("ENVIRONMENT REQUIREMENTS", ""),
        ("", ""),
        ("Environment", "Timemachine (primary): test clock at current date, all roles available, 3,241 requests.\n"
                        "QA-1 (secondary): same build (2.1.26-SNAPSHOT), Cyprus office for cross-calendar testing.\n"
                        "Stage (prod baseline): compare behavior with production-like data."),
        ("", ""),
        ("QASE GAP ANALYSIS", ""),
        ("", ""),
        ("Existing Qase", "19 test cases in 'Calendar day-off transfers' sub-suite (display-oriented, not lifecycle). "
                          "0 lifecycle test cases. 0 ledger mechanics. 0 calendar conflict resolution. "
                          "0 search type testing. 0 permission model testing."),
        ("Gap", "Full lifecycle coverage needed: CRUD operations, ledger mechanics, all 4 conflict paths, "
                "8 search types, permission model, 15 known bugs. This workbook fills 100% of lifecycle gaps."),
        ("", ""),
        ("TEST SUITES", ""),
    ]

    for r, (k, v) in enumerate(plan_rows, 1):
        ws.cell(row=r, column=1, value=k).font = SUBTITLE_FONT if k.isupper() else (TITLE_FONT if r == 1 else BODY_FONT)
        ws.cell(row=r, column=2, value=v).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP

    # Suite hyperlinks
    link_start = len(plan_rows) + 1
    for i, (tab, title, cases) in enumerate(SUITES):
        r = link_start + i
        cell = ws.cell(row=r, column=1)
        cell.value = f"{tab} — {title}"
        cell.hyperlink = f"#'{tab}'!A1"
        cell.font = LINK_FONT
        ws.cell(row=r, column=2, value=f"{len(cases)} cases").font = BODY_FONT

    set_col_widths(ws, [40, 110])

    # ── Feature Matrix ───────────────────────────────────────────────────────
    ws_fm = wb.create_sheet("Feature Matrix")
    ws_fm.sheet_properties.tabColor = TAB_GREEN

    fm_headers = ["Feature Area", "Functional", "Negative", "Boundary", "Data Integrity",
                  "Bug Verification", "Security", "Integration", "Total"]
    for c, h in enumerate(fm_headers, 1):
        ws_fm.cell(row=1, column=c, value=h)
    style_header_row(ws_fm, 1, len(fm_headers))

    for idx, (feat, func, neg, bnd, di, bug, sec, intg) in enumerate(FEATURE_MATRIX):
        r = idx + 2
        total = func + neg + bnd + di + bug + sec + intg
        vals = [feat, func, neg, bnd, di, bug, sec, intg, total]
        for c, v in enumerate(vals, 1):
            cell = ws_fm.cell(row=r, column=c, value=v)
            if c == 1:
                cell.hyperlink = f"#'{SUITES[idx][0]}'!A1"
                cell.font = LINK_FONT
        style_data_row(ws_fm, r, len(fm_headers), idx)

    # Totals row
    total_r = len(FEATURE_MATRIX) + 2
    ws_fm.cell(row=total_r, column=1, value="TOTAL").font = Font(name=ARIAL, bold=True, size=11)
    for c in range(2, len(fm_headers) + 1):
        total_val = sum(ws_fm.cell(row=r, column=c).value or 0 for r in range(2, total_r))
        ws_fm.cell(row=total_r, column=c, value=total_val).font = Font(name=ARIAL, bold=True, size=11)
    for c in range(1, len(fm_headers) + 1):
        ws_fm.cell(row=total_r, column=c).border = THIN_BORDER
        ws_fm.cell(row=total_r, column=c).fill = PatternFill("solid", fgColor="E2EFDA")

    set_col_widths(ws_fm, [35, 12, 12, 12, 15, 16, 12, 14, 10])
    ws_fm.auto_filter.ref = f"A1:{get_column_letter(len(fm_headers))}{total_r}"

    # ── Risk Assessment ──────────────────────────────────────────────────────
    ws_ra = wb.create_sheet("Risk Assessment")
    ws_ra.sheet_properties.tabColor = TAB_GREEN

    ra_headers = ["Risk", "Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    for c, h in enumerate(ra_headers, 1):
        ws_ra.cell(row=1, column=c, value=h)
    style_header_row(ws_ra, 1, len(ra_headers))

    severity_fills = {"Critical": RISK_CRIT, "High": RISK_HIGH, "Medium": RISK_MED, "Low": RISK_LOW}

    for idx, (risk, desc, likelihood, impact, severity, mitigation) in enumerate(RISKS):
        r = idx + 2
        vals = [risk, desc, likelihood, impact, severity, mitigation]
        for c, v in enumerate(vals, 1):
            ws_ra.cell(row=r, column=c, value=v)
        style_data_row(ws_ra, r, len(ra_headers), idx)
        sev_fill = severity_fills.get(severity)
        if sev_fill:
            ws_ra.cell(row=r, column=5).fill = sev_fill

    set_col_widths(ws_ra, [40, 60, 14, 14, 14, 60])
    ws_ra.auto_filter.ref = f"A1:{get_column_letter(len(ra_headers))}{len(RISKS) + 1}"

    # ── Test Suite Tabs ──────────────────────────────────────────────────────
    tc_headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
                  "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    tc_widths = [14, 40, 35, 50, 50, 10, 16, 22, 30, 30]

    for tab_name, suite_title, cases in SUITES:
        ws_ts = wb.create_sheet(tab_name)
        ws_ts.sheet_properties.tabColor = TAB_BLUE

        # Back-link
        back = ws_ts.cell(row=1, column=1, value="\u2190 Back to Plan")
        back.hyperlink = "#'Plan Overview'!A1"
        back.font = BACK_LINK_FONT

        ws_ts.cell(row=2, column=1, value=f"{tab_name}: {suite_title}").font = TITLE_FONT

        header_row = 4
        for c, h in enumerate(tc_headers, 1):
            ws_ts.cell(row=header_row, column=c, value=h)
        style_header_row(ws_ts, header_row, len(tc_headers))

        for idx, case in enumerate(cases):
            r = header_row + 1 + idx
            for c, v in enumerate(case, 1):
                ws_ts.cell(row=r, column=c, value=v)
            style_data_row(ws_ts, r, len(tc_headers), idx)

        set_col_widths(ws_ts, tc_widths)
        last_data = header_row + len(cases)
        ws_ts.auto_filter.ref = f"A{header_row}:{get_column_letter(len(tc_headers))}{last_data}"

    wb.save(OUTPUT)
    total_cases = sum(len(s[2]) for s in SUITES)
    print(f"Generated {OUTPUT}")
    print(f"  Tabs: Plan Overview + Feature Matrix + Risk Assessment + {len(SUITES)} TS-* suites")
    print(f"  Total test cases: {total_cases}")
    print(f"  Risks: {len(RISKS)}")
    for tab, title, cases in SUITES:
        priorities = {}
        for c in cases:
            p = c[5]
            priorities[p] = priorities.get(p, 0) + 1
        pstr = "/".join(f"{priorities.get(k, 0)}" for k in ["Critical", "High", "Medium", "Low"])
        print(f"    {tab}: {len(cases)} cases (C/H/M/L: {pstr})")


if __name__ == "__main__":
    build_workbook()
