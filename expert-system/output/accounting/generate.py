#!/usr/bin/env python3
"""Generate unified accounting.xlsx workbook — supplements to existing Qase 127 cases."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

OUTPUT = "/home/v/Dev/ttt-expert-v1/expert-system/output/accounting/accounting.xlsx"

# -- Styles ---------------------------------------------------------------
ARIAL = "Arial"
HEADER_FONT = Font(name=ARIAL, bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
SUBTITLE_FONT = Font(name=ARIAL, bold=True, size=12)
BODY_FONT = Font(name=ARIAL, size=10)
LINK_FONT = Font(name=ARIAL, size=10, color="0563C1", underline="single")
BACK_LINK_FONT = Font(name=ARIAL, size=9, color="0563C1", underline="single")
ROW_EVEN = PatternFill("solid", fgColor="D6E4F0")
ROW_ODD = PatternFill("solid", fgColor="FFFFFF")
RISK_CRIT = PatternFill("solid", fgColor="FF6B6B")
RISK_HIGH = PatternFill("solid", fgColor="FFA07A")
RISK_MED = PatternFill("solid", fgColor="FFD700")
RISK_LOW = PatternFill("solid", fgColor="90EE90")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TAB_GREEN = "00B050"
TAB_BLUE = "4472C4"


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def style_data_row(ws, row, max_col, idx):
    fill = ROW_EVEN if idx % 2 == 0 else ROW_ODD
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# -- Test Suite Definitions ------------------------------------------------
# (id, title, preconditions, steps, expected, priority, type, req_ref, component, notes)

SUITES = []

# ==========================================================================
# TS-ACC-PeriodEdge -- Period Management Edge Cases & Validation Bugs
# ==========================================================================
SUITES.append(("TS-ACC-PeriodEdge", "Period Management Edge Cases & Validation Bugs", [
    ("TC-ACC-001", "Report period: advance multiple months forward in one operation",
     "ACCOUNTANT or ADMIN user. Salary office with REPORT period = 2026-03-01.",
     "1. PATCH /v1/offices/{officeId}/periods/report with start = 2026-07-01\n2. Verify response\n3. Check office_period table",
     "Period advanced successfully to July. No upper bound or jump-size restriction enforced on report period. DB updated, cache evicted.",
     "High", "Boundary", "PERIOD-REPORT-01", "OfficePeriodServiceImpl.patchReportPeriod",
     "Unlike approve period, report period has no max jump limit"),

    ("TC-ACC-002", "Report period: attempt to set before approve period start",
     "ACCOUNTANT. Office: REPORT=2026-03-01, APPROVE=2026-02-01.",
     "1. PATCH /v1/offices/{id}/periods/report with start = 2026-01-01 (before approve)\n2. Verify error response",
     "400 error. Report period cannot precede approve period (strict < check). Error code indicates constraint violation.",
     "High", "Negative", "PERIOD-REPORT-02", "OfficePeriodServiceImpl",
     "Invariant: APPROVE ≤ REPORT always"),

    ("TC-ACC-003", "Report period: non-first-day-of-month rejected",
     "ACCOUNTANT. Salary office.",
     "1. PATCH /v1/offices/{id}/periods/report with start = 2026-04-15\n2. Verify error response",
     "400 error. getDayOfMonth() != 1 check at line 91 rejects non-first dates. Error message indicates first day required.",
     "Medium", "Validation", "PERIOD-REPORT-03", "OfficePeriodServiceImpl:91", ""),

    ("TC-ACC-004", "BUG: Approve period accepts non-first-day-of-month",
     "ACCOUNTANT. Salary office with APPROVE=2026-02-01.",
     "1. PATCH /v1/offices/{id}/periods/approve with start = 2026-03-15\n2. Verify response (expect error but currently 200)",
     "BUG (BUG-PERIOD-1): 200 OK returned, period set to March 15. Missing getDayOfMonth() != 1 check at patchApprovePeriod():104. Report period has this check but approve does not.",
     "Critical", "Bug", "PERIOD-APPROVE-01, BUG-PERIOD-1", "OfficePeriodServiceImpl:104",
     "HIGH severity known bug. Compare with patchReportPeriod() line 91 which has the validation."),

    ("TC-ACC-005", "BUG: NPE on null start in PATCH body",
     "ACCOUNTANT. Any salary office.",
     "1. PATCH /v1/offices/{id}/periods/report with body {} or {\"start\": null}\n2. Verify response",
     "BUG (BUG-PERIOD-2): 500 NullPointerException at start.getDayOfMonth(). DTO has @NotNull but @RequestBody lacks @Valid annotation. Stack trace returned in response body.",
     "Critical", "Bug", "PERIOD-PATCH-01, BUG-PERIOD-2", "OfficePeriodController",
     "HIGH severity. Both report and approve endpoints affected."),

    ("TC-ACC-006", "Stack trace leakage on invalid date format in period PATCH",
     "ACCOUNTANT. Salary office.",
     "1. PATCH /v1/offices/{id}/periods/report with start = 'not-a-date'\n2. Inspect error response body",
     "BUG (BUG-PERIOD-3): Full Java stack trace (98+ frames) in response body including class names, method names, package structure. Information disclosure risk.",
     "High", "Security", "PERIOD-PATCH-02, BUG-PERIOD-3", "OfficePeriodController",
     "MEDIUM severity. Multiple endpoints affected by stack trace leakage."),

    ("TC-ACC-007", "BUG: Permission inconsistency between report and approve min/max",
     "API token with valid permissions.",
     "1. GET /v1/offices/periods/report/min with API_SECRET_TOKEN → expect 403\n2. GET /v1/offices/periods/approve/min with API_SECRET_TOKEN → expect 200\n3. Compare authorization behavior",
     "BUG (BUG-PERIOD-4): Report min/max requires JWT only (rejects API token with 403). Approve min/max accepts both JWT and API token. Inconsistent @PreAuthorize configuration.",
     "High", "Bug", "PERIOD-AUTH-01, BUG-PERIOD-4", "OfficePeriodController",
     "MEDIUM severity. Auth configuration inconsistency between similar endpoints."),

    ("TC-ACC-008", "Invalid office ID returns 200 with default period data",
     "ACCOUNTANT or ADMIN user.",
     "1. GET /v1/offices/99999/periods/report (nonexistent office)\n2. GET /v1/offices/0/periods/report\n3. GET /v1/offices/-1/periods/report\n4. Verify all return 200 with period data",
     "BUG: All return 200 with valid-looking period response instead of 404. Silent fallback masks invalid input. Contrast with employee login which properly returns 400.",
     "High", "Bug", "PERIOD-INPUT-01", "OfficePeriodController",
     "MEDIUM severity. Should return 404 for nonexistent office."),

    ("TC-ACC-009", "Approve period: max 1-month jump forward enforced",
     "ACCOUNTANT. Office APPROVE=2026-02-01.",
     "1. PATCH /v1/offices/{id}/periods/approve with start = 2026-04-01 (2 months ahead)\n2. Verify error response\n3. PATCH with start = 2026-03-01 (1 month ahead)\n4. Verify success",
     "2-month jump rejected with CODE_ERROR_APPROVE_CHANGE_MORE_THAN_ONE_MONTH. 1-month jump accepted. Maximum 1-month jump in either direction enforced.",
     "High", "Boundary", "PERIOD-APPROVE-02", "OfficePeriodServiceImpl",
     "Critical business rule. Accountants must advance month by month."),

    ("TC-ACC-010", "Approve period: 2-month backward limit enforced",
     "ACCOUNTANT. Office APPROVE=2026-02-01. Current date ~March 2026.",
     "1. PATCH /v1/offices/{id}/periods/approve with start = 2025-12-01 (3 months ago)\n2. Verify error response\n3. PATCH with start = 2026-01-01 (1 month back)\n4. Verify success",
     "3-month revert rejected. Cannot go back more than 2 months from today (today.minusMonths(2).withDayOfMonth(2)). 1-month revert accepted.",
     "High", "Boundary", "PERIOD-APPROVE-03", "OfficePeriodServiceImpl",
     "Lower bound prevents excessive historical reopening."),

    ("TC-ACC-011", "Non-salary office: GET returns default, PATCH returns 404",
     "ADMIN user. Non-salary office (e.g. office.salary=FALSE, id=9).",
     "1. GET /v1/offices/9/periods/report → verify returns computed default\n2. PATCH /v1/offices/9/periods/report with valid start → verify 404\n3. Confirm office.salary=FALSE in DB",
     "GET returns computed default period (today - 1 month). PATCH returns 404 — non-salary offices cannot have periods modified. Filtered by office.salary = TRUE check.",
     "Medium", "Functional", "PERIOD-NONSALARY-01", "OfficePeriodServiceImpl",
     "Non-salary offices (Академгородок, Париж, Дюссельдорф) are legacy/inactive."),

    ("TC-ACC-012", "Period caching: eviction on PATCH, consistency on concurrent GETs",
     "ACCOUNTANT. Salary office.",
     "1. GET /v1/offices/{id}/periods/report → note response\n2. PATCH /v1/offices/{id}/periods/report with new start\n3. Immediately GET /v1/offices/{id}/periods/report\n4. Verify updated value returned",
     "GET after PATCH returns updated period. @Cacheable on getPeriod() evicted per PATCH via SimpleKey(officeId, periodType). No stale cache served.",
     "Medium", "Functional", "PERIOD-CACHE-01", "OfficePeriodServiceImpl",
     "Cache eviction is per (officeId, periodType) key."),

    ("TC-ACC-013", "Concurrent period modifications: race condition handling",
     "Two ACCOUNTANT sessions for the same office.",
     "1. Session A: PATCH approve period to 2026-03-01\n2. Session B: simultaneously PATCH approve period to 2026-04-01\n3. Verify final state consistency",
     "One update wins, final state is consistent. Database-level constraint prevents invalid state. No deadlock or data corruption.",
     "Medium", "Concurrency", "PERIOD-CONCURRENT-01", "OfficePeriodServiceImpl",
     "No explicit optimistic/pessimistic locking found in code — relies on DB transaction isolation."),

    ("TC-ACC-014", "Report period equals approve period: boundary allowed",
     "ACCOUNTANT. Office with different report and approve periods.",
     "1. Set report period to match approve period exactly (both = 2026-02-01)\n2. Verify success\n3. Attempt to set approve period = report period\n4. Verify behavior",
     "Report = approve is allowed (strict < check, not <=). Setting approve to match report also allowed (strict > check allows equality). Equal periods represent 'caught up' state.",
     "Medium", "Boundary", "PERIOD-EQUAL-01", "OfficePeriodServiceImpl",
     "Edge case: approve can be equal to report but not exceed it."),
]))

# ==========================================================================
# TS-ACC-PeriodEffects -- Period Change Cross-Service Side Effects
# ==========================================================================
SUITES.append(("TS-ACC-PeriodEffects", "Period Change Cross-Service Side Effects", [
    ("TC-ACC-015", "Approve period advance triggers PeriodChangedEvent via RabbitMQ",
     "ACCOUNTANT. Office with unapproved reports in closing month. RabbitMQ running.",
     "1. PATCH /v1/offices/{id}/periods/approve to advance by 1 month\n2. Monitor RabbitMQ messages\n3. Check for PeriodChangedEvent publication",
     "PeriodChangedEvent published to RabbitMQ exchange. Event contains officeId and new period start. Consumed by vacation service for recalculation.",
     "Critical", "Integration", "PERIOD-EVENT-01", "OfficePeriodServiceImpl, RabbitMQ",
     "Key cross-service trigger. Vacation service listens for this event."),

    ("TC-ACC-016", "Approve period advance triggers auto-reject of unapproved reports",
     "ACCOUNTANT. Office with REPORTED-state reports in the month being closed.",
     "1. Create reports in REPORTED state for the closing month\n2. PATCH /v1/offices/{id}/periods/approve to advance\n3. Verify reports transitioned to REJECTED state\n4. Check reject table for auto-reject record",
     "All REPORTED-state reports in closing month auto-rejected. Single shared Reject record created with description='auto.reject.state'. Reports FK set to this reject record.",
     "Critical", "Integration", "PERIOD-AUTOREJECT-01", "TaskReportServiceImpl.rejectByOfficeId",
     "Auto-reject is side effect of period advance. Creates entry in reject table with special description."),

    ("TC-ACC-017", "Auto-reject sends email notifications to affected employees",
     "Office with employees having unapproved reports. Email service running.",
     "1. Advance approve period (triggering auto-reject)\n2. Verify email notifications sent\n3. Check notification content matches APPROVE_REJECT template",
     "Email notification sent per affected employee. Uses APPROVE_REJECT template. My Tasks page shows auto-reject warning: 'Unconfirmed hours for task {name} were auto-rejected upon month closure'.",
     "High", "Integration", "PERIOD-AUTOREJECT-02", "AutoRejectedReportsContainer",
     "Warning on My Tasks page stored in localStorage (hiddenAutoRejectWarnings)."),

    ("TC-ACC-018", "Approve period advance triggers vacation day recalculation",
     "ACCOUNTANT. Office with employees having vacation balances.",
     "1. Note employee vacation day balances before period advance\n2. PATCH /v1/offices/{id}/periods/approve to advance\n3. Verify vacation day recalculation triggered via RabbitMQ\n4. Check updated balances",
     "Vacation day balances recalculated. PeriodChangedEvent consumed by vacation service triggers bulk recalculation for office employees. Especially impacts advance-vacation offices (AV=true).",
     "High", "Integration", "PERIOD-RECALC-01", "VacationService, RabbitMQ",
     "Affects offices with advanceVacation=true: norm deviation recalculation."),

    ("TC-ACC-019", "Approve period revert triggers PeriodReopenedEvent",
     "ACCOUNTANT. Office with APPROVE period previously advanced.",
     "1. PATCH /v1/offices/{id}/periods/approve to revert by 1 month\n2. Monitor for PeriodReopenedEvent\n3. Verify employees can now edit reports in reopened month",
     "PeriodReopenedEvent published. Previously closed month reopened for report editing. No auto-reject reversal — rejected reports remain rejected.",
     "High", "Integration", "PERIOD-REOPEN-01", "OfficePeriodServiceImpl",
     "Revert does NOT undo auto-reject. Employees must manually re-report."),

    ("TC-ACC-020", "Extended period blocks approve period advancement",
     "ACCOUNTANT. Employee has active extended report period in the office.",
     "1. Create extended period: PUT /v1/periods/report/employees/{login}\n2. PATCH /v1/offices/{id}/periods/approve to advance\n3. Verify error response",
     "400 error: CODE_ERROR_APPROVE_CHANGE_MORE_THAN_ONE_MONTH or extended period block. Approve period cannot be advanced while any employee has active extended period in the office.",
     "High", "Negative", "PERIOD-EXTENDED-01", "OfficePeriodServiceImpl",
     "Extended periods block office-wide approve advance as safety mechanism."),

    ("TC-ACC-021", "Individual extended period: create and auto-cleanup",
     "ADMIN. Employee without existing extended period.",
     "1. PUT /v1/periods/report/employees/{login} to create extended period\n2. Verify extended period active\n3. Wait for ExtendedPeriodScheduler cleanup (every 5 min)\n4. Verify auto-removal after expiry",
     "Extended period created, visible in Individual Period Changing tab. ExtendedPeriodScheduler runs every 5 minutes — removes expired extended periods automatically.",
     "Medium", "Functional", "PERIOD-EXTENDED-02", "ExtendedPeriodScheduler",
     "Auto-cleanup prevents stale extended periods blocking office operations."),

    ("TC-ACC-022", "Period advance with no reports in closing month: clean advance",
     "ACCOUNTANT. Office with no REPORTED-state reports in closing month.",
     "1. Verify no pending reports for the closing month\n2. PATCH /v1/offices/{id}/periods/approve to advance\n3. Verify clean advance with no auto-reject side effects",
     "Period advanced successfully. No auto-reject triggered (no matching reports). PeriodChangedEvent still published. Vacation recalculation still triggers.",
     "Medium", "Functional", "PERIOD-CLEAN-01", "TaskReportServiceImpl",
     "Edge case: events fire even when no reports are affected."),

    ("TC-ACC-023", "Auto-reject warning display and dismissal on My Tasks page",
     "Employee whose reports were auto-rejected by period advance.",
     "1. Login as employee with auto-rejected reports\n2. Navigate to My Tasks page (/report)\n3. Verify AutoRejectedReportsContainer warning displayed\n4. Click 'Go to the report page' link\n5. Click close button to dismiss\n6. Refresh page — verify warning hidden via localStorage",
     "Warning notification shows at top of Report page: 'Unconfirmed hours for task {taskName} were automatically rejected upon month closure'. Close button stores to localStorage. Warning only queries previous month — older auto-rejections invisible.",
     "Medium", "UI", "PERIOD-AUTOREJECT-03", "AutoRejectedReportsContainer",
     "Design issue: single-month window, BO leak in controller, no manager notification."),

    ("TC-ACC-024", "Period advance effect on confirmation page visibility",
     "MANAGER. Approve period advanced closing a month.",
     "1. Note confirmation page state before period advance\n2. Accountant advances approve period\n3. Refresh confirmation page (/approve)\n4. Verify closed month no longer shows pending items",
     "Reports in closed month no longer appear in confirmation page filters. Week tabs update to reflect new approve period boundary. Previously pending reports now show as REJECTED.",
     "Medium", "Integration", "PERIOD-CONFIRM-01", "Frontend approve module",
     "Managers see closed month items as rejected, can no longer approve them."),

    ("TC-ACC-025", "Report period advance effect on employee report submission",
     "EMPLOYEE. Report period advanced forward.",
     "1. Note current report period\n2. Accountant advances report period forward by 1 month\n3. Employee attempts to report for the previous (now-closed) month\n4. Verify error",
     "400 error: reportDate before office report period start. Employee can only report for months >= report period start. Closing report period blocks historical edits.",
     "High", "Integration", "PERIOD-REPORT-EFFECT-01", "TaskReportServiceImpl",
     "Key business rule: report period gates employee report submission."),
]))

# ==========================================================================
# TS-ACC-PayValidation -- Payment Validation Edge Cases & Bugs
# ==========================================================================
SUITES.append(("TS-ACC-PayValidation", "Payment Validation Edge Cases & Bugs", [
    ("TC-ACC-026", "Payment days mismatch: regular + admin != vacation total",
     "ACCOUNTANT. APPROVED vacation with 5 days.",
     "1. PUT /v1/vacations/pay/{id} with regularDaysPayed=3, administrativeDaysPayed=1 (total=4, vacation=5)\n2. Verify error response",
     "400 error: 'exception.vacation.pay.days.not.equal'. Validation: regularDaysPayed + administrativeDaysPayed must equal vacation.getDays(). Sum check at checkForPayment().",
     "Critical", "Validation", "PAY-VALIDATION-01", "PayVacationService.checkForPayment",
     "Core validation rule. Prevents incorrect day accounting."),

    ("TC-ACC-027", "Payment of already-PAID vacation rejected",
     "ACCOUNTANT. Vacation already in PAID status.",
     "1. PUT /v1/vacations/pay/{id} with correct days\n2. Verify error response",
     "400 error: 'exception.vacation.status.notAllowed'. Only APPROVED vacations can be paid. PAID status blocks re-payment.",
     "High", "Negative", "PAY-VALIDATION-02", "PayVacationService.checkForPayment",
     "Prevents double payment."),

    ("TC-ACC-028", "Payment of APPROXIMATE period vacation rejected",
     "ACCOUNTANT. Vacation with period type APPROXIMATE (not EXACT).",
     "1. Find vacation with APPROXIMATE period (future vacation without exact dates)\n2. PUT /v1/vacations/pay/{id}\n3. Verify error response",
     "400 error. Period type must be EXACT for payment. APPROXIMATE vacations cannot be processed for payment until dates are confirmed.",
     "High", "Negative", "PAY-VALIDATION-03", "PayVacationService.checkForPayment",
     "Period precision check. APPROXIMATE = tentative, EXACT = confirmed dates."),

    ("TC-ACC-029", "BUG: Payment type misalignment — ADMIN vacation paid as regular",
     "ACCOUNTANT. APPROVED ADMINISTRATIVE vacation (paymentType=ADMINISTRATIVE).",
     "1. PUT /v1/vacations/pay/{id} with regularDaysPayed=1, administrativeDaysPayed=0\n2. Verify response (expect error but currently 200)\n3. Check vacation_payment record in DB",
     "BUG (BUG-PAY-2): 200 OK. Payment accepted with wrong type distribution. checkForPayment validates only total match (regular+admin==total), NOT that distribution matches paymentType. Allows incorrect accounting classification.",
     "High", "Bug", "PAY-VALIDATION-04, BUG-PAY-2", "PayVacationService.checkForPayment",
     "MEDIUM severity. Accounting impact: day type reported incorrectly in financial records."),

    ("TC-ACC-030", "Payment with negative days rejected by @Range validation",
     "ACCOUNTANT. APPROVED vacation.",
     "1. PUT /v1/vacations/pay/{id} with regularDaysPayed=-1, administrativeDaysPayed=0\n2. Verify 400 response with validation error",
     "400 error: 'must be between 0 and 366'. @Range(min=0, max=366) annotation on DTO fields. Bean validation triggers before business logic.",
     "Medium", "Validation", "PAY-VALIDATION-05", "VacationPaymentDto",
     "@Range(0-366) on both regularDaysPayed and administrativeDaysPayed."),

    ("TC-ACC-031", "Payment with days > 366 rejected by @Range validation",
     "ACCOUNTANT. APPROVED vacation.",
     "1. PUT /v1/vacations/pay/{id} with regularDaysPayed=400, administrativeDaysPayed=0\n2. Verify 400 response",
     "400 error: 'must be between 0 and 366'. Upper bound of @Range annotation reached. Even if vacation had 400 days (impossible), the field-level validation rejects.",
     "Low", "Boundary", "PAY-VALIDATION-06", "VacationPaymentDto",
     "Defensive upper bound check."),

    ("TC-ACC-032", "Payment with null/empty body rejected",
     "ACCOUNTANT. APPROVED vacation.",
     "1. PUT /v1/vacations/pay/{id} with empty body {} \n2. PUT with body {\"regularDaysPayed\": null}\n3. Verify 400 responses",
     "400 error: 'regularDaysPayed must not be null' and/or 'administrativeDaysPayed must not be null'. @NotNull validation on DTO fields.",
     "Medium", "Validation", "PAY-VALIDATION-07", "VacationPaymentDto",
     "@NotNull on both payment day fields."),

    ("TC-ACC-033", "Payment for nonexistent vacation ID returns 400 (not 404)",
     "ACCOUNTANT.",
     "1. PUT /v1/vacations/pay/999999 with valid body\n2. Verify error response\n3. Note status code (400 vs 404)",
     "400 error: 'Vacation id not found'. Returns 400 instead of REST-standard 404. Error response format inconsistency with other services.",
     "Medium", "Negative", "PAY-VALIDATION-08", "PayVacationService",
     "Error response inconsistency: should be 404 per REST conventions."),

    ("TC-ACC-034", "Payment for CANCELED vacation rejected",
     "ACCOUNTANT. Vacation in CANCELED status.",
     "1. Find CANCELED vacation in DB\n2. PUT /v1/vacations/pay/{id}\n3. Verify error response",
     "400 error: 'exception.vacation.status.notAllowed'. Only APPROVED status allows payment. CANCELED, REJECTED, NEW all blocked.",
     "Medium", "Negative", "PAY-VALIDATION-09", "PayVacationService.checkForPayment",
     "Status check: statusAllowed = [APPROVED] only."),

    ("TC-ACC-035", "BUG: Payment dates endpoint accepts start > end",
     "ACCOUNTANT.",
     "1. GET /v1/paymentdates?vacationStartDate=2026-04-01&vacationEndDate=2026-03-01 (reversed)\n2. Verify response",
     "BUG (BUG-PAY-3): Returns valid results (same as normal range). No validation that vacationStartDate <= vacationEndDate. Should reject reversed date range.",
     "Medium", "Bug", "PAY-VALIDATION-10, BUG-PAY-3", "PaymentDateService",
     "LOW severity. Misleading but not data-corrupting."),

    ("TC-ACC-036", "BUG: Available paid days accepts negative newDays",
     "ACCOUNTANT.",
     "1. GET /v1/vacationdays/available?employeeLogin=X&paymentDate=2026-03-01&newDays=-5\n2. Verify response",
     "BUG (BUG-PAY-5): Returns availablePaidDays=16.0 without error. Negative newDays should be rejected. Non-positive values produce meaningless calculations.",
     "Medium", "Bug", "PAY-VALIDATION-11, BUG-PAY-5", "EmployeeDaysService",
     "LOW severity. Input validation gap."),

    ("TC-ACC-037", "BUG: Stack trace leakage on invalid payment date format",
     "ACCOUNTANT.",
     "1. GET /v1/paymentdates?vacationStartDate=2026-13-01&vacationEndDate=2026-14-01\n2. Inspect response body",
     "BUG (BUG-PAY-6): Full Spring exception with class names and conversion details in response body. Information disclosure risk. Should return clean 400 with message.",
     "Medium", "Security", "PAY-VALIDATION-12, BUG-PAY-6", "Vacation REST controller",
     "Information disclosure. Leaks internal class names and framework details."),

    ("TC-ACC-038", "BUG: DB/API data representation inconsistency for ADMINISTRATIVE vacations",
     "ACCOUNTANT. ADMINISTRATIVE type vacation.",
     "1. Query DB: SELECT regular_days, administrative_days FROM vacation WHERE payment_type='ADMINISTRATIVE'\n2. GET /v2/vacations/{id} via API\n3. Compare regularDays/administrativeDays values",
     "BUG (BUG-PAY-4): DB stores ADMINISTRATIVE vacation days in regular_days column (e.g. regular_days=1, administrative_days=0). API returns them swapped: regularDays=0, administrativeDays=1. DTO conversion transposes based on payment_type. DB queries give wrong day-type breakdown.",
     "High", "Bug", "PAY-VALIDATION-13, BUG-PAY-4", "VacationMapper/DTO conversion",
     "MEDIUM severity. Impacts any direct DB reporting queries."),
]))

# ==========================================================================
# TS-ACC-PayLifecycle -- Payment Lifecycle, Auto-Payment & Day Return
# ==========================================================================
SUITES.append(("TS-ACC-PayLifecycle", "Payment Lifecycle, Auto-Payment & Day Return", [
    ("TC-ACC-039", "Single payment: APPROVED → PAID status transition",
     "ACCOUNTANT. APPROVED REGULAR vacation (e.g. 5 days).",
     "1. PUT /v1/vacations/pay/{id} with regularDaysPayed=5, administrativeDaysPayed=0\n2. Verify 200 response\n3. GET /v2/vacations/{id} — verify status = PAID\n4. Check vacation_payment record in DB",
     "Status transitions to PAID. vacation_payment record created with correct day split. VacationStatusChangedEvent published. Days NOT deducted (already deducted at approval time).",
     "Critical", "Functional", "PAY-LIFECYCLE-01", "PayVacationService.payVacation",
     "Key behavior: days deducted at APPROVAL, not payment. Payment is accounting-only transition."),

    ("TC-ACC-040", "Partial payment: fewer regular days → remainder returned to balance",
     "ACCOUNTANT. APPROVED REGULAR vacation (10 days). Employee has available balance.",
     "1. Note employee vacation day balances (current + next year)\n2. PUT /v1/vacations/pay/{id} with regularDaysPayed=7, administrativeDaysPayed=3\n3. Verify balance changes\n4. Check day return logic",
     "3 days returned to balance. Day return priority: nextYearAvailableDays restored first (up to cap of 20), then currentYearDays. Employee balance increases by 3 unpaid regular days.",
     "Critical", "Functional", "PAY-LIFECYCLE-02", "PayVacationService.payVacation",
     "Day return logic: FIFO reverse — next year first, then current year."),

    ("TC-ACC-041", "Day return: nextYearAvailableDays cap at 20",
     "ACCOUNTANT. Employee with nextYearAvailableDays=19. APPROVED vacation 5 days.",
     "1. Note nextYearAvailableDays=19\n2. Pay with regularDaysPayed=2, administrativeDaysPayed=3 (3 unpaid)\n3. Verify day return: 1 to nextYear (cap at 20), 2 to currentYear",
     "NextYear gets 1 day (19→20, cap reached). Remaining 2 days go to currentYear balance. NEW_YEAR_VACATION_DAYS constant = 20 enforced as ceiling.",
     "High", "Boundary", "PAY-LIFECYCLE-03", "VacationDaysBO",
     "Cap prevents excessive next-year accrual beyond annual entitlement."),

    ("TC-ACC-042", "Batch payment: 'Pay all checked requests' processes multiple vacations",
     "ACCOUNTANT. Multiple APPROVED vacations selected via checkboxes on payment page.",
     "1. Navigate to Vacation Payment page (/vacation/payment)\n2. Select 3+ vacation requests via checkboxes\n3. Click 'Pay all checked requests'\n4. Confirm in popup\n5. Verify all selected vacations transition to PAID",
     "All selected vacations processed. Each generates VacationStatusChangedEvent. Batch payment uses auto-distribution based on paymentType (REGULAR→all regular, ADMINISTRATIVE→all admin). No individual day split selection in batch mode.",
     "High", "Functional", "PAY-LIFECYCLE-04", "PayVacationService",
     "Batch popup cannot customize regular/admin split per vacation — auto-distributes."),

    ("TC-ACC-043", "Concurrent payment: write lock prevents duplicate processing",
     "Two ACCOUNTANT sessions attempting to pay same vacation simultaneously.",
     "1. Session A: PUT /v1/vacations/pay/{id} (in flight)\n2. Session B: PUT /v1/vacations/pay/{id} simultaneously\n3. Verify one succeeds, other fails\n4. Check for data corruption",
     "Write lock acquired by first request. Second request either waits (lock contention) or fails with lock error. No duplicate payment created. vacation_payment table has single entry.",
     "High", "Concurrency", "PAY-LIFECYCLE-05", "PayVacationService (write lock)",
     "Write lock acquired at method entry. Prevents race conditions."),

    ("TC-ACC-044", "Auto-payment cron: payExpiredApproved processes old APPROVED vacations",
     "APPROVED vacations older than 2 months exist in DB.",
     "1. Identify APPROVED vacations with start_date > 2 months ago\n2. Trigger payExpiredApproved (via test API or wait for cron)\n3. Verify auto-paid with correct day distribution\n4. Check ShedLock record",
     "APPROVED vacations older than today.minusMonths(2).withDayOfMonth(2) auto-paid. Day distribution: REGULAR type → all regularDays, ADMINISTRATIVE type → all adminDays. ShedLock prevents concurrent cron execution across instances.",
     "High", "Functional", "PAY-LIFECYCLE-06", "VacationPaymentScheduler",
     "Cron: every 10 min (0 */10 * * * *). ShedLock for distributed safety."),

    ("TC-ACC-045", "BUG: VacationStatusUpdateJob 2-hour orphan window",
     "NEW_FOR_PAID status update entries in status_updates table.",
     "1. Query status_updates for NEW_FOR_PAID entries created > 2 hours ago\n2. Trigger VacationStatusUpdateJob (cron or test API)\n3. Verify orphaned entries NOT processed\n4. Check for stuck entries",
     "BUG (BUG-PAY-1): Entries older than 2 hours permanently orphaned. findRecentNew(now.minusHours(2)) query window excludes old entries. No cleanup/retry mechanism. Found 6 stuck entries for Saturn office.",
     "Critical", "Bug", "PAY-LIFECYCLE-07, BUG-PAY-1", "VacationStatusUpdateJob",
     "HIGH severity. Processing window creates permanent orphans. No alerting or recovery."),

    ("TC-ACC-046", "Payment timeline audit: incomplete event fields",
     "ACCOUNTANT. Pay a vacation and check timeline.",
     "1. PUT /v1/vacations/pay/{id}\n2. Query vacation timeline events\n3. Check VACATION_PAID event fields",
     "Timeline event created with type VACATION_PAID. However: days_used=0, administrative_days_used=0 (always zero). previous_status=NULL. Audit trail for payment is incomplete — cannot reconstruct payment details from timeline alone.",
     "Medium", "Functional", "PAY-LIFECYCLE-08", "VacationTimelineService",
     "Audit gap: payment event doesn't record day split or pre-payment status."),

    ("TC-ACC-047", "Payment page: ADMINISTRATIVE vacations cannot be paid via UI",
     "ACCOUNTANT. Vacation Payment page with ADMINISTRATIVE vacation requests.",
     "1. Navigate to /vacation/payment\n2. Find ADMINISTRATIVE vacation in table\n3. Verify no checkbox, no status, no action buttons",
     "ADMINISTRATIVE vacations displayed in table but have no status column value, no action buttons, no checkbox for selection. Cannot be marked as paid through UI. Must use API or batch payment button.",
     "Medium", "UI", "PAY-LIFECYCLE-09", "Frontend VacationPayment component",
     "UI limitation: ADMINISTRATIVE vacations visible but not actionable individually."),

    ("TC-ACC-048", "Payment month quick tabs navigate correctly",
     "ACCOUNTANT on Vacation Payment page.",
     "1. Navigate to /vacation/payment\n2. Verify month quick tabs (Jan-May 2026 visible)\n3. Click each tab\n4. Verify table filters to selected month\n5. Use month picker for non-quick-tab month",
     "Quick tabs filter payment table to selected month. Month picker allows selection of any month. Unpaid vacation alert banner visible at top when unpaid requests exist for any month.",
     "Low", "UI", "PAY-LIFECYCLE-10", "Frontend VacationPayment component",
     "Quick tabs provide fast navigation; month picker for full range."),
]))

# ==========================================================================
# TS-ACC-DayCorrect -- Vacation Day Correction Edge Cases
# ==========================================================================
SUITES.append(("TS-ACC-DayCorrect", "Vacation Day Correction Edge Cases", [
    ("TC-ACC-049", "Manual day correction with comment creates audit trail",
     "ACCOUNTANT. Employee with vacation day balance.",
     "1. Navigate to /vacation/days-correction\n2. Find employee, click inline edit on 'Vacation days' cell\n3. Change value (e.g. 24 → 20)\n4. Confirm edit with comment explaining reason\n5. Open Events Feed dialog for same employee",
     "Balance updated to new value. Events feed shows correction event with: date, 'Day correction' event type, previous/new value, comment text. PUT /v1/vacationdays/{login} called with updated value.",
     "High", "Functional", "DAYCORR-AUDIT-01", "EmployeeDaysService",
     "Comment is audit trail for accounting corrections. Max 255 chars (backend @Size)."),

    ("TC-ACC-050", "Day correction: negative balance allowed",
     "ACCOUNTANT. Employee with positive vacation days balance.",
     "1. Edit employee vacation days to negative value (e.g. -5)\n2. Verify correction accepted\n3. Check employee vacation days balance",
     "Negative balance accepted and stored. System allows negative vacation days. This differs from regular vacation calculation which shows 0 for negative (RegularCalculation only). AdvanceCalculation offices can legitimately go negative.",
     "High", "Boundary", "DAYCORR-NEGATIVE-01", "EmployeeDaysService",
     "Negative balances: RegularCalc displays 0, AdvanceCalc shows actual negative."),

    ("TC-ACC-051", "Bulk recalculation for salary office",
     "ACCOUNTANT. Salary office with multiple employees.",
     "1. POST /v1/vacationdays/recalculate with office identifier\n2. Verify all employees in office recalculated\n3. Check for balance changes",
     "Bulk recalculation processes all active employees in the salary office. Uses FIFO day consumption logic. Returns all regular days to balance then re-distributes among NEW/APPROVED vacations. If insufficient, auto-converts to ADMINISTRATIVE.",
     "High", "Functional", "DAYCORR-BULK-01", "EmployeeDaysService",
     "Heavy operation. May change balances for all employees in office."),

    ("TC-ACC-052", "Day correction events feed shows complete history",
     "ACCOUNTANT. Employee with vacation history (corrections, vacations, payments).",
     "1. Navigate to /vacation/days-correction\n2. Click events feed button for employee with rich history\n3. Verify dialog content",
     "Events feed dialog shows: employee name, annual vacation days left, work dates (start-end), events table with columns: Date, Event type, Paid/Unpaid days allowance, Paid/Unpaid days used. Total row at bottom. All corrections, approvals, payments, cancellations listed chronologically.",
     "Medium", "Functional", "DAYCORR-EVENTS-01", "VacationTimeline",
     "Events feed is primary audit mechanism for day balance changes."),

    ("TC-ACC-053", "Day correction: dismissed employee filter toggle",
     "ACCOUNTANT. Both active and dismissed employees exist.",
     "1. Navigate to /vacation/days-correction\n2. Default view: only active employees\n3. Toggle 'Show dismissed employees' checkbox\n4. Verify dismissed employees appear in table\n5. Attempt to edit dismissed employee's days",
     "Dismissed employees appear when filter enabled. Day correction should still work for dismissed employees (adjustments may be needed for final accounting). Dismissed employees have distinct visual indication.",
     "Medium", "Functional", "DAYCORR-DISMISSED-01", "Frontend DaysCorrection component",
     "Important for final settlement accounting."),

    ("TC-ACC-054", "Available paid days calculation: binary search mode (newDays=0)",
     "ACCOUNTANT.",
     "1. GET /v1/vacationdays/available?employeeLogin=X&paymentDate=2026-03-01&newDays=0\n2. Verify response",
     "Returns maximum safe vacation duration via binary search. availablePaidDays reflects max days employee can take without insufficient balance. daysNotEnough list shows future vacations at risk if more days used.",
     "Medium", "Functional", "DAYCORR-AVAILABLE-01", "VacationAvailablePaidDaysCalculatorImpl",
     "Binary search: O(N × log(maxDays)) calls to calculate. Computationally expensive."),

    ("TC-ACC-055", "Available paid days: daysNotEnough warning for at-risk vacations",
     "ACCOUNTANT. Employee with multiple future APPROVED vacations near balance limit.",
     "1. GET /v1/vacationdays/available with newDays close to total balance\n2. Check daysNotEnough field in response",
     "Response includes daysNotEnough: list of future vacation IDs at risk of insufficient days if the proposed new vacation is approved. Helps accountant assess impact of day adjustments on future obligations.",
     "Medium", "Functional", "DAYCORR-AVAILABLE-02", "VacationAvailablePaidDaysCalculatorImpl",
     "Warning system for cascading balance impacts."),

    ("TC-ACC-056", "Day correction: maternity special case — all year balances summed",
     "ACCOUNTANT. Employee with maternity=true flag.",
     "1. Find employee with maternity=true in DB\n2. GET /v1/vacationdays/available for this employee\n3. Verify available days = sum of ALL year balances",
     "Maternity employees: available = sum of ALL year balances (no year restriction). Normal employees limited to current + next year. Maternity accumulation allows multi-year rollover.",
     "Medium", "Boundary", "DAYCORR-MATERNITY-01", "VacationCalculationStrategy",
     "Special case: maternity=true bypasses year restrictions."),

    ("TC-ACC-057", "No-pagination on vacation days list (1609 records)",
     "ACCOUNTANT.",
     "1. GET /v1/vacationdays (no pagination params)\n2. Count returned records\n3. Measure response time",
     "Returns all 1609 employee records in single response. No pagination support. Response may be slow and large. Contrast with v2 APIs that use pageSize/totalCount pagination.",
     "Low", "Performance", "DAYCORR-PERF-01", "EmployeeDaysController",
     "Performance risk. No pagination = entire dataset per request."),

    ("TC-ACC-058", "Day correction: cross-year balance redistribution on FIFO recalc",
     "ACCOUNTANT. Employee with multiple year balances and APPROVED vacations.",
     "1. Note per-year balances via GET /v1/vacationdays/{login}/years\n2. Trigger recalculation\n3. Compare per-year balances before and after",
     "FIFO consumption: days consumed from earliest year first. On recalc: all regular days returned to pool, then re-distributed among NEW/APPROVED using FIFO. If insufficient total, later vacations auto-converted to ADMINISTRATIVE.",
     "High", "Functional", "DAYCORR-FIFO-01", "VacationDaysService",
     "FIFO logic can change auto-type assignment of existing vacations."),
]))

# ==========================================================================
# TS-ACC-Notifications -- Accounting Notification Triggers & Templates
# ==========================================================================
SUITES.append(("TS-ACC-Notifications", "Accounting Notification Triggers & Templates", [
    ("TC-ACC-059", "Salary page: 'Notify all managers' button sends bulk notifications",
     "ACCOUNTANT on Salary page (/admin/salary).",
     "1. Navigate to /admin/salary\n2. Click 'Notify all managers' button\n3. Verify confirmation dialog\n4. Confirm\n5. Check email service for sent notifications",
     "Bulk notification sent. Uses APPROVE_REQUEST template. Sent to all managers with unconfirmed employee reports. CC's the triggering accountant. POST /v1/reports/accounting/notifications called.",
     "High", "Functional", "NOTIF-BULK-01", "TaskReportAccountingService.notifyManagers",
     "Permission: ACCOUNTING.NOTIFY required."),

    ("TC-ACC-060", "Salary page: individual manager notification (envelope icon)",
     "ACCOUNTANT on Salary page. Manager with unconfirmed reports.",
     "1. Navigate to /admin/salary\n2. Find row with manager who has unconfirmed reports\n3. Click envelope icon in that row\n4. Verify notification sent to that specific manager",
     "Individual notification sent to selected manager. Uses APPROVE_REQUEST_FOR_EMPLOYEE template. CC's the triggering accountant. More targeted than bulk notification.",
     "Medium", "Functional", "NOTIF-INDIVIDUAL-01", "TaskReportAccountingService",
     "Per-row action button. Template includes specific employee/manager context."),

    ("TC-ACC-061", "Accounting notification email templates: APPROVE_REQUEST and APPROVE_REQUEST_FOR_EMPLOYEE",
     "Email service running. Accountant triggers notifications.",
     "1. Trigger 'Notify all managers' → captures APPROVE_REQUEST template\n2. Trigger individual notification → captures APPROVE_REQUEST_FOR_EMPLOYEE template\n3. Verify template content, subject, recipients",
     "APPROVE_REQUEST: sent to all managers with pending reports. APPROVE_REQUEST_FOR_EMPLOYEE: sent to specific manager for specific employee. Both CC the triggering accountant. Both include salary office, period, and report details.",
     "Medium", "Functional", "NOTIF-TEMPLATE-01", "Email service templates",
     "120 templates total, 70 active. These 2 are accounting-specific."),

    ("TC-ACC-062", "Auto-reject notification: email sent to affected employees",
     "Period advance triggers auto-reject of reports.",
     "1. Set up REPORTED-state reports for an employee\n2. Advance approve period\n3. Check reject table for auto-reject record (description='auto.reject.state')\n4. Wait for sendRejectNotifications scheduler (every 5 min)\n5. Verify email sent",
     "sendRejectNotifications scheduler picks up new reject records every 5 min. Sends APPROVE_REJECT template to executor (employee). executor_notified flag set to true after send. Email contains task name, rejection reason, period info.",
     "High", "Integration", "NOTIF-AUTOREJECT-01", "RejectNotificationScheduler",
     "5-minute scheduler delay between reject creation and notification."),

    ("TC-ACC-063", "Budget over/under-reporting banner notification triggers",
     "ADMIN/PM/SPM role. Employees with over or under-reported hours.",
     "1. Login as PM/ADMIN on Confirmation page (/approve)\n2. Check for non-dismissible banner at top\n3. Verify triggers: over-reporting in current month OR over/under-reporting at month end",
     "Banner visible when: excess > 0% (over, red highlight) or excess < 0% (under, purple highlight). Shows clock icon with tooltip: deviation %, month, DM, projects with PM names. Non-dismissible — persists until resolved.",
     "Medium", "UI", "NOTIF-BUDGET-01", "OverReportedBanner component",
     "Calculation: excess = (reported - budgetNorm) / budgetNorm × 100%. ExcessStatus: HIGH/LOW/NEUTRAL/NA."),

    ("TC-ACC-064", "Forgotten report notification scheduler",
     "Employees with < 90% of personal norm reported.",
     "1. Verify sendReportsForgottenNotifications runs Mon/Fri 16:00\n2. Check employees below 90% threshold\n3. Verify notification emails sent",
     "Employees below 90% personal norm threshold receive forgotten report notification. Runs Mon/Fri at 16:00. Daily retry at 16:30 for deferred notifications. Configurable thresholds in TTT Parameters.",
     "Medium", "Functional", "NOTIF-FORGOTTEN-01", "ReportForgottenNotificationScheduler",
     "Two schedulers: main (Mon/Fri 16:00) and delayed retry (daily 16:30)."),

    ("TC-ACC-065", "Accounting report search: permission and auth behavior",
     "Users with different authentication methods.",
     "1. GET /v1/reports/accounting with JWT (accountant user) → expect 200\n2. GET /v1/reports/accounting with API token (ACCOUNTING.VIEW permission) → verify behavior\n3. Compare responses",
     "JWT: works with ACCOUNTING.VIEW permission. API token: returns 403 despite having ACCOUNTING.VIEW — token-vs-session auth discrepancy. Service-layer permission checks behave differently for different auth mechanisms.",
     "High", "Security", "NOTIF-AUTH-01", "TaskReportAccountingService",
     "Known auth gap: API token user has roles but service-layer permission checks reject."),

    ("TC-ACC-066", "Reports changed notification: scheduled daily for manager-reported hours",
     "Manager has reported hours on behalf of employee.",
     "1. Manager reports hours for another employee via API\n2. Wait for sendReportsChangedNotifications (daily 07:50)\n3. Verify employee receives notification",
     "Daily scheduler (07:50) detects reports changed by someone other than executor. Sends notification to the affected employee. Includes task name, date, hours changed.",
     "Medium", "Functional", "NOTIF-CHANGED-01", "ReportsChangedNotificationScheduler",
     "Detects cross-employee reporting. Daily 07:50 schedule."),
]))

# ==========================================================================
# TS-ACC-SickLeaveAcct -- Sick Leave Accounting Workflow Gaps
# ==========================================================================
SUITES.append(("TS-ACC-SickLeaveAcct", "Sick Leave Accounting Workflow Gaps", [
    ("TC-ACC-067", "Sick leave accounting status workflow: New → Pending → Paid",
     "ACCOUNTANT on Sick Leave Records page (/accounting/sick-leaves).",
     "1. Find sick leave with status 'New'\n2. Change status dropdown to 'Pending'\n3. Verify status updated\n4. Change status to 'Paid'\n5. Verify final status",
     "Status transitions: New → Pending → Paid. Each transition updates accounting_status in DB. Status dropdown shows only valid next states. Paid is terminal state.",
     "High", "Functional", "SL-ACCT-01", "SickLeaveAccountingService",
     "Dual status model: sick leave status (OPEN/CLOSED/DELETED) × accounting status (NEW/PENDING/PAID/REJECTED)."),

    ("TC-ACC-068", "Sick leave accounting status: Reject workflow",
     "ACCOUNTANT. Sick leave with status 'New' or 'Pending'.",
     "1. Find sick leave with status 'New'\n2. Change status to 'Rejected'\n3. Verify status updated\n4. Attempt to change Rejected back to New/Pending\n5. Verify behavior",
     "New → Rejected: allowed. Pending → Rejected: allowed. Rejected is terminal (no transition back to New/Pending). Rejected sick leaves retain original sick leave status (OPEN/CLOSED) unchanged.",
     "High", "Functional", "SL-ACCT-02", "SickLeaveAccountingService",
     "Rejected is terminal for accounting status. Cannot be reopened."),

    ("TC-ACC-069", "Sick leave accounting: Deleted state blocks status changes",
     "ACCOUNTANT. Sick leave with state 'Deleted'.",
     "1. Find deleted sick leave in table (state = Deleted)\n2. Verify status dropdown is disabled/absent\n3. Verify no action buttons available",
     "Deleted sick leaves cannot have accounting status changed. Status dropdown not available. Actions column shows no buttons. Deleted state overrides accounting workflow.",
     "Medium", "Negative", "SL-ACCT-03", "Frontend SickLeaveAccounting component",
     "Deleted = hard block. Accounting cannot process deleted sick leaves."),

    ("TC-ACC-070", "Sick leave accounting: overdue alert display",
     "Sick leaves with overdue state (past end date, still OPEN).",
     "1. Navigate to /accounting/sick-leaves\n2. Filter by state = 'Overdue'\n3. Verify overdue entries highlighted in red\n4. Check overdue count badge in navigation",
     "Overdue sick leaves highlighted with red background. Count reflected in nav badge. Overdue = OPEN + endDate < today. Alert helps accountant prioritize processing.",
     "Medium", "UI", "SL-ACCT-04", "Frontend SickLeaveAccounting component",
     "GET /v1/accounting/sick-leaves/overdue/count endpoint provides count."),

    ("TC-ACC-071", "Sick leave accounting: concurrent status change by two accountants",
     "Two ACCOUNTANT sessions viewing same sick leave.",
     "1. Accountant A: opens sick leave status dropdown\n2. Accountant B: changes status from New to Pending\n3. Accountant A: submits change from New to Paid\n4. Verify final state and conflict handling",
     "Last write wins or conflict error. No optimistic locking visible in sick leave accounting. Potential for race condition where Accountant A's change overwrites B's intermediate state.",
     "Medium", "Concurrency", "SL-ACCT-05", "SickLeaveAccountingService",
     "No explicit concurrency control found in sick leave accounting flow."),

    ("TC-ACC-072", "Sick leave accounting: filter by salary office",
     "ACCOUNTANT with access to multiple salary offices.",
     "1. Navigate to /accounting/sick-leaves\n2. Select salary office filter\n3. Verify table filters to selected office\n4. Verify accountant sees only their assigned offices (if not chief accountant)",
     "Table filtered by selected salary office. ACCOUNTANT sees their office(s) only. CHIEF_ACCOUNTANT sees all offices. ADMIN sees all offices.",
     "Medium", "Functional", "SL-ACCT-06", "Frontend SickLeaveAccounting component",
     "Role-based office visibility: ACC=own, CACC/ADMIN=all."),

    ("TC-ACC-073", "Sick leave accounting: sort by dates (descending default)",
     "ACCOUNTANT on Sick Leave Records page.",
     "1. Navigate to /accounting/sick-leaves\n2. Verify default sort: sick leave dates descending\n3. Click date column header to toggle sort\n4. Verify ascending sort works",
     "Default: descending by sick leave dates (newest first). Click toggles asc/desc. All sortable columns work: Employee, Dates, Days, Work days, Sick note number, Salary office.",
     "Low", "UI", "SL-ACCT-07", "Frontend SickLeaveAccounting component",
     "Existing Qase suite 230 covers sort/filter but not all edge cases."),
]))

# ==========================================================================
# TS-ACC-APIErrors -- API Error Handling, Auth Gaps & Information Disclosure
# ==========================================================================
SUITES.append(("TS-ACC-APIErrors", "API Error Handling, Auth Gaps & Information Disclosure", [
    ("TC-ACC-074", "GET /v1/reports/accounting returns 403 despite valid ACCOUNTING.VIEW permission",
     "API token user with ACCOUNTING.VIEW permission. JWT accountant user.",
     "1. GET /v1/reports/accounting with API_SECRET_TOKEN → verify 403\n2. GET /v1/reports/accounting with JWT (accountant) → verify 200\n3. GET /v1/authentication/permissions → verify ACCOUNTING: [VIEW, NOTIFY]",
     "API token: 403 despite having ACCOUNTING.VIEW permission. JWT: 200 success. Service-layer permission checks behave differently for token vs session auth. AUTHENTICATED_USER authority required but not granted to API tokens.",
     "High", "Security", "API-AUTH-01", "TaskReportAccountingController",
     "Known auth gap. Token has permission but lacks AUTHENTICATED_USER authority."),

    ("TC-ACC-075", "BUG: status=ALL causes 500 NPE on vacation list",
     "Any authenticated user.",
     "1. GET /v2/vacations?status=ALL\n2. Verify 500 response\n3. Check error details",
     "BUG: 500 NullPointerException at VacationRepositoryCustomImpl.buildCommonCondition:433. ALL enum passed as null to repository which doesn't handle null status. Other status values (APPROVED, PAID, etc.) work correctly.",
     "High", "Bug", "API-ERROR-01", "VacationRepositoryCustomImpl:433",
     "NPE in repository layer. Missing null handling for ALL enum value."),

    ("TC-ACC-076", "Information disclosure: multiple endpoints return Java stack traces",
     "Any authenticated user. Invalid inputs.",
     "1. GET /v1/paymentdates with invalid date format → check for stack trace\n2. PATCH period with null body → check for stack trace\n3. GET /v1/offices/*/periods with invalid format → check for stack trace",
     "Multiple endpoints return full Java stack traces (class names, method names, line numbers, package structure). ExceptionHandler doesn't catch all exception types. Leaks internal architecture details.",
     "High", "Security", "API-SECURITY-01", "Global ExceptionHandler",
     "OWASP: Improper Error Handling. Stack traces should never reach client."),

    ("TC-ACC-077", "Pagination inconsistency: v1 vs v2 API responses",
     "Any authenticated user.",
     "1. GET v1 paginated endpoint → check pagination fields (size, totalElements)\n2. GET v2 paginated endpoint → check pagination fields (pageSize, totalCount)\n3. GET /v1/vacationdays → verify no pagination at all\n4. Document inconsistencies",
     "v1 uses: size, totalElements, totalPages. v2 uses: pageSize, totalCount. Vacation days list (/v1/vacationdays) has NO pagination — returns all 1609 records. Inconsistent pagination contracts across API versions.",
     "Medium", "Consistency", "API-CONSISTENCY-01", "Multiple controllers",
     "3 different pagination approaches in same API. Impacts API consumer development."),

    ("TC-ACC-078", "Error response inconsistency: TTT vs Vacation service errors",
     "Various error-triggering requests across services.",
     "1. Trigger error in TTT service → check errorCode field present\n2. Trigger error in Vacation service → check for trace field\n3. Request nonexistent vacation ID → verify 400 (not 404)\n4. Request nonexistent office ID → verify 200 (not 404)",
     "TTT returns structured errors with errorCode. Vacation sometimes includes 'trace' field, sometimes doesn't. Nonexistent vacation: 400 (wrong — should be 404). Nonexistent office: 200 with default data (wrong — should be 404). No consistent error contract.",
     "Medium", "Consistency", "API-CONSISTENCY-02", "TTT/Vacation error handlers",
     "Cross-service error handling inconsistency complicates error parsing."),

    ("TC-ACC-079", "Period min/max endpoints: cross-office aggregation correctness",
     "ACCOUNTANT or ADMIN user. Multiple offices with different period dates.",
     "1. GET /v1/offices/periods/report/min → verify returns earliest report period\n2. GET /v1/offices/periods/report/max → verify returns latest report period\n3. GET /v1/offices/periods/approve/min → verify earliest approve period\n4. GET /v1/offices/periods/approve/max → verify latest approve period\n5. Cross-check with individual office periods",
     "Min returns earliest date across all salary offices. Max returns latest. Currently all 27 offices identical (REPORT=March, APPROVE=Feb), so min=max. Would diverge if offices advanced at different times.",
     "Medium", "Functional", "API-PERIOD-01", "OfficePeriodController",
     "Useful for accounting dashboard to show period range across offices."),

    ("TC-ACC-080", "Individual employee period: GET and extended period interaction",
     "ADMIN. Employee with and without extended periods.",
     "1. GET /v1/periods/report/employees/{login} for normal employee → verify matches office period\n2. Create extended period for employee\n3. GET again → verify extended period returned\n4. GET /v1/periods/report/employees → verify extended period list",
     "Normal employee returns office-level period. Extended period overrides for individual. Extended period list shows all employees with active extensions. Empty list when no extensions exist (currently empty on test envs).",
     "Medium", "Functional", "API-PERIOD-02", "EmployeeExtendedPeriodController",
     "employee_work_period table newly created but unpopulated on test environments."),

    ("TC-ACC-081", "Vacation days grouped by years: per-year breakdown correctness",
     "ACCOUNTANT.",
     "1. GET /v1/vacationdays/{login}/years\n2. Verify per-year breakdown\n3. Check for negative year values (corrections)\n4. Cross-check with total available days",
     "Returns array of {year, days} per employee. Negative values possible for correction years (e.g. {year:2025, days:-60}). Sum of all years should approximately equal total available days (accounting for in-progress vacations).",
     "Medium", "Functional", "API-DAYS-01", "EmployeeDaysController",
     "Negative year values from manual corrections. Cross-check with /v1/vacationdays."),

    ("TC-ACC-082", "Days summary timeline: totalAccrued, totalUsed, totalAdministrative",
     "ACCOUNTANT.",
     "1. GET /v1/timelines/days-summary/{login}\n2. Verify totalAccruedDays matches calculation formula\n3. Verify totalUsedDays accounts for all APPROVED+PAID regular vacations\n4. Verify totalAdministrativeDays accounts for unpaid vacations",
     "Summary provides: totalAccruedDays (formula-based, includes adjustments), totalUsedDays (regular vacation days consumed), totalAdministrativeDays (unpaid/admin days). Values consistent with per-year breakdown and available balance.",
     "Medium", "Functional", "API-DAYS-02", "VacationTimelineService",
     "High-level summary endpoint. Cross-reference with /years and /available."),
]))


# -- Risk Assessment -------------------------------------------------------
RISKS = [
    ("Period advance with unapproved reports", "Auto-reject may silently reject employee work",
     "High", "Critical", "Critical",
     "Test auto-reject thoroughly: TS-ACC-PeriodEffects TC-ACC-016/017"),
    ("VacationStatusUpdateJob orphan window", "Payment status updates permanently lost after 2 hours",
     "Medium", "Critical", "High",
     "BUG-PAY-1: TS-ACC-PayLifecycle TC-ACC-045. No recovery mechanism."),
    ("Missing first-day validation on approve period", "Non-standard period dates corrupt monthly processing",
     "High", "High", "Critical",
     "BUG-PERIOD-1: TS-ACC-PeriodEdge TC-ACC-004"),
    ("NPE on null period start", "Server crash on malformed input",
     "Medium", "High", "High",
     "BUG-PERIOD-2: TS-ACC-PeriodEdge TC-ACC-005"),
    ("Payment type misalignment allowed", "Incorrect accounting classification of vacation days",
     "Medium", "High", "High",
     "BUG-PAY-2: TS-ACC-PayValidation TC-ACC-029"),
    ("DB/API day type representation inconsistency", "Financial reports from DB queries show wrong day types",
     "High", "High", "Critical",
     "BUG-PAY-4: TS-ACC-PayValidation TC-ACC-038"),
    ("Stack trace information disclosure", "Internal architecture exposed to clients",
     "High", "Medium", "High",
     "Multiple endpoints: TS-ACC-PeriodEdge TC-ACC-006, TS-ACC-PayValidation TC-ACC-037"),
    ("API token auth gap for accounting endpoints", "Automated integrations cannot access accounting data",
     "Medium", "Medium", "Medium",
     "TS-ACC-APIErrors TC-ACC-074. AUTHENTICATED_USER not granted to API tokens."),
    ("Concurrent period modifications without locking", "Race condition on critical accounting operation",
     "Low", "High", "Medium",
     "TS-ACC-PeriodEdge TC-ACC-013. No explicit locking found."),
    ("No pagination on vacation days list", "Performance degradation as employee count grows",
     "Medium", "Medium", "Medium",
     "TS-ACC-DayCorrect TC-ACC-057. 1609 records per request."),
    ("Auto-payment cron orphan risk", "APPROVED vacations not auto-paid if timing edge case",
     "Low", "Medium", "Low",
     "TS-ACC-PayLifecycle TC-ACC-044. ShedLock prevents dual execution."),
    ("Audit trail gaps in payment timeline events", "Cannot reconstruct payment details from event log",
     "Medium", "Medium", "Medium",
     "TS-ACC-PayLifecycle TC-ACC-046. days_used=0, previous_status=NULL."),
    ("Sick leave concurrent accounting status change", "Race condition between two accountants",
     "Low", "Medium", "Low",
     "TS-ACC-SickLeaveAcct TC-ACC-071. No optimistic locking."),
]

# -- Feature Matrix --------------------------------------------------------
FEATURES = [
    # (feature, ts_tab, functional, negative, boundary, security, integration, bug, concurrency)
    ("Period Management — Edge Cases", "TS-ACC-PeriodEdge", 3, 2, 3, 1, 0, 4, 1),
    ("Period Change — Cross-Service Effects", "TS-ACC-PeriodEffects", 4, 1, 0, 0, 5, 0, 1),
    ("Payment Validation — Error Handling", "TS-ACC-PayValidation", 0, 3, 2, 1, 0, 5, 0),
    ("Payment Lifecycle — Auto-Payment", "TS-ACC-PayLifecycle", 5, 0, 1, 0, 0, 1, 1),
    ("Day Correction — Edge Cases", "TS-ACC-DayCorrect", 6, 0, 2, 0, 0, 0, 0),
    ("Notifications — Triggers & Templates", "TS-ACC-Notifications", 5, 0, 0, 1, 1, 0, 0),
    ("Sick Leave Accounting — Workflow Gaps", "TS-ACC-SickLeaveAcct", 4, 1, 0, 0, 0, 0, 1),
    ("API Errors — Auth & Disclosure", "TS-ACC-APIErrors", 4, 0, 0, 2, 0, 1, 0),
]

# ==========================================================================
# Build workbook
# ==========================================================================

wb = openpyxl.Workbook()
wb.remove(wb.active)

# -- Plan Overview ---------------------------------------------------------
ws = wb.create_sheet("Plan Overview")
ws.sheet_properties.tabColor = TAB_GREEN
set_col_widths(ws, [20, 90])

plan_rows = [
    ("Test Plan", "Accounting Supplements — Gap Coverage"),
    ("Date", str(date.today())),
    ("Phase", "B — Generation (Supplements to existing 127 Qase cases)"),
    ("Module", "Accounting: Period Management, Vacation Payment, Day Corrections, Notifications, Sick Leave Accounting"),
    ("", ""),
    ("Scope", "This workbook supplements 127 existing Qase test cases (Suite 207, sub-suites 209-234) that cover:\n"
              "• Salary page: search (2), filters (8), table display (6)\n"
              "• Period changes: basic rules (5), dates/sorting (4), individual periods (5)\n"
              "• Vacation payment: table/filters/sorting (26), individual pay popup (7), payment execution (5), batch (5)\n"
              "• Day correction: search (4), events + corrections (18)\n"
              "• Sick leave accounting: sort/filter/actions/alerts (27)\n\n"
              "This supplement focuses on GAPS not covered by Qase:\n"
              "• Period edge cases: validation bugs, boundary conditions, non-first-day bug, NPE on null\n"
              "• Period cross-service effects: auto-reject, vacation recalc, RabbitMQ events, extended periods\n"
              "• Payment validation: error handling, type misalignment bug, DB/API inconsistency\n"
              "• Payment lifecycle: partial payment, day return logic, auto-payment cron, orphan window\n"
              "• Day correction: negative balances, bulk recalc, FIFO redistribution, maternity edge case\n"
              "• Accounting notifications: manager notify, auto-reject emails, forgotten reports\n"
              "• Sick leave accounting: status workflow transitions, concurrent changes, overdue alerts\n"
              "• API-level: auth gaps, information disclosure, pagination inconsistency"),
    ("", ""),
    ("Environments", "Primary: timemachine (ttt-timemachine.noveogroup.com)\n"
                     "Secondary: qa-1 (ttt-qa-1.noveogroup.com)\n"
                     "Production baseline: stage (ttt-stage.noveogroup.com)"),
    ("", ""),
    ("Test Data Strategy", "• Period operations: use timemachine clock manipulation for date-dependent tests\n"
                           "• Payment tests: APPROVED vacations (query DB: SELECT * FROM vacation WHERE status='APPROVED')\n"
                           "• Day correction: employees with non-zero balances\n"
                           "• Sick leave accounting: mix of NEW/PENDING/PAID/REJECTED statuses\n"
                           "• Cross-service: set up REPORTED-state reports before period advance\n"
                           "• API tokens: use API_SECRET_TOKEN header with test API key"),
    ("", ""),
    ("Key Accounts", "• perekrest: CHIEF_ACCOUNTANT + ACCOUNTANT (all offices)\n"
                     "• accountant users: office-specific access\n"
                     "• admin: full access\n"
                     "• regular employees: for report/vacation targets"),
    ("", ""),
    ("Known Bugs Referenced", "• BUG-PERIOD-1: Missing first-day-of-month validation on approve period (HIGH)\n"
                              "• BUG-PERIOD-2: NPE on null start in PATCH body (HIGH)\n"
                              "• BUG-PERIOD-3: Stack trace leakage on invalid date format (MEDIUM)\n"
                              "• BUG-PERIOD-4: Permission inconsistency report vs approve min/max (MEDIUM)\n"
                              "• BUG-PAY-1: VacationStatusUpdateJob 2-hour orphan window (HIGH)\n"
                              "• BUG-PAY-2: Payment type misalignment allowed (MEDIUM)\n"
                              "• BUG-PAY-3: Payment dates start > end accepted (MEDIUM)\n"
                              "• BUG-PAY-4: DB/API day type representation inconsistency (MEDIUM)\n"
                              "• BUG-PAY-5: Available days accepts negative newDays (LOW)\n"
                              "• BUG-PAY-6: Stack trace leakage on invalid payment date (MEDIUM)\n"
                              "• NPE on status=ALL: VacationRepositoryCustomImpl:433 (HIGH)"),
    ("", ""),
    ("Test Suite Links", ""),
]

r = 1
for label, value in plan_rows:
    ws.cell(row=r, column=1, value=label).font = SUBTITLE_FONT if label else BODY_FONT
    ws.cell(row=r, column=2, value=value).font = BODY_FONT
    ws.cell(row=r, column=2).alignment = WRAP
    r += 1

suite_link_start = r
for suite_name, suite_title, cases in SUITES:
    cell = ws.cell(row=r, column=2,
                   value=f"{suite_name}: {suite_title} — {len(cases)} cases")
    cell.font = LINK_FONT
    cell.hyperlink = f"#'{suite_name}'!A1"
    r += 1

# -- Feature Matrix --------------------------------------------------------
ws2 = wb.create_sheet("Feature Matrix")
ws2.sheet_properties.tabColor = TAB_GREEN
set_col_widths(ws2, [35, 22, 12, 12, 12, 12, 12, 10, 14, 10])

feat_headers = ["Feature Area", "Test Suite Tab", "Functional", "Negative",
                "Boundary", "Security", "Integration", "Bug", "Concurrency", "Total"]
for c, h in enumerate(feat_headers, 1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(feat_headers))

for idx, (feat, tab, func, neg, bnd, sec, integ, bug, conc) in enumerate(FEATURES):
    r = idx + 2
    total = func + neg + bnd + sec + integ + bug + conc
    vals = [feat, tab, func, neg, bnd, sec, integ, bug, conc, total]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        if c == 2:
            cell.font = LINK_FONT
            cell.hyperlink = f"#'{tab}'!A1"
    style_data_row(ws2, r, len(feat_headers), idx)

# Totals row
total_r = len(FEATURES) + 2
ws2.cell(row=total_r, column=1, value="TOTAL").font = Font(name=ARIAL, bold=True, size=10)
for c in range(3, 11):
    total = sum(ws2.cell(row=r, column=c).value or 0 for r in range(2, total_r))
    ws2.cell(row=total_r, column=c, value=total).font = Font(name=ARIAL, bold=True, size=10)
for c in range(1, 11):
    ws2.cell(row=total_r, column=c).border = THIN_BORDER
    ws2.cell(row=total_r, column=c).fill = PatternFill("solid", fgColor="E2EFDA")

ws2.auto_filter.ref = f"A1:J{total_r}"

# -- Risk Assessment -------------------------------------------------------
ws3 = wb.create_sheet("Risk Assessment")
ws3.sheet_properties.tabColor = TAB_GREEN
set_col_widths(ws3, [40, 50, 12, 12, 12, 55])

risk_headers = ["Risk", "Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
for c, h in enumerate(risk_headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, len(risk_headers))

sev_fills = {"Critical": RISK_CRIT, "High": RISK_HIGH, "Medium": RISK_MED, "Low": RISK_LOW}

for idx, (risk, desc, likelihood, impact, severity, mitigation) in enumerate(RISKS):
    r = idx + 2
    for c, v in enumerate([risk, desc, likelihood, impact, severity, mitigation], 1):
        ws3.cell(row=r, column=c, value=v)
    style_data_row(ws3, r, len(risk_headers), idx)
    ws3.cell(row=r, column=5).fill = sev_fills.get(severity, ROW_ODD)

ws3.auto_filter.ref = f"A1:F{len(RISKS) + 1}"

# -- Test Suite Tabs -------------------------------------------------------
TC_HEADERS = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
              "Priority", "Type", "Requirement Ref", "Module / Component", "Notes"]
TC_WIDTHS = [14, 40, 30, 45, 45, 10, 12, 22, 30, 35]

for suite_name, suite_title, cases in SUITES:
    ws = wb.create_sheet(suite_name)
    ws.sheet_properties.tabColor = TAB_BLUE

    # Back link
    cell = ws.cell(row=1, column=1, value="← Back to Plan Overview")
    cell.font = BACK_LINK_FONT
    cell.hyperlink = "#'Plan Overview'!A1"

    # Title
    ws.cell(row=2, column=1, value=f"{suite_name}: {suite_title}").font = TITLE_FONT
    ws.cell(row=3, column=1, value=f"{len(cases)} test cases").font = BODY_FONT

    # Headers
    for c, h in enumerate(TC_HEADERS, 1):
        ws.cell(row=5, column=c, value=h)
    style_header_row(ws, 5, len(TC_HEADERS))

    # Cases
    for idx, case in enumerate(cases):
        r = 6 + idx
        for c, v in enumerate(case, 1):
            ws.cell(row=r, column=c, value=v)
        style_data_row(ws, r, len(TC_HEADERS), idx)

    set_col_widths(ws, TC_WIDTHS)
    ws.auto_filter.ref = f"A5:{get_column_letter(len(TC_HEADERS))}{5 + len(cases)}"

wb.save(OUTPUT)

total_cases = sum(len(cases) for _, _, cases in SUITES)
print(f"Generated {OUTPUT}")
print(f"Total: {len(SUITES)} test suites, {total_cases} test cases")
print(f"Tabs: Plan Overview, Feature Matrix, Risk Assessment + {len(SUITES)} TS- tabs")
for sn, st, cases in SUITES:
    print(f"  {sn}: {len(cases)} cases — {st}")
