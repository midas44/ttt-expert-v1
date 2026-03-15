#!/usr/bin/env python3
"""Generate unified sick-leave.xlsx workbook with test plan + all test suites."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

OUTPUT = "/home/v/Dev/ttt-expert-v1/expert-system/output/sick-leave/sick-leave.xlsx"

# ── Styles ──────────────────────────────────────────────────────────────────
ARIAL = "Arial"
HEADER_FONT = Font(name=ARIAL, bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
SUBTITLE_FONT = Font(name=ARIAL, bold=True, size=12)
BODY_FONT = Font(name=ARIAL, size=10)
LINK_FONT = Font(name=ARIAL, size=10, color="0563C1", underline="single")
BACK_LINK_FONT = Font(name=ARIAL, size=9, color="0563C1", underline="single")
ROW_EVEN = PatternFill("solid", fgColor="D6E4F0")
ROW_ODD = PatternFill("solid", fgColor="FFFFFF")
RISK_CRIT = PatternFill("solid", fgColor="FF6B6B")
RISK_HIGH = PatternFill("solid", fgColor="FFA07A")
RISK_MED = PatternFill("solid", fgColor="FFD700")
RISK_LOW = PatternFill("solid", fgColor="90EE90")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TAB_GREEN = "00B050"
TAB_BLUE = "4472C4"


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def style_data_row(ws, row, max_col, idx):
    fill = ROW_EVEN if idx % 2 == 0 else ROW_ODD
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Test Suite Definitions ──────────────────────────────────────────────────
# Each suite: (tab_name, title, cases_list)
# Each case: (id, title, preconditions, steps, expected, priority, type, req_ref, component, notes)

SUITES = []

# ── TS-SL-CRUD ──────────────────────────────────────────────────────────────
SUITES.append(("TS-SL-CRUD", "Sick Leave CRUD Lifecycle", [
    ("TC-SL-001", "Create sick leave with minimum required fields",
     "Logged in as employee. No active sick leaves overlapping target dates.",
     "1. Navigate to /sick-leave/my\n2. Click 'Add a sick note'\n3. Select Start date (today)\n4. Select End date (today + 3 days)\n5. Click Save",
     "Sick leave created with status OPEN/Started, accounting_status NEW. Calendar days auto-calculated. Appears in list.",
     "High", "Functional", "REQ-sick-leave §Create", "frontend-sick-leave-module, sick-leave-service", ""),
    ("TC-SL-002", "Create sick leave with all optional fields",
     "Logged in as employee.",
     "1. Click 'Add a sick note'\n2. Fill Start/End dates\n3. Enter Number field (e.g. 'BL-2026-001')\n4. Upload 1 PDF file\n5. Add 2 notifyAlso recipients\n6. Click Save",
     "Sick leave created with number, 1 file attached, 2 extra notification recipients stored. All fields visible in detail modal.",
     "High", "Functional", "REQ-sick-leave", "frontend-sick-leave-module", ""),
    ("TC-SL-003", "Calendar days auto-calculation on date selection",
     "Create modal open.",
     "1. Select Start date = 2026-03-16 (Monday)\n2. Select End date = 2026-03-22 (Sunday)\n3. Observe Calendar days field",
     "Calendar days shows 7 (all calendar days including weekends). Field is read-only, auto-updated.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "total_days = end - start + 1 (calendar)"),
    ("TC-SL-004", "Work days auto-calculation",
     "Sick leave created spanning Mon-Sun (7 calendar days).",
     "1. Via API GET /v1/sick-leaves/{id}\n2. Check work_days field",
     "work_days = 5 (Mon-Fri, excluding Sat/Sun per production calendar). Uses office's production calendar for calculation.",
     "Medium", "Functional", "", "sick-leave-service", "Work days visible only in accounting view column"),
    ("TC-SL-005", "Create single-day sick leave (start = end)",
     "Logged in as employee.",
     "1. Click 'Add a sick note'\n2. Select Start date\n3. Verify End date auto-fills with same date\n4. Save",
     "Sick leave created with total_days=1. Start date selection auto-fills end date.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Confirmed behavior: start auto-fills end"),
    ("TC-SL-006", "Create long-duration sick leave (>30 days)",
     "Logged in as employee.",
     "1. Create sick leave: Start = 2026-01-01, End = 2026-04-30 (120 days)\n2. Save",
     "Sick leave created successfully. No max duration validation. Calendar days = 120. DB shows max observed = 140-141 days.",
     "Low", "Boundary", "", "sick-leave-service", "No upper limit on duration"),
    ("TC-SL-007", "Create future sick leave — SCHEDULED status",
     "Logged in as employee.",
     "1. Create sick leave with Start date = today + 7 days\n2. Save\n3. Check State in list",
     "State shows 'Planned' (computed SCHEDULED from OPEN + future start_date). Stored status = OPEN.",
     "High", "Functional", "", "sick-leave-service", "Computed status, not stored"),
    ("TC-SL-008", "Edit sick leave — change dates",
     "Existing OPEN sick leave in list.",
     "1. Click detail icon on sick leave row\n2. Click Edit\n3. Change End date to 3 days later\n4. Save",
     "Dates updated. Calendar days recalculated. NOTIFY_SICKLEAVE_DATES_CHANGED email sent.",
     "High", "Functional", "", "frontend-sick-leave-module", "Two-step PATCH if files also changed"),
    ("TC-SL-009", "Edit sick leave — add number field",
     "Existing OPEN sick leave without number.",
     "1. Edit sick leave\n2. Enter Number = 'BL-123'\n3. Save",
     "Number field updated. NOTIFY_SICKLEAVE_NUMBER_CHANGED email sent.",
     "Medium", "Functional", "", "sick-leave-service", "Number optional until close"),
    ("TC-SL-010", "Edit sick leave — change number field",
     "Existing sick leave with number 'BL-123'.",
     "1. Edit sick leave\n2. Change Number to 'BL-456'\n3. Save",
     "Number updated. Change notification sent with new number.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-011", "Edit sick leave — add files (two-step update)",
     "Existing sick leave with no attachments.",
     "1. Edit sick leave\n2. Attach 2 PDF files via file uploader\n3. Save",
     "Files uploaded via POST /v1/files/upload first, then PATCH with filesIds. NOTIFY_SICKLEAVE_FILES_ADDED sent. Files visible in attachment panel.",
     "High", "Functional", "", "frontend-sick-leave-module", "Two-step: upload then patch. FilesAddedEvent on patch only."),
    ("TC-SL-012", "Edit sick leave — remove files",
     "Existing sick leave with 2 file attachments.",
     "1. Edit sick leave\n2. Remove 1 file (click X on file chip)\n3. Save",
     "File association removed (diff-and-sync). Remaining file still accessible. Removed file's DB record deleted.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-013", "Close (end) sick leave — with number",
     "Existing OPEN sick leave.",
     "1. Click end/close action on sick leave row\n2. Enter Number = 'BL-CLOSE-001' (required for close)\n3. Confirm",
     "Status changes to CLOSED. State shows 'Ended'. Number stored. NOTIFY_SICKLEAVE_CLOSED sent.",
     "High", "Functional", "REQ-sick-leave §Close", "sick-leave-service", "Number becomes required (*) during close flow"),
    ("TC-SL-014", "Close sick leave — without number (validation error)",
     "Existing OPEN sick leave.",
     "1. Click end/close action\n2. Leave Number field empty\n3. Try to confirm",
     "Validation error: Number field is required for closing. Close operation blocked.",
     "High", "Negative", "", "sick-leave-service", "Frontend Yup validation enforces required on close mode"),
    ("TC-SL-015", "Close sick leave — verify end dialog title bug",
     "Existing OPEN sick leave.",
     "1. Click end/close action\n2. Observe dialog title",
     "BUG: Dialog title says 'Delete the sick note?' instead of 'End the sick note?'. Functionally closes (not deletes).",
     "Medium", "Bug verification", "BUG-SL (S31)", "frontend-sick-leave-module", "Known bug from session 31 exploration"),
    ("TC-SL-016", "Reopen closed sick leave",
     "Existing CLOSED sick leave.",
     "1. Via API: PATCH /v1/sick-leaves/{id} with status=OPEN\n2. Verify state changes",
     "Status changes back to OPEN. State recalculated based on dates (Started/Overdue/Planned).",
     "Medium", "Functional", "", "sick-leave-service", "Direct status overwrite, no guard conditions"),
    ("TC-SL-017", "Delete sick leave — via Detail modal",
     "Existing non-PAID sick leave.",
     "1. Click detail icon on row\n2. Click Delete button in detail modal\n3. Confirm deletion",
     "Status set to DELETED. Row disappears from default view. NOTIFY_SICKLEAVE_DELETE sent. Soft-delete only.",
     "High", "Functional", "", "frontend-sick-leave-module", "Delete button only in Detail modal, not table row"),
    ("TC-SL-018", "Delete PAID sick leave — blocked",
     "Existing sick leave with accounting_status = PAID.",
     "1. Open detail modal\n2. Attempt to delete",
     "Delete button should be disabled/hidden OR backend returns error. Cannot delete PAID sick leave.",
     "High", "Negative", "", "sick-leave-service", "Guard: accounting_status != PAID"),
    ("TC-SL-019", "Verify soft-delete — DELETED status retained in DB",
     "Delete a sick leave from UI.",
     "1. Delete sick leave\n2. Query DB: SELECT status FROM sick_leave WHERE id = {id}",
     "status = 'DELETED' in DB. Record not physically removed. File associations and notify_also records remain (orphaned).",
     "Low", "Data integrity", "BUG-SL-5", "sick-leave-service", "Known debt: orphaned file/notify_also records"),
    ("TC-SL-020", "Create overlapping sick leave — force=false (conflict)",
     "Existing OPEN sick leave: 2026-03-10 to 2026-03-15.",
     "1. Create new sick leave: 2026-03-12 to 2026-03-18\n2. Backend receives force=false",
     "409 Conflict response. Overlap dialog shown in UI asking to force or cancel.",
     "High", "Negative", "", "sick-leave-service", "Frontend handles 409 with force dialog"),
    ("TC-SL-021", "Create overlapping sick leave — force=true (allowed)",
     "Existing OPEN sick leave overlapping target dates. Overlap dialog shown.",
     "1. Click 'Force create' in overlap dialog\n2. Backend receives force=true",
     "Sick leave created despite overlap. Both active sick leaves coexist.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-022", "Create sick leave overlapping active vacation — crossing notification",
     "Employee has active vacation 2026-03-10 to 2026-03-20.",
     "1. Create sick leave: 2026-03-12 to 2026-03-18\n2. Force=true to bypass overlap check",
     "Sick leave created. NOTIFY_EMPLOYEE_SICKLEAVE_OVERLAPS_VACATION email sent to relevant parties.",
     "High", "Functional", "REQ-sick-leave §VacationCrossing", "sick-leave-service", "Vacation crossing check separate from overlap"),
]))

# ── TS-SL-DualStatus ───────────────────────────────────────────────────────
SUITES.append(("TS-SL-DualStatus", "Sick Leave Dual Status Model", [
    ("TC-SL-023", "Initial status on creation — OPEN + NEW",
     "Create a new sick leave.",
     "1. Create sick leave via UI\n2. Query DB: SELECT status, accounting_status FROM sick_leave WHERE id = {new_id}",
     "status = 'OPEN', accounting_status = 'NEW'. Employee view shows State = Started (if current dates).",
     "High", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-024", "Computed status SCHEDULED — future start date",
     "Existing OPEN sick leave with start_date > today.",
     "1. GET /v1/sick-leaves?statuses=SCHEDULED\n2. Verify returned records",
     "Records with OPEN + future start_date returned with status = SCHEDULED. UI shows State = 'Planned'.",
     "High", "Functional", "", "sick-leave-service", "Computed at query time via CASE WHEN, never stored"),
    ("TC-SL-025", "Computed status OVERDUE — past end date",
     "Existing OPEN sick leave with end_date < today.",
     "1. GET /v1/sick-leaves?statuses=OVERDUE\n2. Check UI manager view",
     "Records with OPEN + past end_date returned with status = OVERDUE. UI shows State = 'Overdue'. Green checkmark action in manager view.",
     "High", "Functional", "", "sick-leave-service", "OverdueSickLeaveCommand per-request check"),
    ("TC-SL-026", "Read/write asymmetry — PATCH rejects SCHEDULED/OVERDUE",
     "Existing OPEN sick leave with computed status SCHEDULED.",
     "1. PATCH /v1/sick-leaves/{id} with status=SCHEDULED\n2. Observe response",
     "Error response: SCHEDULED is not a valid write value. Only OPEN and CLOSED accepted for PATCH. Undocumented behavior.",
     "Medium", "Negative", "", "sick-leave-service", "Design debt: computed statuses not writable"),
    ("TC-SL-027", "Accounting: NEW → PROCESSING transition",
     "Sick leave with accounting_status = NEW. Logged in as accountant.",
     "1. Open accounting page\n2. Find sick leave record\n3. Change Status dropdown from New to Pending/Processing",
     "accounting_status updated to PROCESSING. accountant FK set to current user. Main status remains OPEN.",
     "High", "Functional", "", "sick-leave-service", "Pending in UI = PROCESSING in DB"),
    ("TC-SL-028", "Accounting: PROCESSING → PAID — auto-closes main status",
     "Sick leave: status=OPEN, accounting_status=PROCESSING.",
     "1. Change Status dropdown to Paid\n2. Verify both statuses",
     "accounting_status = PAID. Main status auto-changed to CLOSED. State shows 'Ended'.",
     "Critical", "Functional", "", "sick-leave-service", "Status coupling: PAID → auto-close"),
    ("TC-SL-029", "Accounting: PROCESSING → REJECTED — sets main REJECTED",
     "Sick leave: status=OPEN, accounting_status=PROCESSING.",
     "1. Change Status dropdown to Rejected\n2. Verify both statuses",
     "accounting_status = REJECTED. Main status auto-changed to REJECTED. NOTIFY_SICKLEAVE_REJECTED sent.",
     "Critical", "Functional", "", "sick-leave-service", "Status coupling: REJECTED → main REJECTED"),
    ("TC-SL-030", "Accounting: PROCESSING → OPEN reopens closed sick leave",
     "Sick leave: status=CLOSED, accounting_status=NEW.",
     "1. Change accounting status to Processing\n2. Verify main status",
     "accounting_status = PROCESSING. Main status changes from CLOSED to OPEN (reopened). State recalculated.",
     "High", "Functional", "", "sick-leave-service", "Coupling: PROCESSING → OPEN reopens"),
    ("TC-SL-031", "Unrestricted transition: PAID → NEW (no guardrails)",
     "Sick leave: accounting_status = PAID (main status = CLOSED).",
     "1. Change Status dropdown from Paid to New\n2. Verify result",
     "BUG: Transition succeeds. accounting_status = NEW. Main status may remain CLOSED (inconsistent state). No guardrails prevent backward transition.",
     "High", "Bug verification", "BUG-SL-3", "sick-leave-service", "Any-to-any accounting transitions allowed"),
    ("TC-SL-032", "Unrestricted transition: REJECTED → PAID",
     "Sick leave: accounting_status = REJECTED.",
     "1. Change Status dropdown from Rejected to Paid\n2. Verify result",
     "BUG: Transition succeeds. REJECTED → PAID. Main status auto-set to CLOSED. No intermediate PROCESSING required.",
     "Medium", "Bug verification", "BUG-SL-3", "sick-leave-service", ""),
    ("TC-SL-033", "Status combination matrix — valid data patterns",
     "Access to timemachine database.",
     "1. Query: SELECT status, accounting_status, COUNT(*) FROM sick_leave GROUP BY 1,2\n2. Compare with expected combinations",
     "Valid combinations: OPEN/NEW, CLOSED/NEW, CLOSED/PAID, REJECTED/REJECTED, DELETED/NEW. 62% are CLOSED/NEW backlog.",
     "Low", "Data integrity", "", "sick-leave-service", "348 total records in timemachine"),
    ("TC-SL-034", "Employee view — State column, no Status column",
     "Logged in as employee with sick leaves.",
     "1. Navigate to /sick-leave/my\n2. Observe table columns",
     "Columns: Sick leave dates, Calendar days, Number, Accountant, State, Actions. NO Status (accounting) column visible.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Employees cannot see accounting status"),
    ("TC-SL-035", "Manager view — State + Status as plain text",
     "Logged in as PM/DM. Navigate to /vacation/sick-leaves-of-employees.",
     "1. Observe table columns\n2. Try to interact with Status column",
     "Both State and Status columns present. Status shown as plain text (not editable dropdown). Managers see but cannot change accounting status.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-036", "Accounting view — Status as inline dropdown",
     "Logged in as accountant. Navigate to /accounting/sick-leaves.",
     "1. Observe Status column\n2. Click on Status cell value",
     "Status column shows inline dropdown for direct status change. No confirmation modal — changes apply immediately on selection.",
     "High", "Functional", "", "frontend-sick-leave-module", "SickLeaveAccountingStatusCell component"),
    ("TC-SL-037", "Accounting status change persists accountant FK",
     "Logged in as accountant 'perekrest'. Change any accounting status.",
     "1. Change status from New to Processing\n2. Check Accountant column\n3. Query DB: SELECT accountant FROM sick_leave WHERE id = {id}",
     "Accountant column shows current user. DB accountant FK points to current user's employee ID.",
     "Medium", "Functional", "", "sick-leave-service", "Auto-assigned on any accounting status change"),
]))

# ── TS-SL-Accounting ───────────────────────────────────────────────────────
SUITES.append(("TS-SL-Accounting", "Sick Leave Accounting Workflow", [
    ("TC-SL-038", "Accounting page access — accountant role required",
     "1. User with accountant role\n2. User without accountant role",
     "1. Login as accountant → navigate to /accounting/sick-leaves\n2. Login as regular employee → attempt same URL",
     "Accountant: page loads with table. Non-accountant: redirected or access denied.",
     "High", "Functional", "", "frontend-sick-leave-module", "Route requires VACATIONS:SICK_LEAVE_ACCOUNTING_VIEW"),
    ("TC-SL-039", "Accounting page — 10 columns displayed",
     "Logged in as accountant.",
     "1. Navigate to /accounting/sick-leaves\n2. Verify all columns",
     "10 columns: Employee, Sick leave dates, Days, Work days, Sick note, Accountant, Salary office, State, Status, Actions.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Richer than employee/manager views"),
    ("TC-SL-040", "Salary office filter — 27 offices",
     "Logged in as accountant.",
     "1. Click Salary office filter\n2. Count available options\n3. Select specific office\n4. Verify table filters",
     "Filter shows available salary offices (up to 27 on timemachine). Table updates to show only selected office's sick leaves.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Default scoped to accountant's salary office"),
    ("TC-SL-041", "State filter — 7 values",
     "Logged in as accountant.",
     "1. Click State filter dropdown\n2. Verify options: All, Started, Ended, Planned, Overdue, Rejected, Deleted",
     "7 filter values available. Each filters correctly based on computed main status.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-042", "Status filter — 4 values + 'Rejected Rejected' bug",
     "Logged in as accountant.",
     "1. Click Status filter dropdown\n2. Observe options",
     "BUG: Filter shows 'Rejected Rejected' (duplicated label) for the REJECTED accounting status. Other values: New, Pending, Paid.",
     "Low", "Bug verification", "BUG-SL-4", "frontend-sick-leave-module", "Known bug from session 11"),
    ("TC-SL-043", "Change accounting status — New → Processing via dropdown",
     "Sick leave with accounting_status = NEW in accounting view.",
     "1. Click Status dropdown on row\n2. Select 'Pending' (= PROCESSING)\n3. Verify status change",
     "Status changes immediately to Processing/Pending. No confirmation dialog. accountant FK set to current user.",
     "High", "Functional", "", "frontend-sick-leave-module", "Direct PATCH on dropdown change"),
    ("TC-SL-044", "Change accounting status — Processing → Paid with auto-close",
     "Sick leave with accounting_status = PROCESSING.",
     "1. Change dropdown to 'Paid'\n2. Verify both State and Status",
     "Status = Paid. State changes to 'Ended' (main status auto-closed to CLOSED). Accountant column updated.",
     "Critical", "Functional", "", "sick-leave-service", "Critical coupling behavior"),
    ("TC-SL-045", "Add comment via inline tooltip",
     "Sick leave record in accounting view.",
     "1. Click comment/speech-bubble action button\n2. Enter comment text in tooltip textarea\n3. Save",
     "Comment saved. Tooltip shows saved comment on hover. DB accountant_comment field updated.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Inline Tooltip, not modal"),
    ("TC-SL-046", "Edit comment — update existing",
     "Sick leave with existing accountant comment.",
     "1. Click comment action\n2. Modify text\n3. Save",
     "Comment updated. Previous text replaced entirely.",
     "Low", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-047", "View sick note details — edit dialog",
     "Sick leave record in accounting view.",
     "1. Click pencil (edit) action button\n2. Observe edit dialog fields",
     "Dialog shows: Employee (read-only), Start date, End date, Calendar days (auto-calc), Sick note number. No accounting status field.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Accounting status managed only via inline dropdown"),
    ("TC-SL-048", "Edit sick leave dates from accounting view",
     "Non-PAID sick leave in accounting view.",
     "1. Click edit action\n2. Change End date\n3. Save",
     "Dates updated. Calendar days recalculated. NOTIFY_SICKLEAVE_DATES_CHANGED sent.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-049", "Edit PAID sick leave — admin only",
     "Sick leave with accounting_status = PAID. Test with accountant and admin roles.",
     "1. As accountant: try to edit PAID sick leave\n2. As admin: try to edit PAID sick leave",
     "Accountant: edit blocked (button disabled/hidden). Admin: edit allowed (admin override for PAID sick leaves).",
     "High", "Functional", "", "sick-leave-service", "PM for own reports; admin/accountant for PAID"),
    ("TC-SL-050", "Overdue records — green checkmark action",
     "OPEN sick leave with end_date in the past (OVERDUE state).",
     "1. Navigate to accounting page\n2. Filter State = Overdue\n3. Check action column",
     "Green checkmark action button visible. OverdueSickLeaveCommand per-request check surfaces warning.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-051", "Accounting page — no create button",
     "Logged in as accountant on /accounting/sick-leaves.",
     "1. Check for 'Add a sick note' button",
     "No create button present. Accounting view is read/process-only, not for creating new sick leaves.",
     "Low", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-052", "Column sorting — all sortable columns",
     "Accounting page with multiple records.",
     "1. Click 'Sick leave dates' column header\n2. Click 'Days' column header\n3. Click 'Employee' column header",
     "Table sorts by clicked column. Sort direction toggles asc/desc. Sort encoding uses +/- prefix.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-053", "Accounting backlog visibility — CLOSED/NEW records",
     "Timemachine data with 215 CLOSED/NEW records.",
     "1. Filter State = Ended, Status = New\n2. Observe count",
     "Shows 215 records (62% backlog) — sick leaves ended but never processed by accounting. Visible for processing.",
     "Low", "Data integrity", "", "sick-leave-service", "Data quality observation"),
    ("TC-SL-054", "NoveoAI widget overlap with Status/Actions columns",
     "Accounting page on narrow viewport.",
     "1. Observe right side of table\n2. Check if floating NoveoAI widget overlaps",
     "BUG: NoveoAI floating widget overlaps Status and Actions columns, blocking interaction.",
     "Low", "Bug verification", "BUG-SL-8", "frontend-sick-leave-module", "Known UI overlap issue"),
    ("TC-SL-055", "View attachments from accounting — sick note files",
     "Sick leave with file attachments in accounting view.",
     "1. Click clipboard/attachment action button\n2. Verify files displayed",
     "Attachment panel opens showing uploaded files with view/download links. Files accessible to accountant role.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "View sick note action button"),
]))

# ── TS-SL-Files ────────────────────────────────────────────────────────────
SUITES.append(("TS-SL-Files", "Sick Leave File Handling", [
    ("TC-SL-056", "Upload single PDF file during creation",
     "Create sick leave modal open.",
     "1. Click file upload area\n2. Select 1 PDF file (<5MB)\n3. Verify file chip appears\n4. Save sick leave",
     "File uploaded via POST /v1/files/upload (multipart). UUID-renamed on disk. Junction record in sick_leave_file. File visible in attachment panel.",
     "High", "Functional", "", "frontend-sick-leave-module", "react-dropzone with hidden input[type=file]"),
    ("TC-SL-057", "Upload PNG and JPEG image files",
     "Create sick leave modal open.",
     "1. Upload 1 PNG file\n2. Upload 1 JPEG file\n3. Save",
     "Both files accepted. Frontend MIME check passes for PNG/JPEG. Files stored with UUID names.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Accepted types: PDF, PNG, JPEG"),
    ("TC-SL-058", "Upload maximum 5 files",
     "Create sick leave modal open.",
     "1. Upload 5 files (mix of PDF/PNG/JPEG)\n2. Save",
     "All 5 files accepted. sick_leave_file junction has 5 records. @Size(max=5) on filesIds.",
     "High", "Boundary", "", "sick-leave-service", "Max 5 files per sick leave"),
    ("TC-SL-059", "Upload 6th file — rejected",
     "Sick leave with 5 files attached.",
     "1. Edit sick leave\n2. Try to add 6th file",
     "Frontend blocks 6th file addition. If bypassed, backend rejects with @Size(max=5) validation error.",
     "High", "Negative", "", "sick-leave-service", "Boundary: max files enforcement"),
    ("TC-SL-060", "Upload file exceeding 5MB — rejected",
     "Create sick leave modal open.",
     "1. Try to upload 6MB file",
     "Frontend client-side check rejects >5MB file. If bypassed, backend @MultipartSizeValid rejects.",
     "High", "Negative", "", "sick-leave-service", "Spring multipart limits are -1 (unlimited), relies on custom annotation"),
    ("TC-SL-061", "Add file during edit — two-step PATCH",
     "Existing sick leave with no attachments.",
     "1. Edit sick leave\n2. Attach 1 PDF\n3. Save",
     "Step 1: POST /v1/files/upload → UUID. Step 2: PATCH with filesIds. NOTIFY_SICKLEAVE_FILES_ADDED sent (only on patch, not create).",
     "High", "Functional", "", "frontend-sick-leave-module", "FilesAddedEvent triggered on patch only"),
    ("TC-SL-062", "Remove file during edit",
     "Existing sick leave with 2 file attachments.",
     "1. Edit sick leave\n2. Click X on file chip to remove 1 file\n3. Save",
     "Diff-and-sync: removed file's junction record deleted. Remaining file intact.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-063", "View file attachments — inline panel",
     "Sick leave with attachments in list.",
     "1. Click attachment icon (data-testid='sickleave-action-attachments')\n2. Observe inline panel",
     "Panel expands showing file thumbnails/links. 'View sick note' link opens file. File names shown (original or UUID-based).",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-064", "Download file attachment",
     "Sick leave with file attachment visible in panel.",
     "1. Click file link/download button\n2. Via API: GET /v1/files/{id}/download",
     "File downloads successfully. Content matches uploaded file.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-065", "File cascade on sick leave delete",
     "Sick leave with 2 file attachments.",
     "1. Delete sick leave from Detail modal\n2. Check DB: SELECT * FROM sick_leave_file WHERE sick_leave_id = {id}",
     "BUG: Junction records remain orphaned. Sick leave soft-deleted (DELETED status) but file associations NOT cleaned up.",
     "Medium", "Bug verification", "BUG-SL-5", "sick-leave-service", "Known debt: orphaned file records"),
    ("TC-SL-066", "Backend MIME bypass — upload non-allowed file type",
     "Disable frontend validation (DevTools or direct API call).",
     "1. Upload .exe or .html file directly to POST /v1/files/upload\n2. Include in sick leave PATCH",
     "SECURITY BUG: Backend has NO MIME type validation. Only frontend checks file types. Any file content accepted if size passes.",
     "High", "Security", "BUG (S15)", "sick-leave-service", "No backend MIME validation — potential malicious upload"),
    ("TC-SL-067", "File upload via API token — 403 expected",
     "API token with 21 permissions (no AUTHENTICATED_USER).",
     "1. POST /v1/files/upload with API_SECRET_TOKEN header\n2. Observe response",
     "403 Forbidden. File upload requires AUTHENTICATED_USER (JWT). API tokens insufficient. Full Java stack trace leaked in error response.",
     "Medium", "Security", "BUG (S15)", "sick-leave-service", "Stack trace leakage on 403"),
]))

# ── TS-SL-Notifications ───────────────────────────────────────────────────
SUITES.append(("TS-SL-Notifications", "Sick Leave Notifications", [
    ("TC-SL-068", "Create notification — employee creates own",
     "Employee creates own sick leave.",
     "1. Create sick leave as employee\n2. Check email notifications sent",
     "NOTIFY_SICKLEAVE_OPEN template sent. Recipients: employee's manager + optional approvers + notifyAlso + per-office receivers.",
     "High", "Functional", "REQ-sick-leave §Notify", "sick-leave-service", "Async after commit"),
    ("TC-SL-069", "Create notification — manager creates for subordinate",
     "Manager creates sick leave for subordinate employee.",
     "1. Manager navigates to /vacation/sick-leaves-of-employees\n2. Creates sick leave for subordinate\n3. Check notifications",
     "NOTIFY_SICKLEAVE_OPEN_BY_SUPERVISOR template sent (different template from employee-created).",
     "High", "Functional", "", "sick-leave-service", "Editor type determines template variant"),
    ("TC-SL-070", "Close notification — NOTIFY_SICKLEAVE_CLOSED",
     "Close an existing OPEN sick leave.",
     "1. Close sick leave with number\n2. Check email",
     "NOTIFY_SICKLEAVE_CLOSED template sent. Includes sick leave details and closing number.",
     "Medium", "Functional", "", "sick-leave-service", "Chain-of-responsibility dispatch"),
    ("TC-SL-071", "Dates changed notification",
     "Edit sick leave dates.",
     "1. Edit sick leave, change End date\n2. Check email",
     "NOTIFY_SICKLEAVE_DATES_CHANGED template sent. Includes old and new date values.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-072", "Number changed notification",
     "Edit sick leave number.",
     "1. Edit sick leave, change Number field\n2. Check email",
     "NOTIFY_SICKLEAVE_NUMBER_CHANGED template sent.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-073", "Rejected notification",
     "Accounting sets status to REJECTED.",
     "1. Change accounting status to Rejected\n2. Check email",
     "NOTIFY_SICKLEAVE_REJECTED template sent. Main status also changes to REJECTED.",
     "High", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-074", "Delete notification — employee deletes own",
     "Employee deletes own sick leave.",
     "1. Delete sick leave from Detail modal as employee\n2. Check email",
     "NOTIFY_SICKLEAVE_DELETE template sent.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-075", "Delete notification — manager deletes",
     "Manager deletes subordinate's sick leave.",
     "1. Manager deletes sick leave\n2. Check email",
     "NOTIFY_SICKLEAVE_DELETE_BY_SUPERVISOR template sent (different from employee-deleted).",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-076", "Files added notification",
     "Add files to existing sick leave via edit.",
     "1. Edit sick leave, attach new files\n2. Save\n3. Check email",
     "NOTIFY_SICKLEAVE_FILES_ADDED template sent. Only triggered on PATCH (not on initial create).",
     "Medium", "Functional", "", "sick-leave-service", "FilesAddedEvent on patch only, not create"),
    ("TC-SL-077", "Vacation overlap notification",
     "Create sick leave overlapping active vacation.",
     "1. Create sick leave overlapping employee's active vacation (force=true)\n2. Check email",
     "NOTIFY_EMPLOYEE_SICKLEAVE_OVERLAPS_VACATION template sent to relevant parties.",
     "High", "Functional", "", "sick-leave-service", "Separate from overlap check"),
    ("TC-SL-078", "Notification recipients — full list verification",
     "Employee with manager, notifyAlso, per-office receivers configured.",
     "1. Create sick leave with 2 notifyAlso recipients\n2. Verify all email recipients",
     "Email sent to: employee's manager + optional approvers + 2 notifyAlso + per-office notification receivers. All receive same template.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-079", "notifyAlso multiselect — add/view recipients",
     "Create modal open.",
     "1. Click notifyAlso multiselect\n2. Search for 2 employees\n3. Add them\n4. Save\n5. Verify in detail modal",
     "notifyAlso field stores selected logins. Visible in detail modal as 'Notify also' field.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Async employee search in multiselect"),
    ("TC-SL-080", "Editor type detection bug — reference equality",
     "Create sick leave. Check notification logs.",
     "1. Create sick leave as employee\n2. Inspect notification template used (via email test API or DB)",
     "BUG: getEditorType() uses `==` (reference equality) instead of .equals() on BO instances. Always falls to default editor type. Notification template may be incorrect variant.",
     "Medium", "Bug verification", "BUG-SL-2", "sick-leave-service", "employee == currentEmployee always false"),
    ("TC-SL-081", "Chain-of-responsibility dispatch — multi-field edit",
     "Edit sick leave changing both dates AND number.",
     "1. Edit sick leave: change end date + change number\n2. Check emails sent",
     "Chain dispatches DATES_CHANGED and NUMBER_CHANGED. Verify correct templates selected based on which fields changed.",
     "Low", "Functional", "", "sick-leave-service", ""),
]))

# ── TS-SL-Validation ──────────────────────────────────────────────────────
SUITES.append(("TS-SL-Validation", "Sick Leave Validation Rules", [
    ("TC-SL-082", "Start date required — empty field validation",
     "Create modal open.",
     "1. Leave Start date empty\n2. Fill End date\n3. Try to save",
     "Validation error: Start date is required. Save blocked by Yup schema (create mode).",
     "High", "Negative", "", "frontend-sick-leave-module", "Yup validation in create mode"),
    ("TC-SL-083", "End date required — empty field validation",
     "Create modal open.",
     "1. Fill Start date\n2. Leave End date empty\n3. Try to save",
     "Validation error: End date is required. Save blocked.",
     "High", "Negative", "", "frontend-sick-leave-module", ""),
    ("TC-SL-084", "Start date > end date — auto-adjustment",
     "Create modal with Start date already set.",
     "1. Select End date earlier than Start date\n2. Observe Start date field",
     "Start date auto-adjusts to match End date. No validation error — UI auto-corrects.",
     "High", "Functional", "", "frontend-sick-leave-module", "Confirmed behavior: end < start → start = end"),
    ("TC-SL-085", "Start date auto-fills end date on selection",
     "Create modal open, no dates selected.",
     "1. Select Start date = 2026-03-16\n2. Check End date field",
     "End date auto-fills with same value (2026-03-16). Calendar days = 1.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Convenience auto-fill"),
    ("TC-SL-086", "Calendar days recalculates on date change",
     "Create modal with dates set.",
     "1. Set Start = 2026-03-01, End = 2026-03-10 (10 days)\n2. Change End to 2026-03-15\n3. Observe Calendar days",
     "Calendar days updates live: 10 → 15. Read-only field, auto-calculated as endDate - startDate + 1.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-087", "Work days excludes weekends per production calendar",
     "Create sick leave spanning 2 weeks (14 calendar days, includes 4 weekend days).",
     "1. Create sick leave: Mon 2026-03-02 to Sun 2026-03-15\n2. Check work_days via API",
     "work_days = 10 (14 calendar - 4 weekend days). Uses office's production calendar, not just Sat/Sun.",
     "Medium", "Functional", "", "sick-leave-service", "Work days visible in accounting view"),
    ("TC-SL-088", "Number field — max 40 characters",
     "Create/edit modal with Number field.",
     "1. Enter 41-character string in Number field\n2. Save",
     "Backend rejects: number exceeds max length 40. Frontend may truncate or show validation error.",
     "Medium", "Boundary", "", "sick-leave-service", "VARCHAR(40) in DB, bean validation"),
    ("TC-SL-089", "Number field — whitespace trimming",
     "Edit modal with Number field.",
     "1. Enter '  BL-123  ' (leading/trailing spaces)\n2. Save\n3. Verify stored value",
     "Number stored as 'BL-123' (trimmed). Backend trims whitespace.",
     "Low", "Functional", "", "sick-leave-service", "Confirmed in investigation"),
    ("TC-SL-090", "Number required on close — not on create",
     "Create modal and Close dialog.",
     "1. Create sick leave WITHOUT number → succeeds\n2. Try to close same sick leave WITHOUT number",
     "Create: number optional (no asterisk). Close: number required (asterisk shown). Yup switches validation mode.",
     "High", "Functional", "", "frontend-sick-leave-module", "Mode-dependent Yup schema: create vs close"),
    ("TC-SL-091", "Overlap detection — active sick leaves only",
     "1 OPEN + 1 DELETED sick leave covering same dates.",
     "1. Create new sick leave overlapping both\n2. Observe overlap check",
     "Overlap check only considers active sick leaves (OPEN, CLOSED). DELETED and REJECTED excluded from overlap detection.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-092", "Manager overlap — client-side 100-record cap",
     "Manager creating sick leave for employee with >100 existing sick leaves.",
     "1. Manager creates sick leave\n2. Observe client-side overlap check",
     "Client-side overlap check pre-fetches max 100 records. If overlap is with record #101+, it is missed. force=true hardcoded bypasses backend check too.",
     "Medium", "Bug verification", "", "frontend-sick-leave-module", "Overlap cap + force=true = no validation for managers"),
    ("TC-SL-093", "Overlap — server-side 409 response handling",
     "Employee creates overlapping sick leave (force=false).",
     "1. Create sick leave overlapping existing active one\n2. Observe 409 response handling",
     "Frontend shows crossing/overlap dialog from 409 response. Options: force create or cancel.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-094", "Vacation crossing — force=false triggers dialog",
     "Employee with active vacation. Creating sick leave overlapping it.",
     "1. Create sick leave overlapping active vacation\n2. Backend force=false → 409 crossing check",
     "Vacation crossing dialog shown. Employee can choose to force-create or cancel.",
     "High", "Functional", "", "sick-leave-service", "Separate from sick leave overlap"),
    ("TC-SL-095", "Date picker — readonly inputs, calendar widget only",
     "Create modal date pickers.",
     "1. Try to type in date input field\n2. Click date picker calendar",
     "Input fields are readonly — must use calendar widget (click cells). No keyboard date entry possible.",
     "Low", "Usability", "", "frontend-sick-leave-module", "Verified via Playwright: <input readonly>"),
    ("TC-SL-096", "Vacation overlap dialog re-triggers on every save",
     "Sick leave overlapping vacation. Overlap dialog already dismissed.",
     "1. Create sick leave overlapping vacation → dismiss overlap dialog\n2. Edit same sick leave (no date change)\n3. Save",
     "BUG: Overlap dialog re-triggers on every save (create AND edit), not just when dates actually overlap. Nuisance dialog.",
     "Low", "Bug verification", "BUG (S31)", "frontend-sick-leave-module", ""),
]))

# ── TS-SL-Permissions ─────────────────────────────────────────────────────
SUITES.append(("TS-SL-Permissions", "Sick Leave Permissions & Security", [
    ("TC-SL-097", "Employee — view own sick leaves",
     "Logged in as regular employee.",
     "1. Navigate to /sick-leave/my\n2. Verify table shows own sick leaves",
     "Employee sees own sick leaves only. Table populated with personal records.",
     "High", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-098", "Employee — create own sick leave",
     "Logged in as regular employee.",
     "1. Click 'Add a sick note'\n2. Create sick leave (dates auto-scoped to self)",
     "Employee can create sick leave for themselves. No employee selector — implicitly own login.",
     "High", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-099", "Employee — edit own OPEN sick leave",
     "Employee with OPEN sick leave.",
     "1. Edit own sick leave\n2. Change dates\n3. Save",
     "Edit succeeds. Employee can modify own OPEN sick leaves.",
     "High", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-100", "Employee — cannot edit PAID sick leave",
     "Employee with PAID sick leave.",
     "1. Try to edit PAID sick leave",
     "Edit button disabled/hidden. Only admin can edit PAID sick leaves.",
     "Medium", "Negative", "", "sick-leave-service", ""),
    ("TC-SL-101", "PM — view subordinates' sick leaves",
     "Logged in as PM with subordinates.",
     "1. Navigate to /vacation/sick-leaves-of-employees\n2. Check My department / My projects tabs",
     "PM sees sick leaves of employees in their department/projects. Two tab views available.",
     "High", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-102", "PM — create sick leave for subordinate",
     "Logged in as PM.",
     "1. Click 'Add a sick note' on manager page\n2. Search for subordinate employee\n3. Fill dates\n4. Save",
     "PM can create sick leave for any searched employee. Employee async-search with no subordination filter.",
     "High", "Functional", "", "frontend-sick-leave-module", "No file upload in manager create"),
    ("TC-SL-103", "No creation permission check — any user for any employee (BUG-SL-1)",
     "API access (curl/Postman) as any authenticated user.",
     "1. POST /v1/sick-leaves with login of unrelated employee\n2. Observe response",
     "SECURITY BUG: Sick leave created for any employee regardless of relationship. No authorization check beyond AUTHENTICATED_USER.",
     "Critical", "Security", "BUG-SL-1", "sick-leave-service", "Any authenticated user can create for any employee"),
    ("TC-SL-104", "Accountant — access accounting page",
     "Logged in as user with ROLE_ACCOUNTANT.",
     "1. Navigate to /accounting/sick-leaves",
     "Page loads. 10-column table with inline status dropdown. Salary office filter scoped to accountant's office.",
     "High", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-105", "Accountant — change accounting status",
     "Logged in as accountant with sick leave records visible.",
     "1. Change Status dropdown on any row\n2. Verify PATCH succeeds",
     "Status updated. Only accountant/chief_accountant/admin can change accounting status.",
     "High", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-106", "Admin — edit PAID sick leave",
     "Logged in as admin. Sick leave with accounting_status = PAID.",
     "1. Edit PAID sick leave from accounting view\n2. Change dates\n3. Save",
     "Admin can edit PAID sick leaves (override). Regular accountant/PM cannot.",
     "High", "Functional", "", "sick-leave-service", "Admin privilege for PAID records"),
    ("TC-SL-107", "Route /sick-leave/my — no router-level permission check (BUG-SL-6)",
     "User with no SICK_LEAVE:VIEW permission.",
     "1. Navigate directly to /sick-leave/my",
     "BUG: Route has TODO comment for permission check but none implemented. Any authenticated user can access the page.",
     "Medium", "Bug verification", "BUG-SL-6", "frontend-sick-leave-module", "TODO in code, no guard"),
    ("TC-SL-108", "API token — CRUD endpoints return 403",
     "API token with 21 permissions (standard set).",
     "1. GET /v1/sick-leaves/{id} with API_SECRET_TOKEN\n2. POST /v1/sick-leaves with API_SECRET_TOKEN\n3. PATCH /v1/sick-leaves/{id}\n4. DELETE /v1/sick-leaves/{id}",
     "All return 403 Forbidden. Sick leave CRUD requires AUTHENTICATED_USER (JWT only). API tokens have no alternative permission path. Only search/count work.",
     "High", "Security", "", "sick-leave-service", "Inconsistent with vacation module pattern"),
    ("TC-SL-109", "Department manager — My department tab",
     "Logged in as department manager.",
     "1. Navigate to /vacation/sick-leaves-of-employees\n2. Check My department tab",
     "DM sees sick leaves for all employees in their department. State + Status filters available.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
]))

# ── TS-SL-ManagerView ──────────────────────────────────────────────────────
SUITES.append(("TS-SL-ManagerView", "Sick Leave Manager View", [
    ("TC-SL-110", "Manager page — My department tab content",
     "Logged in as PM/DM with subordinates having sick leaves.",
     "1. Navigate to /vacation/sick-leaves-of-employees\n2. Default tab: My department\n3. Verify content",
     "Table shows subordinates' sick leaves. Columns: Employee, Sick leave dates, Calendar days, State, Status, Actions.",
     "High", "Functional", "", "frontend-sick-leave-module", "Default redirect to /my-department"),
    ("TC-SL-111", "Manager page — My projects tab",
     "PM with projects containing employees with sick leaves.",
     "1. Click 'My projects' tab\n2. Verify content",
     "Shows sick leaves of employees on PM's projects. Pagination enabled (~20/page).",
     "High", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-112", "Manager page — State filter",
     "Manager view with mixed-state sick leaves.",
     "1. Click State filter\n2. Select 'Overdue'\n3. Verify filtered results",
     "Table filters to show only OVERDUE sick leaves (OPEN + past end_date).",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-113", "Manager page — Status filter",
     "Manager view with records in different accounting statuses.",
     "1. Click Status filter\n2. Select 'Paid'\n3. Verify filtered results",
     "Table filters to show only PAID accounting status records. Status shown as plain text (not dropdown).",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-114", "Manager create — employee search + date fields",
     "Manager create modal open.",
     "1. Type employee name in search field\n2. Select from dropdown\n3. Fill Start/End dates\n4. Optionally fill Number\n5. Save",
     "Employee async-search populates. Sick leave created for selected employee with manager as creator.",
     "High", "Functional", "", "frontend-sick-leave-module", "No subordination filter on employee search"),
    ("TC-SL-115", "Manager create — no file upload capability",
     "Manager create modal open.",
     "1. Observe modal fields",
     "No file upload area in manager create modal. Only employee, dates, number fields. File upload is employee-only capability.",
     "Medium", "Functional", "", "frontend-sick-leave-module", "Capability gap vs employee view"),
    ("TC-SL-116", "Manager create — force=true hardcoded (bypasses overlap)",
     "Manager creates sick leave overlapping existing one for employee.",
     "1. Manager creates overlapping sick leave\n2. Monitor API request",
     "force=true hardcoded in manager saga. Backend overlap check bypassed. No 409/conflict dialog shown to manager.",
     "High", "Bug verification", "", "frontend-sick-leave-module", "Combined with 100-record client cap = no overlap validation for managers"),
    ("TC-SL-117", "Manager view — overdue green checkmark action",
     "Manager view with OVERDUE sick leave records.",
     "1. Locate overdue record\n2. Observe action column",
     "Green checkmark action button visible for overdue records. OverdueSickLeaveCommand per-request check.",
     "Medium", "Functional", "", "sick-leave-service", ""),
    ("TC-SL-118", "Manager view — pagination",
     "My projects tab with >20 sick leave records.",
     "1. Navigate to My projects tab\n2. Scroll to bottom\n3. Check pagination controls",
     "Pagination controls shown. ~20 records per page. Navigate between pages.",
     "Low", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-119", "Manager view — sorting by date column",
     "Manager view with multiple records.",
     "1. Click 'Sick leave dates' column header\n2. Click again to reverse",
     "Default sort: descending by date. Click toggles direction. Sort encoding +/- prefix.",
     "Low", "Functional", "", "frontend-sick-leave-module", ""),
    ("TC-SL-120", "Manager view — detail modal from action icon",
     "Manager view with sick leave record.",
     "1. Click detail action icon on row\n2. Observe modal content",
     "Detail modal shows: Employee, Accountant, State, Status, Period, Calendar days, Number, Notify also. Read-only for manager.",
     "Medium", "Functional", "", "frontend-sick-leave-module", ""),
]))

# ── Risk Assessment ────────────────────────────────────────────────────────
RISKS = [
    ("No creation permission check", "Any authenticated user can create sick leave for any employee via API", "High", "Critical", "Critical",
     "Test with unrelated users creating for others. Verify no authorization beyond AUTHENTICATED_USER."),
    ("Unrestricted accounting status transitions", "Any-to-any transitions (PAID→NEW) allowed — no state machine guardrails", "High", "High", "Critical",
     "Test all backward transitions. Verify coupling effects (PAID→NEW leaves main CLOSED)."),
    ("Backend MIME bypass", "No server-side file type validation — any file accepted", "Medium", "Critical", "High",
     "Upload non-PDF/PNG/JPEG files via direct API call. Verify no MIME check."),
    ("Status coupling inconsistency", "PAID→CLOSED auto-close but PAID→NEW leaves CLOSED (orphaned state)", "High", "High", "High",
     "Test full status matrix with coupling side effects."),
    ("API token exclusion from CRUD", "Sick leave CRUD endpoints reject API tokens (AUTHENTICATED_USER only)", "High", "Medium", "High",
     "Automation blocked. Test all CRUD endpoints with API token — all should 403."),
    ("Manager force=true bypass", "Manager create hardcodes force=true + 100-record client cap = no overlap validation", "Medium", "High", "High",
     "Manager creates overlapping sick leave. Verify no conflict dialog shown."),
    ("Orphaned records on soft-delete", "File and notifyAlso associations remain after sick leave deletion", "Medium", "Low", "Medium",
     "Delete sick leave with files. Query DB for orphaned junction records."),
    ("Reference equality in getEditorType()", "== on BO instances always false — wrong notification template variant", "High", "Low", "Medium",
     "Create sick leave, check if correct OPEN vs OPEN_BY_SUPERVISOR template used."),
    ("End dialog title bug", "Close dialog says 'Delete the sick note?' instead of 'End the sick note?'", "High", "Low", "Medium",
     "Open close/end dialog, verify title text."),
    ("Router-level permission gap", "/sick-leave/my has no route-level permission check (TODO comment)", "Medium", "Medium", "Medium",
     "Access route as user without SICK_LEAVE:VIEW permission."),
    ("Stack trace leakage on 403", "Full Java stack trace (90+ frames) in error response for API token 403", "Low", "Medium", "Low",
     "Call file upload endpoint with API token, inspect error response body."),
    ("NoveoAI widget overlap", "Floating widget overlaps Status/Actions columns in accounting view", "High", "Low", "Low",
     "Open accounting page on standard viewport. Check widget position."),
]

# ── Feature Matrix ─────────────────────────────────────────────────────────
FEATURES = [
    # (feature, functional, negative, boundary, security, bug_verif, data_integrity, total, suite_ref)
    ("CRUD Lifecycle", 14, 3, 1, 0, 1, 1, 20, "TS-SL-CRUD"),
    ("Dual Status Model", 10, 1, 0, 0, 2, 1, 14, "TS-SL-DualStatus"),
    ("Accounting Workflow", 13, 0, 0, 0, 2, 1, 16, "TS-SL-Accounting"),
    ("File Handling", 6, 2, 1, 2, 1, 0, 12, "TS-SL-Files"),
    ("Notifications", 11, 0, 0, 0, 1, 0, 12, "TS-SL-Notifications"),
    ("Validation Rules", 8, 3, 2, 0, 1, 0, 14, "TS-SL-Validation"),
    ("Permissions & Security", 8, 1, 0, 3, 1, 0, 13, "TS-SL-Permissions"),
    ("Manager View", 9, 0, 0, 0, 1, 0, 10, "TS-SL-ManagerView"),
]

# ── Build Workbook ─────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════
# TAB 1: Plan Overview
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Plan Overview"
ws.sheet_properties.tabColor = TAB_GREEN

plan_content = [
    ("Sick Leave Module — Test Plan", TITLE_FONT),
    ("", None),
    ("Scope", SUBTITLE_FONT),
    ("Comprehensive testing of the Sick Leave module covering employee CRUD lifecycle,", BODY_FONT),
    ("dual status model (main + accounting), accounting workflow, file handling,", BODY_FONT),
    ("email notifications, form validation, role-based permissions, and manager view.", BODY_FONT),
    ("", None),
    ("Objectives", SUBTITLE_FONT),
    ("1. Validate complete sick leave lifecycle: create, edit, close, reopen, delete", BODY_FONT),
    ("2. Verify dual status system coupling (main status ↔ accounting status)", BODY_FONT),
    ("3. Test accounting workflow: inline status changes, comments, filters", BODY_FONT),
    ("4. Verify file upload/download with security validation", BODY_FONT),
    ("5. Validate 5 notification event types with correct templates and recipients", BODY_FONT),
    ("6. Test form validation rules (Yup frontend + bean backend)", BODY_FONT),
    ("7. Verify role-based access control and known security gaps", BODY_FONT),
    ("8. Test manager-specific creation flow and view functionality", BODY_FONT),
    ("", None),
    ("Approach", SUBTITLE_FONT),
    ("Testing combines UI (Playwright), API (curl/Swagger), and DB (PostgreSQL) actions.", BODY_FONT),
    ("UI testing on timemachine environment as multiple users (employee, manager, accountant, admin).", BODY_FONT),
    ("API testing via JWT session auth (API tokens rejected for CRUD — only search/count work).", BODY_FONT),
    ("Known bugs (8 total, 1 Critical, 2 High, 3 Medium, 4 Low) verified inline.", BODY_FONT),
    ("", None),
    ("Test Data Strategy", SUBTITLE_FONT),
    ("- Timemachine env: 348 existing sick leaves, ~104-114/year, max duration 140-141 days", BODY_FONT),
    ("- Use employee logins from test-data-landscape: perekrest (accountant), dergachev (manager)", BODY_FONT),
    ("- Create/delete within session to maintain environment cleanliness", BODY_FONT),
    ("- DB mining: SELECT from ttt_vacation.sick_leave for specific status combinations", BODY_FONT),
    ("- Overlap testing: create temporary sick leaves with known overlapping dates", BODY_FONT),
    ("", None),
    ("Environment Requirements", SUBTITLE_FONT),
    ("- Primary: timemachine (ttt-timemachine.noveogroup.com)", BODY_FONT),
    ("- Secondary: qa-1 (ttt-qa-1.noveogroup.com)", BODY_FONT),
    ("- Cross-env: stage for comparison (ttt-stage.noveogroup.com)", BODY_FONT),
    ("- Browser: Chrome via Playwright (VPN required)", BODY_FONT),
    ("- Database: PostgreSQL on port 5433, schema ttt_vacation", BODY_FONT),
    ("", None),
    ("Qase Existing Coverage", SUBTITLE_FONT),
    ("57 existing cases cover display/notification only (0 lifecycle CRUD cases):", BODY_FONT),
    ("- Color indication in My Tasks / Employee Tasks (6 cases)", BODY_FONT),
    ("- Confirmation table display (6 cases)", BODY_FONT),
    ("- Planner display (6 cases)", BODY_FONT),
    ("- Email notifications (14 cases)", BODY_FONT),
    ("- Accounting sort/filter/table/actions (25 cases)", BODY_FONT),
    ("- My Sick Leaves / Employee Sick Leaves suites: empty placeholders (0 cases)", BODY_FONT),
    ("THIS TEST PLAN fills the lifecycle CRUD gap with 121 new test cases.", BODY_FONT),
    ("", None),
    ("Test Suite Index", SUBTITLE_FONT),
]

row = 1
for text, font in plan_content:
    cell = ws.cell(row=row, column=1, value=text)
    if font:
        cell.font = font
    cell.alignment = WRAP
    row += 1

# Suite hyperlinks
suite_info = []
for tab_name, title, cases in SUITES:
    count = len(cases)
    suite_info.append((tab_name, title, count))

for tab_name, title, count in suite_info:
    cell = ws.cell(row=row, column=1)
    cell.value = f"{tab_name}: {title} — {count} cases"
    cell.font = LINK_FONT
    cell.hyperlink = f"#'{tab_name}'!A1"
    row += 1

ws.column_dimensions["A"].width = 100

# ═══════════════════════════════════════════════════════════════════════════
# TAB 2: Feature Matrix
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Feature Matrix")
ws2.sheet_properties.tabColor = TAB_GREEN

fm_headers = ["Feature Area", "Functional", "Negative", "Boundary", "Security",
              "Bug Verification", "Data Integrity", "Total", "Test Suite"]
for c, h in enumerate(fm_headers, 1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(fm_headers))

for i, (feat, func, neg, bnd, sec, bug, data, total, ref) in enumerate(FEATURES):
    r = i + 2
    ws2.cell(row=r, column=1, value=feat)
    ws2.cell(row=r, column=2, value=func)
    ws2.cell(row=r, column=3, value=neg)
    ws2.cell(row=r, column=4, value=bnd)
    ws2.cell(row=r, column=5, value=sec)
    ws2.cell(row=r, column=6, value=bug)
    ws2.cell(row=r, column=7, value=data)
    ws2.cell(row=r, column=8, value=total)
    link_cell = ws2.cell(row=r, column=9, value=ref)
    link_cell.font = LINK_FONT
    link_cell.hyperlink = f"#'{ref}'!A1"
    style_data_row(ws2, r, len(fm_headers), i)

# Totals row
total_r = len(FEATURES) + 2
ws2.cell(row=total_r, column=1, value="TOTAL").font = Font(name=ARIAL, bold=True, size=10)
for c in range(2, 9):
    val = sum(row[c - 1] for row in FEATURES)  # index 1-7 maps to columns 2-8
    ws2.cell(row=total_r, column=c, value=val).font = Font(name=ARIAL, bold=True, size=10)
for c in range(1, len(fm_headers) + 1):
    ws2.cell(row=total_r, column=c).border = THIN_BORDER
    ws2.cell(row=total_r, column=c).fill = PatternFill("solid", fgColor="B4C6E7")

set_col_widths(ws2, [22, 12, 10, 10, 10, 16, 14, 8, 18])
ws2.auto_filter.ref = f"A1:I{total_r}"

# ═══════════════════════════════════════════════════════════════════════════
# TAB 3: Risk Assessment
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Risk Assessment")
ws3.sheet_properties.tabColor = TAB_GREEN

risk_headers = ["Risk", "Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
for c, h in enumerate(risk_headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, len(risk_headers))

for i, (risk, desc, like, impact, sev, mit) in enumerate(RISKS):
    r = i + 2
    ws3.cell(row=r, column=1, value=risk)
    ws3.cell(row=r, column=2, value=desc)
    ws3.cell(row=r, column=3, value=like)
    ws3.cell(row=r, column=4, value=impact)
    sev_cell = ws3.cell(row=r, column=5, value=sev)
    ws3.cell(row=r, column=6, value=mit)
    style_data_row(ws3, r, len(risk_headers), i)
    # Color severity
    if sev == "Critical":
        sev_cell.fill = RISK_CRIT
    elif sev == "High":
        sev_cell.fill = RISK_HIGH
    elif sev == "Medium":
        sev_cell.fill = RISK_MED
    elif sev == "Low":
        sev_cell.fill = RISK_LOW

set_col_widths(ws3, [30, 55, 12, 12, 12, 55])
ws3.auto_filter.ref = f"A1:{get_column_letter(len(risk_headers))}{len(RISKS) + 1}"

# ═══════════════════════════════════════════════════════════════════════════
# TABs 4+: Test Suite sheets (TS-*)
# ═══════════════════════════════════════════════════════════════════════════
TC_HEADERS = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
              "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
TC_WIDTHS = [12, 35, 30, 45, 45, 10, 16, 22, 30, 30]

for tab_name, title, cases in SUITES:
    ws_tc = wb.create_sheet(tab_name)
    ws_tc.sheet_properties.tabColor = TAB_BLUE

    # Back-link to Plan Overview
    back = ws_tc.cell(row=1, column=1, value="← Back to Plan Overview")
    back.font = BACK_LINK_FONT
    back.hyperlink = "#'Plan Overview'!A1"

    # Headers at row 2
    for c, h in enumerate(TC_HEADERS, 1):
        ws_tc.cell(row=2, column=c, value=h)
    style_header_row(ws_tc, 2, len(TC_HEADERS))

    # Cases
    for i, case in enumerate(cases):
        r = i + 3
        for c, val in enumerate(case, 1):
            ws_tc.cell(row=r, column=c, value=val)
        style_data_row(ws_tc, r, len(TC_HEADERS), i)

    set_col_widths(ws_tc, TC_WIDTHS)
    ws_tc.auto_filter.ref = f"A2:{get_column_letter(len(TC_HEADERS))}{len(cases) + 2}"

# ═══════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT)

total_cases = sum(len(cases) for _, _, cases in SUITES)
print(f"Generated {OUTPUT}")
print(f"Tabs: Plan Overview + Feature Matrix + Risk Assessment + {len(SUITES)} TS-* sheets")
print(f"Total test cases: {total_cases}")
print(f"Risks: {len(RISKS)}")
for tab_name, title, cases in SUITES:
    print(f"  {tab_name}: {title} — {len(cases)} cases")
