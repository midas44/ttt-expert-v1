#!/usr/bin/env python3
"""Generate unified security.xlsx workbook with test plan + all test suites."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

OUTPUT = "/home/v/Dev/ttt-expert-v1/expert-system/output/security/security.xlsx"

# -- Styles ---------------------------------------------------------------
ARIAL = "Arial"
HEADER_FONT = Font(name=ARIAL, bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
SUBTITLE_FONT = Font(name=ARIAL, bold=True, size=12)
BODY_FONT = Font(name=ARIAL, size=10)
LINK_FONT = Font(name=ARIAL, size=10, color="0563C1", underline="single")
BACK_LINK_FONT = Font(name=ARIAL, size=9, color="0563C1", underline="single")
ROW_EVEN = PatternFill("solid", fgColor="D6E4F0")
ROW_ODD = PatternFill("solid", fgColor="FFFFFF")
RISK_CRIT = PatternFill("solid", fgColor="FF6B6B")
RISK_HIGH = PatternFill("solid", fgColor="FFA07A")
RISK_MED = PatternFill("solid", fgColor="FFD700")
RISK_LOW = PatternFill("solid", fgColor="90EE90")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TAB_GREEN = "00B050"
TAB_BLUE = "4472C4"


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def style_data_row(ws, row, max_col, idx):
    fill = ROW_EVEN if idx % 2 == 0 else ROW_ODD
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# -- Test Suite Definitions ------------------------------------------------
# Each case: (id, title, preconditions, steps, expected, priority, type, req_ref, component, notes)

SUITES = []

# ==========================================================================
# TS-SEC-JWTAuth -- JWT Authentication & Session Management
# ==========================================================================
SUITES.append(("TS-SEC-JWTAuth", "JWT Authentication & Session Management", [
    ("TC-SEC-001", "Obtain JWT token via CAS login flow",
     "Valid CAS credentials. Browser session.",
     "1. Navigate to TTT login page\n2. Enter username (CAS SSO)\n3. Observe redirect flow\n4. Check TTT_JWT_TOKEN cookie/header in subsequent requests",
     "JWT token issued after successful CAS authentication. Token contains claims: csId, login, officeId, managerId, employeeName, authorities. RS256 signed.",
     "Critical", "Functional", "AUTH-JWT-001", "common-security-jwt, JwtTokenProcessor", ""),

    ("TC-SEC-002", "JWT token contains correct authorities for multi-role user",
     "User with multiple global roles (e.g. ROLE_ADMIN + ROLE_ACCOUNTANT + ROLE_EMPLOYEE).",
     "1. Login as multi-role user (e.g. perekrest)\n2. Decode JWT token (base64)\n3. Inspect 'authorities' claim array",
     "All assigned global roles present in authorities claim. AUTHENTICATED_USER authority included. Roles match employee_global_roles table entries.",
     "High", "Functional", "AUTH-JWT-002", "JwtTokenProcessor, employee_global_roles", "Verify against DB: SELECT role FROM employee_global_roles WHERE employee_id = (SELECT id FROM employee WHERE login = 'perekrest')"),

    ("TC-SEC-003", "JWT token expiration after 1 day",
     "Valid JWT token obtained.",
     "1. Obtain JWT token and note issuance time\n2. Decode token, verify 'exp' claim = iat + 86400s\n3. Wait for expiration (or manipulate clock on timemachine)\n4. Make API request with expired token",
     "Token rejected after expiry. Response: 401 Unauthorized. No automatic refresh — user must re-authenticate via CAS.",
     "High", "Functional", "AUTH-JWT-003", "JwtTokenAuthenticationFilter", "1-day hardcoded in JwtTokenProcessor. No refresh mechanism exists."),

    ("TC-SEC-004", "JWT token with tampered signature rejected",
     "Valid JWT token.",
     "1. Take valid JWT token\n2. Modify payload (change login to another user)\n3. Re-encode with incorrect key\n4. Send request with tampered token in TTT_JWT_TOKEN header",
     "Request rejected. RS256 signature verification fails. 401 Unauthorized returned.",
     "Critical", "Security", "AUTH-JWT-004", "JwtTokenAuthenticationFilter", "RSA key pair — private key signs, public key verifies"),

    ("TC-SEC-005", "JWT token with wrong header name ignored",
     "Valid JWT token.",
     "1. Send API request with JWT in 'Authorization: Bearer' header instead of TTT_JWT_TOKEN\n2. Check response",
     "Request not authenticated via JWT filter — falls through to API token filter, then to anonymous. Expect 401 or 403 depending on endpoint.",
     "Medium", "Negative", "AUTH-JWT-005", "JwtTokenAuthenticationFilter", "Custom header name TTT_JWT_TOKEN, not standard Authorization"),

    ("TC-SEC-006", "JWT grants AUTHENTICATED_USER authority",
     "Valid JWT token for any active employee.",
     "1. Login via CAS, get JWT\n2. Call endpoint requiring hasAuthority('AUTHENTICATED_USER') — e.g. GET /v1/sick-leaves/{id}\n3. Verify access",
     "Access granted. JWT users always have AUTHENTICATED_USER authority per TokenConst design.",
     "High", "Functional", "AUTH-JWT-006", "TokenConst, JwtTokenAuthenticationFilter", "Key distinction: JWT gets AUTHENTICATED_USER, API tokens do NOT"),

    ("TC-SEC-007", "Empty JWT token header returns 401",
     "No authentication.",
     "1. Send API request with TTT_JWT_TOKEN header set to empty string\n2. Check response",
     "401 Unauthorized. Filter skips empty token, falls through to unauthorized handler.",
     "Medium", "Negative", "AUTH-JWT-007", "JwtTokenAuthenticationFilter", ""),

    ("TC-SEC-008", "Malformed JWT token (invalid base64) returns 401",
     "No valid credentials.",
     "1. Send API request with TTT_JWT_TOKEN = 'not-a-valid-jwt-token'\n2. Check response",
     "401 Unauthorized. JWT parsing fails gracefully without server error.",
     "Medium", "Negative", "AUTH-JWT-008", "JwtTokenProcessor", "Verify no 500 error or stack trace in response"),

    ("TC-SEC-009", "JWT token for deactivated employee rejected",
     "Employee account deactivated in DB (employee.active = false or dismissed).",
     "1. Obtain JWT token before deactivation\n2. Deactivate employee\n3. Make API request with pre-existing token\n4. Check response",
     "EmployeeActivationResolver rejects inactive employee. Request fails with 401 or 403.",
     "High", "Security", "AUTH-JWT-009", "EmployeeActivationResolver", "Tests token revocation on account changes"),

    ("TC-SEC-010", "Concurrent JWT tokens for same user both valid",
     "Same user authenticated from two different browsers/sessions.",
     "1. Login as userA from Browser 1, get JWT-1\n2. Login as userA from Browser 2, get JWT-2\n3. Use JWT-1 to make API request\n4. Use JWT-2 to make API request",
     "Both tokens work simultaneously. Stateless session policy — no server-side session invalidation.",
     "Medium", "Functional", "AUTH-JWT-010", "SecurityConfiguration", "STATELESS session creation policy"),

    ("TC-SEC-011", "JWT filter chain order: JWT checked before API token",
     "Request with both TTT_JWT_TOKEN and API_SECRET_TOKEN headers.",
     "1. Send request with valid JWT in TTT_JWT_TOKEN and valid API token in API_SECRET_TOKEN\n2. Check which identity is used\n3. Verify authorities",
     "JWT identity used (first match wins). Authorities include AUTHENTICATED_USER from JWT, not API token permissions. Filter chain: JWT → API token → Authorization.",
     "Medium", "Functional", "AUTH-JWT-011", "SecurityConfiguration filter chain", "Filter chain order matters for dual-header edge case"),

    ("TC-SEC-012", "readOnly claim in JWT restricts write operations",
     "User with readOnly=true flag (e.g. VIEW_ALL role or specific config).",
     "1. Login as readOnly user\n2. Attempt write operation (e.g. create vacation, create report)\n3. Check response",
     "Write operations blocked. readOnly flag checked in permission providers. VIEW permissions granted, CREATE/EDIT denied.",
     "High", "Functional", "AUTH-JWT-012", "PermissionProvider classes", "readOnly flag in JWT claims, enforced at permission provider level"),
]))

# ==========================================================================
# TS-SEC-APIToken -- API Token Authentication & Permissions
# ==========================================================================
SUITES.append(("TS-SEC-APIToken", "API Token Authentication & Permissions", [
    ("TC-SEC-013", "API token via header (API_SECRET_TOKEN) grants specific permissions",
     "Valid EMPLOYEE-type API token with REPORTS_VIEW permission.",
     "1. GET /v1/reports with API_SECRET_TOKEN header\n2. Verify 200 response with report data",
     "Access granted. Only REPORTS_VIEW permission available — other endpoints requiring different permissions return 403.",
     "Critical", "Functional", "AUTH-API-001", "ApiTokenAuthenticationFilter, DatabaseApiTokenResolver", ""),

    ("TC-SEC-014", "API token via URL query param (?token=) accepted",
     "Valid API token.",
     "1. GET /v1/reports?token={secret_key}&startDate=...&endDate=...\n2. Verify response",
     "Access granted. Token accepted from query parameter. SECURITY RISK: token visible in server logs, browser history, HTTP Referer headers.",
     "High", "Security", "AUTH-API-002", "ApiTokenAuthenticationFilter", "Known issue: DI-46. Token exposed in URL — information disclosure risk."),

    ("TC-SEC-015", "API token does NOT grant AUTHENTICATED_USER authority",
     "Valid API token with all 21 permissions.",
     "1. Use API token to call endpoint requiring hasAuthority('AUTHENTICATED_USER') only — e.g. GET /v1/sick-leaves/{id}\n2. Check response",
     "403 Forbidden. API tokens never receive AUTHENTICATED_USER — only explicit ApiPermission enum values from token_permissions table.",
     "Critical", "Functional", "AUTH-API-003", "TokenConst, ApiTokenAuthenticationFilter", "Core security boundary between JWT and API token auth"),

    ("TC-SEC-016", "EMPLOYEE-type token (21 permissions average) vs APPLICATION-type token (4.7 permissions average)",
     "One EMPLOYEE token and one APPLICATION token configured.",
     "1. Query token_permissions for each token type\n2. Compare permission sets\n3. Test endpoint access with each",
     "EMPLOYEE tokens have broader permissions (avg 21). APPLICATION tokens are narrowly scoped (avg 4.7). Access correctly restricted per token's permission set.",
     "High", "Functional", "AUTH-API-004", "token, token_permissions tables", "DB: 1833 EMPLOYEE tokens, 23 APPLICATION tokens"),

    ("TC-SEC-017", "Invalid API token returns 401",
     "No valid token.",
     "1. Send request with API_SECRET_TOKEN = 'invalid-uuid-token'\n2. Check response",
     "401 Unauthorized. DatabaseApiTokenResolver finds no matching token in DB.",
     "High", "Negative", "AUTH-API-005", "DatabaseApiTokenResolver", ""),

    ("TC-SEC-018", "API token with VACATIONS_CREATE can create vacation but not sick leave",
     "API token with VACATIONS_CREATE permission.",
     "1. POST /v1/vacations with valid payload — expect 201\n2. POST /v1/sick-leaves with valid payload — expect 403",
     "Vacation created (Pattern A: hasAnyAuthority AUTHENTICATED_USER OR VACATIONS_CREATE). Sick leave fails (Pattern B: hasAuthority AUTHENTICATED_USER only).",
     "Critical", "Functional", "AUTH-API-006", "VacationController, SickLeaveController", "Key inconsistency: sick leave uses Pattern B, vacation uses Pattern A"),

    ("TC-SEC-019", "API token permission REPORTS_EDIT allows report creation",
     "API token with REPORTS_EDIT permission.",
     "1. POST /v1/reports with valid task report payload\n2. Verify 201 Created",
     "Report created. REPORTS_EDIT permission enables create, update, and delete for task reports via Pattern A endpoints.",
     "High", "Functional", "AUTH-API-007", "TaskReportController", ""),

    ("TC-SEC-020", "API token permission REPORTS_APPROVE allows report state changes",
     "API token with REPORTS_APPROVE permission.",
     "1. PATCH /v1/reports/{id} with state=APPROVED\n2. Verify state change",
     "Report approved. REPORTS_APPROVE grants confirmation operations via API token.",
     "High", "Functional", "AUTH-API-008", "TaskReportController.patch", ""),

    ("TC-SEC-021", "All 21 ApiPermission enum values map to correct endpoint groups",
     "API token with single permission at a time.",
     "For each of the 21 ApiPermission values:\n1. Create token with only that permission\n2. Test relevant endpoint group\n3. Verify access granted for matching endpoints only\n4. Verify 403 for non-matching endpoints",
     "Each permission grants access to its designated endpoint group only: EMPLOYEES_VIEW→employee endpoints, OFFICES_VIEW→office endpoints, STATISTICS_VIEW→statistics, etc.",
     "High", "Functional", "AUTH-API-009", "ApiPermission enum", "21 permissions: EMPLOYEES, OFFICES, STATISTICS, SUGGESTIONS, PROJECTS, TASKS, ASSIGNMENTS, REPORTS, VACATIONS, VACATION_DAYS, CALENDAR, FILES + VIEW/EDIT/APPROVE/CREATE variants"),

    ("TC-SEC-022", "API token for deleted/deactivated employee token owner",
     "EMPLOYEE-type token whose owner employee is deactivated.",
     "1. Note employee with active API token\n2. Deactivate employee\n3. Use token for API call\n4. Check response",
     "Token should be invalidated or access denied. Verify DatabaseApiTokenResolver checks owner activation status.",
     "High", "Security", "AUTH-API-010", "DatabaseApiTokenResolver, token.owner FK", ""),

    ("TC-SEC-023", "Simultaneous JWT and API token — JWT takes precedence",
     "Valid JWT and valid API token for different users.",
     "1. Send request with TTT_JWT_TOKEN for userA and API_SECRET_TOKEN for userB\n2. Check which identity resolves\n3. Verify authorities match JWT user",
     "JWT filter executes first (filter chain order). Request processed as userA with AUTHENTICATED_USER + roles. API token ignored.",
     "Medium", "Functional", "AUTH-API-011", "SecurityConfiguration filter chain", "Same as TC-SEC-011 but emphasizing identity resolution"),

    ("TC-SEC-024", "API token permission boundary: CALENDAR_VIEW on calendar service",
     "API token with CALENDAR_VIEW permission.",
     "1. GET /api/calendar/v1/offices with API token\n2. Check response\n3. Note calendar service uses compound auth: hasAnyRole AND hasAnyAuthority",
     "Calendar service uses Pattern C (compound): requires BOTH a role AND an authority. API tokens may fail if compound check requires role that API token doesn't carry.",
     "High", "Functional", "AUTH-API-012", "Calendar SecurityConfiguration", "Pattern C is stricter — unique to calendar service"),
]))

# ==========================================================================
# TS-SEC-RoleAccess -- Role-Based UI Page Access
# ==========================================================================
SUITES.append(("TS-SEC-RoleAccess", "Role-Based UI Page Access Matrix", [
    ("TC-SEC-025", "Employee (ROLE_EMPLOYEE) — default pages accessible",
     "Logged in as user with only ROLE_EMPLOYEE.",
     "1. Navigate to My Tasks (/)\n2. Navigate to Planner (/planner)\n3. Navigate to FAQ (/faq)\n4. Navigate to Statistics (/statistics)\n5. Navigate to My Vacations (/vacation/my)\n6. Navigate to Vacation Chart (/vacation/chart)\n7. Navigate to Account (/admin/account)",
     "All 7 pages load successfully. These are available to all authenticated users without specific permission requirements.",
     "High", "Functional", "RBAC-PAGE-001", "frontend PrivateRoute, roles-permissions", ""),

    ("TC-SEC-026", "Employee cannot access manager/admin pages",
     "Logged in as user with only ROLE_EMPLOYEE (no PM, DM, ACC, ADM).",
     "1. Navigate to /approve (Confirmation)\n2. Navigate to /admin/projects\n3. Navigate to /admin/employees\n4. Navigate to /admin/settings\n5. Navigate to /admin/calendar\n6. Navigate to /admin/api",
     "All pages blocked. PrivateRoute redirects to home page or shows 'Access Denied'. Requires TASKS:VIEW_APPROVES, PROJECTS:VIEW, EMPLOYEES:VIEW, SETTINGS:VIEW, CALENDARS:VIEW, TOKENS:VIEW respectively.",
     "Critical", "Security", "RBAC-PAGE-002", "frontend PrivateRoute", ""),

    ("TC-SEC-027", "Project Manager page access (ROLE_PROJECT_MANAGER)",
     "Logged in as PM-only user.",
     "1. Navigate to /approve — expect access (TASKS:VIEW_APPROVES)\n2. Navigate to /admin/projects — expect access (PROJECTS:VIEW)\n3. Navigate to /admin/employees — expect access (EMPLOYEES:VIEW)\n4. Navigate to /vacation/request — expect access (VACATIONS:VIEW_APPROVES)\n5. Navigate to /admin/settings — expect denied (SETTINGS:VIEW = ADM only)\n6. Navigate to /admin/calendar — expect denied (CALENDARS:VIEW = CACC/ADM)",
     "PM has access to Confirmation, Projects, Employees, Vacation Requests. Denied for Settings, Calendar, API Tokens, Export.",
     "High", "Functional", "RBAC-PAGE-003", "PermissionProvider.projectManager()", ""),

    ("TC-SEC-028", "Accountant page access (ROLE_ACCOUNTANT)",
     "Logged in as ACC-only user.",
     "1. Navigate to /approve — expect access (TASKS:VIEW_APPROVES)\n2. Navigate to /vacation/payment — expect access (VACATIONS:VIEW_PAYMENTS)\n3. Navigate to /vacation/days-correction — expect access (VACATIONS:VIEW_DAYS)\n4. Navigate to /vacation/sick-leaves-of-employees — expect access (SICK_LEAVE_VIEW)\n5. Navigate to /admin/offices — expect access (OFFICES:VIEW)\n6. Navigate to /admin/projects — expect denied (PROJECTS:VIEW = PM/DM/ADM)",
     "Accountant has accounting-oriented access: payments, days correction, sick leaves, offices. No project management access.",
     "High", "Functional", "RBAC-PAGE-004", "PermissionProvider.accountant()", ""),

    ("TC-SEC-029", "Admin has access to all pages (ROLE_ADMIN)",
     "Logged in as ROLE_ADMIN user.",
     "1. Navigate to all 23+ pages listed in page access matrix\n2. Verify each page loads with full functionality",
     "Admin can access every page including Settings, Calendar, API Tokens, Export. All permission classes granted.",
     "High", "Functional", "RBAC-PAGE-005", "PermissionProvider.admin()", "8 admin users in DB"),

    ("TC-SEC-030", "VIEW_ALL role grants read-only access to all pages",
     "Logged in as ROLE_VIEW_ALL user.",
     "1. Navigate to all view-related pages\n2. Attempt to create/edit on pages that allow it\n3. Verify readOnly flag blocks writes",
     "VIEW_ALL user can view all pages (same view permissions as admin). Write operations blocked by readOnly flag. Create buttons disabled/hidden.",
     "High", "Functional", "RBAC-PAGE-006", "PermissionProvider.viewAll()", "13 VIEW_ALL users in DB. readOnly=true in JWT claims."),

    ("TC-SEC-031", "SECURITY GAP: Accounting subroutes lack frontend permission checks",
     "Logged in as regular employee (ROLE_EMPLOYEE only).",
     "1. Navigate directly to /accounting/salary\n2. Navigate to /accounting/vacation-payout\n3. Navigate to /accounting/days-correction\n4. Navigate to /accounting/periods",
     "BUG: All 4 pages load despite no ACCOUNTING:VIEW permission. Frontend routes missing permission guards. Backend API still requires auth, so data may not load, but page renders.",
     "Critical", "Bug verification", "GAP-ACCT-001", "frontend AccountingRoute, role-permission-matrix", "Known gap: 4 of 5 accounting subroutes unprotected at frontend level"),

    ("TC-SEC-032", "SECURITY GAP: Sick leave route lacks permission check (TODO in code)",
     "Logged in as regular employee.",
     "1. Navigate to /sick-leave\n2. Check if page loads\n3. Check source code comment",
     "BUG: Page loads without permission check. Source code contains Russian TODO comment: 'Add permission check. Component created for future sprints.' Functional but unguarded.",
     "High", "Bug verification", "GAP-SL-001", "SickLeaveRoute/index.js", "Known issue DI-33. TODO remains in code."),

    ("TC-SEC-033", "Department Manager (ROLE_DEPARTMENT_MANAGER) — vacation employees page",
     "Logged in as DM user.",
     "1. Navigate to /vacation/vacation-days — expect access (VACATIONS:VIEW_EMPLOYEES, DM exclusive)\n2. Navigate to /budget-notifications — expect access (BUDGET_NOTIF:VIEW)\n3. Navigate to /admin/projects — expect access (PROJECTS:VIEW)\n4. Navigate to /admin/settings — expect denied",
     "DM has unique access to Vacation Days page (DM-exclusive). Budget Notifications also accessible. Settings denied.",
     "High", "Functional", "RBAC-PAGE-007", "PermissionProvider.departmentManager()", "29 DMs in DB"),

    ("TC-SEC-034", "Contractor (ROLE_CONTRACTOR) — minimal access",
     "Logged in as user with ROLE_CONTRACTOR only.",
     "1. Navigate to My Tasks, Planner, FAQ, Statistics — expect access\n2. Navigate to /vacation/my — check access\n3. Navigate to /approve, /admin/projects, /vacation/request — expect denied",
     "Contractor gets basic personal access only. No explicit permissions granted in any PermissionProvider. Access limited to pages with no permission requirement.",
     "High", "Functional", "RBAC-PAGE-008", "PermissionProvider (no contractor handler)", "159 contractors in DB. Unclear permission model is a gap."),

    ("TC-SEC-035", "Chief Accountant (ROLE_CHIEF_ACCOUNTANT) — calendar admin access",
     "Logged in as CACC user.",
     "1. Navigate to /admin/calendar — expect access (CALENDARS:VIEW = CACC/ADM/VALL)\n2. Navigate to /admin/offices — expect access (OFFICES:VIEW)\n3. Navigate to /admin/salary — expect access (ACCOUNTING:NOTIFY)\n4. Navigate to /admin/projects — expect denied",
     "CACC has accounting + calendar admin access. Calendar admin is unique to CACC/ADM/VALL.",
     "Medium", "Functional", "RBAC-PAGE-009", "PermissionProvider.chiefAccountant()", "Only 2 CACC users in DB"),

    ("TC-SEC-036", "Multi-role user gets union of all role permissions",
     "User with ROLE_ADMIN + ROLE_ACCOUNTANT + ROLE_PROJECT_MANAGER.",
     "1. Login as multi-role user\n2. Navigate to all pages\n3. Verify access is union (not intersection) of all assigned roles",
     "All pages accessible. Permission is additive — having ROLE_ADMIN already grants all permissions. Multi-role grants union of permission sets.",
     "Medium", "Functional", "RBAC-PAGE-010", "PermissionProvider composition", ""),

    ("TC-SEC-037", "Employee Reports page — self-access without EMPLOYEES:VIEW",
     "Logged in as regular employee (no EMPLOYEES:VIEW permission).",
     "1. Navigate to /employee-reports/{own-login}\n2. Verify own reports visible\n3. Navigate to /employee-reports/{other-login}\n4. Check response",
     "Own reports accessible (self-access exception). Other employees' reports blocked — requires EMPLOYEES:VIEW permission.",
     "High", "Functional", "RBAC-PAGE-011", "EmployeeReports component", "Self-access bypass for personal data"),

    ("TC-SEC-038", "Statistics page — tab visibility based on permissions",
     "Logged in as user with specific role (e.g. PM vs employee).",
     "1. Navigate to /statistics\n2. Check which tabs are visible\n3. Compare with statisticsPermissionsTabName mapping",
     "BUG: Most tab-to-permission mappings commented out in code. currentTab defaults to null. Users with specific permissions see no default tab selected on page load.",
     "Medium", "Bug verification", "GAP-STAT-001", "permissions/constants.js", "Known DI-35: statisticsPermissionsTabName mappings commented out"),
]))

# ==========================================================================
# TS-SEC-EndpointPerm -- API Endpoint Authorization Enforcement
# ==========================================================================
SUITES.append(("TS-SEC-EndpointPerm", "API Endpoint Authorization Enforcement", [
    ("TC-SEC-039", "Pattern A endpoints accept both JWT and API token",
     "Valid JWT and valid API token (with matching permission).",
     "1. GET /v1/reports with JWT token — expect 200\n2. GET /v1/reports with API token (REPORTS_VIEW) — expect 200\n3. GET /v1/vacations with JWT — expect 200\n4. GET /v1/vacations with API token (VACATIONS_VIEW) — expect 200",
     "Both auth methods work. Pattern A: hasAnyAuthority('AUTHENTICATED_USER', 'SPECIFIC_PERMISSION') — correct dual-auth pattern.",
     "High", "Functional", "ENDP-AUTH-001", "TaskReportController, VacationController", "Pattern A is the recommended pattern per auth doc"),

    ("TC-SEC-040", "Pattern B endpoints reject API token with 403",
     "API token with all 21 permissions.",
     "1. GET /v1/sick-leaves/{id} with API token — expect 403\n2. POST /v1/sick-leaves with API token — expect 403\n3. PATCH /v1/sick-leaves/{id} with API token — expect 403\n4. DELETE /v1/sick-leaves/{id} with API token — expect 403\n5. GET /v1/reports/accounting with API token — expect 403\n6. GET /v1/task-report-warnings with API token — expect 403",
     "All endpoints return 403. Pattern B: hasAuthority('AUTHENTICATED_USER') only — API tokens excluded by design (no AUTHENTICATED_USER).",
     "Critical", "Functional", "ENDP-AUTH-002", "SickLeaveController, TaskReportController", "Major inconsistency: sick leave CRUD inaccessible via API automation"),

    ("TC-SEC-041", "Pattern C (calendar service) compound role+authority check",
     "JWT user with calendar-related roles. API token with CALENDAR_VIEW.",
     "1. GET /api/calendar/v1/offices with JWT — check access\n2. GET /api/calendar/v1/offices with API token (CALENDAR_VIEW) — check access\n3. Compare behavior",
     "Calendar service uses compound auth: hasAnyRole(...) AND hasAnyAuthority(...). Stricter than Pattern A — requires both conditions met simultaneously.",
     "High", "Functional", "ENDP-AUTH-003", "Calendar SecurityConfiguration", "Unique to calendar service — different from TTT and vacation services"),

    ("TC-SEC-042", "Missing @PreAuthorize on /reports/effort endpoint",
     "Any valid authenticated request (minimal permissions).",
     "1. Create API token with only EMPLOYEES_VIEW (not REPORTS-related)\n2. GET /v1/reports/effort?taskId={valid-id}\n3. Check response",
     "BUG: Access granted despite no REPORTS permission. Endpoint missing @PreAuthorize annotation entirely. Any authenticated request can access.",
     "High", "Bug verification", "BUG-REPORT-6", "TaskReportController.effort", "Known DI-53. Missing auth annotation."),

    ("TC-SEC-043", "Missing @PreAuthorize on /reports/employee-projects endpoint",
     "Any valid authenticated request (minimal permissions).",
     "1. Create API token with only EMPLOYEES_VIEW\n2. GET /v1/reports/employee-projects\n3. Check response",
     "BUG: Access granted despite no REPORTS permission. Endpoint missing @PreAuthorize annotation.",
     "High", "Bug verification", "BUG-REPORT-6", "TaskReportController", "Same issue as TC-SEC-042"),

    ("TC-SEC-044", "Sick leave search+count work via API token but CRUD does not",
     "API token with VACATIONS_VIEW permission.",
     "1. GET /v1/sick-leaves (search) — expect 200\n2. GET /v1/sick-leaves/count — expect 200\n3. GET /v1/sick-leaves/{id} — expect 403\n4. POST /v1/sick-leaves — expect 403",
     "Search and count use Pattern A (AUTHENTICATED_USER OR VACATIONS_VIEW) — works. CRUD uses Pattern B (AUTHENTICATED_USER only) — blocks API token. Inconsistent permission design.",
     "High", "Functional", "ENDP-AUTH-004", "SickLeaveController", "Known DI-45. Blocks API test automation for entire sick leave module."),

    ("TC-SEC-045", "Period API permission inconsistency: report min/max vs approve min/max",
     "API token with OFFICES_VIEW permission.",
     "1. GET /v1/offices/periods/report/min — expect 403 (JWT only)\n2. GET /v1/offices/periods/report/max — expect 403\n3. GET /v1/offices/periods/approve/min — expect 200 (accepts API token)\n4. GET /v1/offices/periods/approve/max — expect 200",
     "BUG: Report period min/max missing OFFICES_VIEW fallback, unlike approve counterparts. Inconsistent API behavior.",
     "High", "Bug verification", "BUG-PERIOD-4", "OfficePeriodController", "Known DI-92/97. Permission mismatch between symmetric endpoints."),

    ("TC-SEC-046", "Day-off endpoints — all AUTHENTICATED_USER only",
     "API token with VACATIONS_VIEW and VACATIONS_CREATE permissions.",
     "1. GET /v1/employee-dayOff (search) with API token\n2. POST /v1/employee-dayOff with API token\n3. PUT /v1/employee-dayOff/approve/{id} with API token",
     "Verify day-off endpoint auth pattern. Check if Pattern A or Pattern B. If Pattern B, same gap as sick leave module.",
     "High", "Functional", "ENDP-AUTH-005", "EmployeeDayOffController", "Determine actual pattern — not yet verified for all day-off endpoints"),

    ("TC-SEC-047", "Vacation CRUD endpoints accept API token (Pattern A)",
     "API token with VACATIONS_CREATE permission.",
     "1. POST /v1/vacations with valid payload — expect 201\n2. PUT /v1/vacations/approve/{id} — expect 200\n3. PUT /v1/vacations/reject/{id} — expect 200\n4. DELETE /v1/vacations/{id} — expect 200 (if status allows)",
     "All vacation lifecycle endpoints use Pattern A: hasAnyAuthority('AUTHENTICATED_USER', 'VACATIONS_CREATE'). API token access works correctly.",
     "High", "Functional", "ENDP-AUTH-006", "VacationController", "Reference implementation of correct Pattern A usage"),

    ("TC-SEC-048", "Report confirmation endpoints — REPORTS_APPROVE permission required",
     "API token with REPORTS_APPROVE permission.",
     "1. PATCH /v1/reports/{id} with state=APPROVED — expect 200\n2. PATCH /v1/reports (batch) with state changes — expect 200",
     "Confirmation operations work via API token with REPORTS_APPROVE. Pattern A correctly allows both JWT and API token access.",
     "Medium", "Functional", "ENDP-AUTH-007", "TaskReportController", ""),

    ("TC-SEC-049", "Token management endpoints — JWT only",
     "API token (not JWT). Admin API token with all permissions.",
     "1. GET /v1/tokens with API token — expect 403\n2. POST /v1/tokens with API token — expect 403\n3. PATCH /v1/tokens with API token — expect 403\n4. DELETE /v1/tokens with API token — expect 403",
     "All token management endpoints require hasAuthority('AUTHENTICATED_USER') — JWT only. API tokens cannot self-manage. Security-correct design.",
     "Medium", "Functional", "ENDP-AUTH-008", "TokenController", "Pattern B used intentionally — correct for token self-management"),

    ("TC-SEC-050", "Statistics endpoints — 10 JSON + 10 CSV, all with proper auth",
     "API token with STATISTICS_VIEW permission.",
     "1. Test each of 10 JSON statistic endpoints\n2. Test each of 10 CSV mirror endpoints\n3. Verify consistent auth pattern across all 20",
     "All 20 statistics endpoints should use consistent auth pattern. Verify no gaps in the 10+10 mirror pattern.",
     "Medium", "Functional", "ENDP-AUTH-009", "StatisticReportController", "20 endpoints total — verify all have @PreAuthorize"),

    ("TC-SEC-051", "Suggest/autocomplete endpoints — permission mapping",
     "API token with different permission levels.",
     "1. GET /v1/suggest/employees with EMPLOYEES_VIEW — expect 200\n2. GET /v1/suggest/projects with PROJECTS_VIEW — expect 200\n3. GET /v1/suggest/tasks without TASKS_VIEW — expect 403\n4. Verify each suggest endpoint maps to correct permission",
     "6 suggest endpoints each mapped to their respective VIEW permission. Verify consistent enforcement.",
     "Medium", "Functional", "ENDP-AUTH-010", "SuggestionController", "6 suggest endpoints: employees, projects, tasks, customers, offices"),
]))

# ==========================================================================
# TS-SEC-SoDuties -- Segregation of Duties & Self-Approval
# ==========================================================================
SUITES.append(("TS-SEC-SoDuties", "Segregation of Duties & Self-Approval", [
    ("TC-SEC-052", "Vacation self-approval by DEPARTMENT_MANAGER",
     "Logged in as user with ROLE_DEPARTMENT_MANAGER who is their own manager.",
     "1. POST /v1/vacations to create vacation request\n2. Observe auto-assigned approver = self\n3. PUT /v1/vacations/approve/{id}\n4. Check status",
     "BUG: Self-approval succeeds. Same person creates and approves. No segregation of duties check. Approver auto-assigned to self for DM.",
     "Critical", "Security", "SOD-VAC-001", "VacationController, VacationServiceImpl", "Known DI-44. No creator != approver validation."),

    ("TC-SEC-053", "Report self-approval via API — executor approves own report",
     "API token owner (pvaynmaster) is both executor and approver.",
     "1. POST /v1/reports with executorLogin=pvaynmaster\n2. PATCH /v1/reports/{id} with state=APPROVED\n3. Verify approverLogin = executorLogin",
     "BUG: Self-approval works. Same person submits and approves their time report. No executor != approver check.",
     "Critical", "Security", "SOD-RPT-001", "TaskReportController.patch", "Known DI-50. Same pattern as vacation module."),

    ("TC-SEC-054", "Direct create-as-APPROVED bypasses approval workflow",
     "API token with REPORTS_EDIT permission.",
     "1. POST /v1/reports with state=APPROVED in request body\n2. Verify report created in APPROVED state\n3. Check approverLogin = creator",
     "BUG: Report created directly in APPROVED state, completely bypassing the approval workflow. No validation that initial state must be REPORTED.",
     "Critical", "Security", "SOD-RPT-002", "TaskReportController.create", "Known DI-51. Most severe SoD violation."),

    ("TC-SEC-055", "Cross-employee reporting — API token owner reports for another",
     "API token owned by pvaynmaster. Target employee: sterekhin.",
     "1. POST /v1/reports with executorLogin=sterekhin (different from token owner)\n2. Verify report created\n3. Check reporter vs executor fields",
     "BUG: Report created with executor=sterekhin, reporter=pvaynmaster. No validation that API token owner can only report for themselves.",
     "High", "Security", "SOD-RPT-003", "TaskReportController.create", "Known DI-52. Any API token owner can report hours for any employee."),

    ("TC-SEC-056", "Day-off CPO self-approval pattern",
     "User with PROJECT role (CPO). Self is own manager.",
     "1. POST /v1/employee-dayOff to create day-off request\n2. Check request status immediately\n3. Verify auto-approval",
     "Request may be auto-approved on creation for CPO users (approver=self). Same SoD concern as vacation. Verify whether manual approval step is skipped.",
     "High", "Security", "SOD-DO-001", "EmployeeDayOffServiceImpl", "CPO self-approve: approver=self, manager→optional"),

    ("TC-SEC-057", "Sick leave — any user can create for any employee",
     "Logged in as regular employee with no management role. Target: unrelated employee.",
     "1. POST /v1/sick-leaves with login={unrelated-employee}\n2. Check if sick leave created\n3. Verify no relation check",
     "BUG: Sick leave created for any employee regardless of requestor's relationship. No creation permission check beyond AUTHENTICATED_USER.",
     "Critical", "Security", "SOD-SL-001", "SickLeaveServiceImpl.createSickLeave()", "Known DI-26. Access control gap — no org-hierarchy check."),

    ("TC-SEC-058", "Batch report approval — self-approve multiple reports at once",
     "API token owner is executor of multiple pending reports.",
     "1. POST multiple reports with executorLogin=self\n2. PATCH /v1/reports (batch) with state=APPROVED for all\n3. Verify all approved by self",
     "BUG: Batch self-approval works. Self-approval bug compounds with batch operations — single API call approves many self-reported hours.",
     "High", "Security", "SOD-RPT-004", "TaskReportController.batchPatch", "Amplifies DI-50 impact — batch scale self-approval"),

    ("TC-SEC-059", "Re-approval of rejected vacation without edit",
     "Vacation in REJECTED state.",
     "1. Check permissions field on REJECTED vacation\n2. Verify APPROVE is available\n3. PUT /v1/vacations/approve/{id} without editing\n4. Check status changes to APPROVED",
     "REJECTED vacations can be re-approved without any modification. Permissions still show APPROVE action. May be intentional but bypasses rejection reason.",
     "Medium", "Functional", "SOD-VAC-002", "VacationController", "Potential design issue: rejection can be overridden without addressing reason"),

    ("TC-SEC-060", "Approval change — verify old approver loses approval capability",
     "Vacation with approver=userA. Change to approver=userB.",
     "1. PUT /v1/vacations/change-approver/{id}/{userB}\n2. As userA, attempt to approve — expect 403\n3. As userB, approve — expect success",
     "After approver change, old approver cannot approve. New approver has full approval capability. Old approver moved to optional approvers.",
     "High", "Functional", "SOD-VAC-003", "VacationController.changeApprover", "Proper SoD enforcement after delegation"),
]))

# ==========================================================================
# TS-SEC-InfoLeak -- Information Disclosure & Error Handling
# ==========================================================================
SUITES.append(("TS-SEC-InfoLeak", "Information Disclosure & Error Handling", [
    ("TC-SEC-061", "Stack trace in 400 error — vacation payment dates",
     "Missing required parameter.",
     "1. GET /v1/paymentdates without required params\n2. Inspect response body",
     "BUG: Full Java stack trace returned in 400 response. Exposes class names (VacationPaymentDatesController), method names, package structure, framework versions.",
     "High", "Security", "LEAK-STACK-001", "VacationPaymentDatesController", "Known DI-77. Multiple endpoints affected."),

    ("TC-SEC-062", "Stack trace in 400 error — period patch with invalid date",
     "Invalid date format in PATCH request.",
     "1. PATCH /v1/offices/{id}/periods/report with start='invalid-date'\n2. Inspect response body",
     "BUG: 98+ frame Java stack trace in response body. Reveals OfficePeriodController, Spring internals, Jackson deserialization chain.",
     "High", "Security", "LEAK-STACK-002", "OfficePeriodController", "Known DI-96. Server should return structured error without stack trace."),

    ("TC-SEC-063", "Stack trace in 500 error — vacation status=ALL NPE",
     "status=ALL query parameter.",
     "1. GET /v2/vacations?status=ALL\n2. Inspect 500 response body",
     "BUG: NullPointerException stack trace returned. Reveals VacationRepositoryCustomImpl.buildCommonCondition:433, JPA/Hibernate internals.",
     "Critical", "Bug verification", "LEAK-NPE-001", "VacationRepositoryCustomImpl", "Known accounting API bug. 500 + full stack trace = double issue."),

    ("TC-SEC-064", "CORS allows all origins — cross-origin API request",
     "Request from external origin.",
     "1. From JavaScript on external site (e.g. localhost:8080), make fetch() to TTT API\n2. Check CORS headers in response\n3. Verify Access-Control-Allow-Origin",
     "BUG: CORS enabled for ALL origins without restriction. Access-Control-Allow-Origin: * (or reflects any origin). Any website can make cross-origin API requests.",
     "Critical", "Security", "LEAK-CORS-001", "SecurityConfiguration (CORS config)", "Known DI-47. Enables CSRF-style attacks if tokens are accessible."),

    ("TC-SEC-065", "API token visible in URL query parameter",
     "Valid API token.",
     "1. Make API call with ?token={secret_key} in URL\n2. Check server access logs for token value\n3. Check browser history for URL with token\n4. Check HTTP Referer header on subsequent navigation",
     "Token accepted via URL but visible in: server access logs, browser history, HTTP Referer headers, proxy logs, CDN logs. Information disclosure risk.",
     "High", "Security", "LEAK-TOKEN-001", "ApiTokenAuthenticationFilter", "Known DI-46. Token exposure in URL. Should be header-only."),

    ("TC-SEC-066", "Credentials committed to source control",
     "Access to Git repository.",
     "1. Check vacation/app/application.yml in repository\n2. Look for api-token and password values\n3. Verify if values are actual secrets vs placeholders",
     "BUG: API token value and DB password (password=123456) committed to source control in application.yml. Should use environment variables or vault.",
     "Critical", "Security", "LEAK-CRED-001", "vacation/app/application.yml", "Known DI-4. Credentials in source code."),

    ("TC-SEC-067", "Error response format inconsistency — TTT vs Vacation service",
     "Trigger errors on both services.",
     "1. Trigger 400 on TTT service — check response format\n2. Trigger 400 on Vacation service — check response format\n3. Compare error structures",
     "TTT returns structured errors with errorCode. Vacation sometimes includes 'trace' field, sometimes doesn't. Non-existent vacation IDs return 400 (not 404), non-existent offices return 200 with default data.",
     "Medium", "Functional", "LEAK-ERR-001", "TTT ErrorHandler vs Vacation ErrorHandler", "Inconsistent error handling across services"),

    ("TC-SEC-068", "Invalid office ID returns 200 with default period data",
     "Non-existent office ID.",
     "1. GET /v1/offices/99999/periods/report\n2. GET /v1/offices/0/periods/report\n3. GET /v1/offices/-1/periods/report\n4. Check responses",
     "BUG: All return 200 with default period data instead of 404. Silent fallback masks invalid input. Could cause downstream logic errors.",
     "Medium", "Bug verification", "LEAK-VAL-001", "OfficePeriodController", "Known accounting API bug. Invalid input accepted silently."),

    ("TC-SEC-069", "GraalVM sandbox bypass potential — tracker script injection",
     "Admin access to configure tracker integration scripts.",
     "1. Create custom script for tracker with obfuscated loop (avoiding keyword detection)\n2. Attempt script variants: Unicode homoglyphs for 'while'/'for', eval-based construction\n3. Check if infinite loop or resource exhaustion is possible",
     "GraalVM JavaScript sandbox uses string keyword matching (while, for, do, goto) not AST analysis. Potentially bypassable with Unicode or eval tricks. Resource limit should prevent infinite execution.",
     "High", "Security", "LEAK-SANDBOX-001", "CustomScriptService", "Known DI-110. String matching vs AST analysis for loop prevention."),

    ("TC-SEC-070", "Vacation days list — no pagination on 1609 records",
     "API call to vacation days endpoint.",
     "1. GET /v1/vacationdays — returns all employees\n2. Check response size\n3. Check for pagination parameters",
     "Returns 1609 employee records in single response without pagination. Potential for information over-exposure — any authenticated user can see all employees' vacation day balances.",
     "Medium", "Security", "LEAK-DATA-001", "VacationDaysController", "Known accounting finding. No pageSize/page params."),
]))

# ==========================================================================
# TS-SEC-ObjPerm -- Object-Level Permissions
# ==========================================================================
SUITES.append(("TS-SEC-ObjPerm", "Object-Level Permissions", [
    ("TC-SEC-071", "Planner close-tag: admin can create/edit/delete",
     "Logged in as ROLE_ADMIN user. Existing project with tags.",
     "1. POST /v1/planner/close-tags — create new tag for project\n2. PATCH /v1/planner/close-tags/{id} — rename tag\n3. DELETE /v1/planner/close-tags/{id}\n4. Verify all operations succeed",
     "Admin has CREATE+EDIT+DELETE permissions on all project close-tags. All-or-nothing grant model.",
     "High", "Functional", "OBJ-TAG-001", "PlannerCloseTagPermissionService", "#2724 Sprint 15"),

    ("TC-SEC-072", "Planner close-tag: project manager can manage own project tags",
     "Logged in as PM of specific project.",
     "1. POST /v1/planner/close-tags for own project — expect 201\n2. POST /v1/planner/close-tags for different project — expect 403\n3. Verify object-level scoping",
     "PM can manage tags for their own project only. Permission granted via employee.id == project.managerId check.",
     "High", "Functional", "OBJ-TAG-002", "PlannerCloseTagPermissionService", ""),

    ("TC-SEC-073", "Planner close-tag: senior manager and project owner access",
     "Logged in as senior manager (project.seniorManagerId) or owner (project.ownerId).",
     "1. As senior manager, create tag for project — expect 201\n2. As project owner, create tag — expect 201\n3. Verify both roles get full CREATE+EDIT+DELETE",
     "Both senior manager and project owner have full tag management permissions. Same all-or-nothing model as admin and PM.",
     "Medium", "Functional", "OBJ-TAG-003", "PlannerCloseTagPermissionService", ""),

    ("TC-SEC-074", "Planner close-tag: regular employee denied (403)",
     "Logged in as regular employee who is NOT PM/senior manager/owner/admin.",
     "1. GET /v1/planner/close-tags — list tags — expect 200 (read allowed)\n2. POST /v1/planner/close-tags — create — expect 403\n3. PATCH /v1/planner/close-tags/{id} — edit — expect 403\n4. DELETE /v1/planner/close-tags/{id} — delete — expect 403",
     "Employee can list/view tags (read access) but cannot create, edit, or delete. Empty permission set returned.",
     "High", "Functional", "OBJ-TAG-004", "PlannerCloseTagPermissionService", "Key test: non-authorized user blocked"),

    ("TC-SEC-075", "Planner close-tag: readOnly user denied write operations",
     "Logged in as user with readOnly=true flag (e.g. VIEW_ALL).",
     "1. Attempt to create close-tag — expect 403\n2. Attempt to edit — expect 403\n3. Attempt to delete — expect 403",
     "readOnly users get empty permission set regardless of role. Write operations blocked.",
     "Medium", "Functional", "OBJ-TAG-005", "PlannerCloseTagPermissionService", "readOnly check comes before role check"),

    ("TC-SEC-076", "Planner close-tag: cross-project manipulation blocked",
     "PM of Project A. Tag belongs to Project B.",
     "1. PATCH /v1/planner/close-tags/{tagOfProjectB} with projectId=A — expect error\n2. DELETE /v1/planner/close-tags/{tagOfProjectB} — expect 403",
     "Cross-project validation prevents modifying tags from other projects. Tag verified to belong to specified project.",
     "High", "Security", "OBJ-TAG-006", "PlannerCloseTagServiceImpl", "Cross-project guard in delete/update"),

    ("TC-SEC-077", "Vacation approval — only assigned approver can approve",
     "Vacation assigned to approver=userA. Logged in as userB (PM but not assigned).",
     "1. PUT /v1/vacations/approve/{id} as userB (not assigned approver)\n2. Check response",
     "Approval denied. Only the assigned primary approver (or admin) can approve. Object-level permission on vacation entity.",
     "High", "Functional", "OBJ-VAC-001", "VacationServiceImpl", "Object-level: vacation.approver check"),

    ("TC-SEC-078", "Vacation permissions field varies by status and role",
     "Different users viewing same vacation at different statuses.",
     "1. As owner, check permissions on NEW vacation — expect EDIT, DELETE\n2. As approver, check permissions — expect APPROVE, REJECT, CHANGE_APPROVER\n3. On PAID vacation, check permissions — expect empty (immutable)\n4. As unrelated user — check permissions — expect empty",
     "Permissions field dynamically computed per user+status combination. PAID status = fully immutable (empty permissions for all).",
     "High", "Functional", "OBJ-VAC-002", "VacationPermissionService", "Dynamic permission resolution per request"),

    ("TC-SEC-079", "Report approval — only assigned managers can approve in UI",
     "Logged in as PM. Viewing confirmation page.",
     "1. Navigate to /approve\n2. Check which employees' reports are visible\n3. Verify only reports from project members (PM's projects) are shown",
     "Confirmation page scoped to PM's project members only. Object-level filtering based on project_member relationship.",
     "Medium", "Functional", "OBJ-RPT-001", "ConfirmationController, frontend-approve", "UI-level scoping — API-level may differ"),
]))

# ==========================================================================
# TS-SEC-InputVal -- Security-Relevant Input Validation
# ==========================================================================
SUITES.append(("TS-SEC-InputVal", "Security-Relevant Input Validation", [
    ("TC-SEC-080", "Report effort upper bound — no maximum validation",
     "API token with REPORTS_EDIT.",
     "1. POST /v1/reports with effort=1500 (25 hours)\n2. POST /v1/reports with effort=99999 (1666 hours)\n3. Check both accepted\n4. Verify summary counts",
     "BUG: No upper bound validation on effort field. Values exceeding 24h/day accepted. Min validated (>=1 min), max is not.",
     "High", "Bug verification", "VAL-RPT-001", "TaskReportController.create", "Known BUG-REPORT-1. Minimum enforced, maximum missing."),

    ("TC-SEC-081", "Report future date — accepted without restriction",
     "API token with REPORTS_EDIT.",
     "1. POST /v1/reports with reportDate=2026-12-31 (9 months ahead)\n2. Verify accepted\n3. Check period validation (only closed periods blocked)",
     "Future dates accepted without restriction. Only closed-period dates rejected. No forward-looking date limit.",
     "Medium", "Bug verification", "VAL-RPT-002", "TaskReportController.create", "Known BUG-REPORT-5. May be intentional design."),

    ("TC-SEC-082", "Vacation payment — reversed date range accepted",
     "Valid auth.",
     "1. GET /v1/paymentdates?startDate=2026-06-01&endDate=2026-01-01 (end before start)\n2. Check response",
     "BUG: Reversed dates accepted without validation. Response may contain incorrect payment dates or unexpected results.",
     "Medium", "Bug verification", "VAL-PAY-001", "VacationPaymentDatesController", "Known accounting bug."),

    ("TC-SEC-083", "Day-off past personalDate accepted",
     "Valid auth, existing public holiday.",
     "1. PATCH /v1/employee-dayOff/{id} with personalDate=2026-02-01 (past date)\n2. Check response",
     "BUG: Past date accepted. No server-side validation that personalDate >= today. May allow backdating compensatory days.",
     "High", "Bug verification", "VAL-DO-001", "EmployeeDayOffController", "Known BUG-DO-4."),

    ("TC-SEC-084", "Day-off weekend personalDate accepted",
     "Valid auth.",
     "1. POST /v1/employee-dayOff with personalDate on Saturday/Sunday\n2. Check response",
     "BUG: Weekend date accepted. No working-day validation. Employee could take compensatory day on non-working day.",
     "Medium", "Bug verification", "VAL-DO-002", "EmployeeDayOffController", "Known BUG-DO-5."),

    ("TC-SEC-085", "Period PATCH with null start — NPE",
     "Admin/accountant auth.",
     "1. PATCH /v1/offices/{id}/periods/report with body {} (null start)\n2. Check response",
     "BUG: 500 NullPointerException. DTO has @NotNull annotation but @Valid missing on @RequestBody. Server crashes instead of 400 validation error.",
     "High", "Bug verification", "VAL-PERIOD-001", "OfficePeriodController", "Known period bug. @NotNull present but @Valid missing."),

    ("TC-SEC-086", "Approve period — first-day-of-month validation missing",
     "Admin/accountant auth.",
     "1. PATCH /v1/offices/{id}/periods/approve with start=2026-03-15 (mid-month)\n2. Compare with PATCH periods/report with same date — expect 400",
     "BUG: Approve period accepts any day of month. Report period validates first-day-of-month. Inconsistent validation between symmetric endpoints.",
     "High", "Bug verification", "VAL-PERIOD-002", "OfficePeriodServiceImpl", "Known period bug-1. Report validates at line 91, approve missing at line 104."),

    ("TC-SEC-087", "Day-off findAll without type parameter — NPE",
     "Valid auth.",
     "1. GET /v1/employee-dayOff without type parameter\n2. Check response",
     "BUG: 500 NullPointerException at EmployeeDayOffSearchServiceImpl.java:134. ordinal() called on null EmployeeDayOffTypeFilter. Server crashes on missing optional parameter.",
     "High", "Bug verification", "VAL-DO-003", "EmployeeDayOffSearchServiceImpl", "Known BUG-DO-1. NPE on null enum."),

    ("TC-SEC-088", "Day-off list endpoint — NPE in cache",
     "Valid auth.",
     "1. GET /v1/employee-dayOff/list\n2. Check response",
     "BUG: 500 NPE in Caffeine cache computeIfAbsent (InternalEmployeeService.java:160). Null key passed to cache.",
     "High", "Bug verification", "VAL-DO-004", "InternalEmployeeService", "Known BUG-DO-2."),

    ("TC-SEC-089", "Vacation create — null paymentMonth causes NPE",
     "Valid auth.",
     "1. POST /v1/vacations with paymentMonth omitted (null)\n2. Check response",
     "BUG: NPE at VacationAvailablePaidDaysCalculatorImpl.java:73 — paymentDate.getYear() on null. Field documented as optional but effectively required.",
     "High", "Bug verification", "VAL-VAC-001", "VacationAvailablePaidDaysCalculatorImpl", "Known BUG-VACATION-2."),

    ("TC-SEC-090", "Vacation create — null optionalApprovers causes NPE",
     "Valid auth.",
     "1. POST /v1/vacations without optionalApprovers field (null)\n2. Check response",
     "BUG: NPE at VacationServiceImpl.java:155 — getOptionalApprovers().add() on null list. Service assumes list is initialized.",
     "High", "Bug verification", "VAL-VAC-002", "VacationServiceImpl", "Known BUG-VACATION-3."),

    ("TC-SEC-091", "SQL injection attempt on search endpoints",
     "Valid auth.",
     "1. GET /v1/suggest/employees?query='; DROP TABLE employee;--\n2. GET /v1/reports?executorLogin='; SELECT 1;--\n3. Check responses for DB errors",
     "No SQL injection. Spring Data JPA uses parameterized queries. Verify: no raw SQL concatenation, proper escaping, no DB error messages in response.",
     "Medium", "Security", "VAL-SQLI-001", "Various controllers", "Standard OWASP check. JPA should prevent, but verify."),

    ("TC-SEC-092", "XSS attempt in comment/reason fields",
     "Valid auth.",
     "1. POST /v1/reports with comment='<script>alert(1)</script>'\n2. POST /v1/vacations with comment containing HTML tags\n3. View data in UI — check if script executes",
     "No XSS execution. React auto-escapes JSX output. Verify: stored data is rendered safely, no dangerouslySetInnerHTML on user data.",
     "Medium", "Security", "VAL-XSS-001", "React frontend, various DTOs", "React provides default XSS protection via JSX escaping."),
]))

# ==========================================================================
# Risk Assessment Data
# ==========================================================================
RISKS = [
    ("Self-approval across modules (DI-44, DI-50, DI-51)", "Department managers can approve their own vacations. Any API user can self-approve reports or create reports directly as APPROVED. No segregation of duties enforcement.",
     "High", "Critical", "Critical",
     "Test self-approval flows in vacation, reports, and day-off modules. Verify creator != approver validation exists (or document absence). Test direct create-as-APPROVED bypass."),

    ("Sick leave access control gap (DI-26, DI-45)", "Any authenticated user can create sick leave for any employee. Sick leave CRUD uses Pattern B (JWT only) — blocks API automation entirely.",
     "High", "Critical", "Critical",
     "Test sick leave creation across user relationships. Verify API token access denied for CRUD but allowed for search/count."),

    ("Accounting routes unprotected (GAP-ACCT)", "4 of 5 accounting subroutes lack frontend permission checks. Backend requires auth but UI page renders for any user.",
     "High", "High", "Critical",
     "Navigate to /accounting/* routes as regular employee. Verify page renders but data may fail to load."),

    ("CORS unrestricted (DI-47)", "Cross-origin requests accepted from any origin. Combined with token exposure, enables cross-site data theft.",
     "Medium", "Critical", "Critical",
     "Test cross-origin fetch from external origin. Check Access-Control-Allow-Origin headers."),

    ("Credentials in source control (DI-4)", "API token and DB password committed to application.yml. Accessible to anyone with repo access.",
     "High", "Critical", "Critical",
     "Review vacation/app/application.yml for hardcoded secrets. Verify environment-based config used in production."),

    ("Stack trace information disclosure (DI-77, DI-96)", "Multiple endpoints return full Java stack traces in error responses, exposing class names, packages, framework versions.",
     "High", "High", "High",
     "Trigger errors on vacation, period, and report endpoints. Verify no stack traces in 400/500 responses."),

    ("Missing @PreAuthorize on endpoints (DI-53)", "Two report endpoints (/effort, /employee-projects) have no authorization annotation. Any authenticated request can access.",
     "High", "Medium", "High",
     "Test both endpoints with minimal-permission API token. Verify unrestricted access."),

    ("API token in URL query param (DI-46)", "Tokens accepted via ?token= query parameter. Visible in server logs, browser history, HTTP Referer.",
     "High", "High", "High",
     "Test token in URL. Check server logs for token presence."),

    ("Permission inconsistency — report vs approve period (DI-92)", "Report period min/max requires JWT only while approve period min/max accepts API token. Symmetric endpoints with asymmetric auth.",
     "Medium", "Medium", "Medium",
     "Test both endpoint pairs with API token. Verify inconsistent 403 behavior."),

    ("GraalVM sandbox bypass potential (DI-110)", "Tracker script sandbox uses string keyword matching not AST analysis. Potentially bypassable with Unicode or eval tricks.",
     "Low", "Critical", "Medium",
     "Attempt script injection via admin tracker config. Test keyword evasion techniques."),

    ("Cross-employee reporting (DI-52)", "API token owner can create reports for other employees without authorization check.",
     "Medium", "High", "High",
     "Test POST /reports with different executorLogin. Verify no ownership check."),

    ("Contractor permissions undefined (GAP-CONTRACTOR)", "ROLE_CONTRACTOR has no explicit permissions in any PermissionProvider. Access model unclear.",
     "Medium", "Medium", "Medium",
     "Login as contractor-only user. Map actual accessible pages and API endpoints."),

    ("Sick leave route TODO (DI-33)", "Frontend /sick-leave route missing permission check. TODO comment in code — planned but not implemented.",
     "High", "Low", "Medium",
     "Navigate to /sick-leave as employee without sick leave permissions. Verify page loads."),
]

# ==========================================================================
# Feature Matrix Data
# ==========================================================================
# (feature, functional, negative, boundary, security, bug_verification, integration, data_integrity)
FEATURE_MATRIX = [
    ("JWT Authentication", 8, 4, 0, 0, 0, 0, 0),
    ("API Token Authentication", 7, 1, 0, 4, 0, 0, 0),
    ("Role-Based Page Access", 10, 0, 0, 4, 0, 0, 0),
    ("Endpoint Authorization", 9, 0, 0, 4, 0, 0, 0),
    ("Segregation of Duties", 3, 0, 0, 6, 0, 0, 0),
    ("Information Disclosure", 2, 0, 0, 5, 3, 0, 0),
    ("Object-Level Permissions", 6, 0, 0, 3, 0, 0, 0),
    ("Input Validation (Security)", 0, 0, 0, 2, 11, 0, 0),
]


# ==========================================================================
# WORKBOOK GENERATION
# ==========================================================================
def build_workbook():
    wb = openpyxl.Workbook()

    # -- Plan Overview -------------------------------------------------------
    ws = wb.active
    ws.title = "Plan Overview"
    ws.sheet_properties.tabColor = TAB_GREEN

    total_cases = sum(len(s[2]) for s in SUITES)

    plan_rows = [
        ("Test Plan: Security & Permissions (Cross-Cutting)", ""),
        ("", ""),
        ("Document", "security.xlsx"),
        ("Module", "Security & Permissions — Authentication, Authorization, Access Control (Cross-Cutting)"),
        ("Version", "1.0"),
        ("Date", date.today().isoformat()),
        ("Branch", "release/2.1"),
        ("Environment", "timemachine (primary dev), qa-1 (secondary dev), stage (prod baseline)"),
        ("", ""),
        ("SCOPE & OBJECTIVES", ""),
        ("", ""),
        ("Scope", "Cross-cutting security testing across all TTT services. "
                  "Three auth mechanisms: JWT (CAS login), API token (external integration), CAS SSO. "
                  "11 global roles (employee, contractor, PM, DM, TL, accountant, chief accountant, "
                  "director, HR, admin, view_all). "
                  "233 endpoints across 4 services (TTT 128, Vacation 79, Calendar 21, Email 5). "
                  "Three authorization patterns: Pattern A (dual JWT/API), Pattern B (JWT only), "
                  "Pattern C (compound role+authority). "
                  "Object-level permissions (planner close-tag, vacation approval, report confirmation). "
                  "21 known security design issues. "
                  "21 ApiPermission enum values for API token scoping."),
        ("Out of scope", "Performance testing (covered in separate plan), "
                         "email notification content (covered in email module), "
                         "deployment/infrastructure security, penetration testing beyond application layer."),
        ("", ""),
        ("Objectives", "1. Verify JWT authentication lifecycle (CAS login, token generation, expiry, claims)\n"
                       "2. Verify API token authentication (header/URL, permission resolution, 21 permission types)\n"
                       "3. Validate role-based page access for all 11 roles × 23+ pages\n"
                       "4. Test 3 endpoint authorization patterns (A/B/C) across 233 endpoints\n"
                       "5. Detect segregation of duties violations (self-approval in vacation, reports, day-off)\n"
                       "6. Verify information disclosure controls (stack traces, CORS, credentials, error formats)\n"
                       "7. Test object-level permissions (planner tags, vacation approval, report confirmation)\n"
                       "8. Validate security-relevant input handling (injection, XSS, NPE triggers)"),
        ("", ""),
        ("APPROACH", ""),
        ("", ""),
        ("Test types", f"Functional ({sum(r[1] for r in FEATURE_MATRIX)}), "
                       f"Negative ({sum(r[2] for r in FEATURE_MATRIX)}), "
                       f"Boundary ({sum(r[3] for r in FEATURE_MATRIX)}), "
                       f"Security ({sum(r[4] for r in FEATURE_MATRIX)}), "
                       f"Bug verification ({sum(r[5] for r in FEATURE_MATRIX)})"),
        ("Total cases", f"{total_cases} test cases across {len(SUITES)} test suites"),
        ("", ""),
        ("AUTH ARCHITECTURE REFERENCE", ""),
        ("", ""),
        ("Authentication", "Filter chain: JwtTokenAuthFilter → ApiTokenAuthFilter → Authorization (first match wins).\n"
                           "JWT: header TTT_JWT_TOKEN, RS256, 1-day expiry, no refresh. Grants AUTHENTICATED_USER.\n"
                           "API: header API_SECRET_TOKEN (also ?token= query). Only grants ApiPermission enum values.\n"
                           "CAS: SSO via CasAuthenticationFilter → generates JWT for subsequent calls."),
        ("Authorization patterns", "Pattern A (correct): hasAnyAuthority('AUTHENTICATED_USER', 'SPECIFIC_PERM') — both JWT and API token.\n"
                                   "Pattern B (restrictive): hasAuthority('AUTHENTICATED_USER') — JWT only, blocks API tokens.\n"
                                   "Pattern C (compound): hasAnyRole(...) AND hasAnyAuthority(...) — calendar service only."),
        ("", ""),
        ("TEST DATA STRATEGY", ""),
        ("", ""),
        ("Test data", "Timemachine env: 1833 EMPLOYEE tokens (avg 21 perms), 23 APPLICATION tokens (avg 4.7 perms). "
                      "11 global roles: EMPLOYEE(1683), CONTRACTOR(159), PM(136), HR(50), DM(29), TL(19), ACC(18), "
                      "VALL(13), ADM(8), CACC(2), DIR(1). "
                      "Key test users: pvaynmaster (multi-role: PM+DM+EMPLOYEE), perekrest (admin+accountant), "
                      "asmirnov (regular employee, Cyprus office). "
                      "SQL queries for test data:\n"
                      "  SELECT r.role, COUNT(*) FROM employee_global_roles r GROUP BY r.role ORDER BY COUNT(*) DESC\n"
                      "  SELECT t.token_type, COUNT(*) FROM token t GROUP BY t.token_type\n"
                      "  SELECT tp.apipermission, COUNT(*) FROM token_permissions tp GROUP BY tp.apipermission\n"
                      "  SELECT login FROM employee WHERE id IN (SELECT employee_id FROM employee_global_roles WHERE role = 'ROLE_ADMIN')"),
        ("", ""),
        ("ENVIRONMENT REQUIREMENTS", ""),
        ("", ""),
        ("Environment", "Timemachine (primary): test clock manipulation, all roles available, 233 API endpoints.\n"
                        "QA-1 (secondary): same build, cross-environment permission consistency check.\n"
                        "Stage (prod baseline): verify security controls match dev environments."),
        ("", ""),
        ("QASE GAP ANALYSIS", ""),
        ("", ""),
        ("Existing Qase", "0 security-focused test cases. 10 API key cases (basic token operations only). "
                          "0 RBAC/permissions tests. 0 cross-browser/responsive/performance tests. "
                          "No dedicated security or auth test suites."),
        ("Gap", "Complete security coverage needed: authentication flows, 11-role RBAC matrix, "
                "3 authorization patterns, segregation of duties, information disclosure, "
                "object-level permissions, input validation. This workbook fills 100% of security testing gaps."),
        ("", ""),
        ("TEST SUITES", ""),
    ]

    for r, (k, v) in enumerate(plan_rows, 1):
        ws.cell(row=r, column=1, value=k).font = SUBTITLE_FONT if k.isupper() else (TITLE_FONT if r == 1 else BODY_FONT)
        ws.cell(row=r, column=2, value=v).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP

    link_start = len(plan_rows) + 1
    for i, (tab, title, cases) in enumerate(SUITES):
        r = link_start + i
        cell = ws.cell(row=r, column=1)
        cell.value = f"{tab} -- {title}"
        cell.hyperlink = f"#'{tab}'!A1"
        cell.font = LINK_FONT
        ws.cell(row=r, column=2, value=f"{len(cases)} cases").font = BODY_FONT

    set_col_widths(ws, [45, 110])

    # -- Feature Matrix -------------------------------------------------------
    ws_fm = wb.create_sheet("Feature Matrix")
    ws_fm.sheet_properties.tabColor = TAB_GREEN

    fm_headers = ["Feature Area", "Functional", "Negative", "Boundary",
                  "Security", "Bug Verification", "Integration", "Data Integrity", "Total"]
    for c, h in enumerate(fm_headers, 1):
        ws_fm.cell(row=1, column=c, value=h)
    style_header_row(ws_fm, 1, len(fm_headers))

    for idx, (feat, func, neg, bnd, sec, bug, intg, di) in enumerate(FEATURE_MATRIX):
        r = idx + 2
        total = func + neg + bnd + sec + bug + intg + di
        vals = [feat, func, neg, bnd, sec, bug, intg, di, total]
        for c, v in enumerate(vals, 1):
            cell = ws_fm.cell(row=r, column=c, value=v)
            if c == 1:
                cell.hyperlink = f"#'{SUITES[idx][0]}'!A1"
                cell.font = LINK_FONT
        style_data_row(ws_fm, r, len(fm_headers), idx)

    total_r = len(FEATURE_MATRIX) + 2
    ws_fm.cell(row=total_r, column=1, value="TOTAL").font = Font(name=ARIAL, bold=True, size=11)
    for c in range(2, len(fm_headers) + 1):
        total_val = sum(ws_fm.cell(row=r, column=c).value or 0 for r in range(2, total_r))
        ws_fm.cell(row=total_r, column=c, value=total_val).font = Font(name=ARIAL, bold=True, size=11)
    for c in range(1, len(fm_headers) + 1):
        ws_fm.cell(row=total_r, column=c).border = THIN_BORDER
        ws_fm.cell(row=total_r, column=c).fill = PatternFill("solid", fgColor="E2EFDA")

    set_col_widths(ws_fm, [35, 12, 12, 12, 12, 16, 14, 15, 10])
    ws_fm.auto_filter.ref = f"A1:{get_column_letter(len(fm_headers))}{total_r}"

    # -- Risk Assessment -------------------------------------------------------
    ws_ra = wb.create_sheet("Risk Assessment")
    ws_ra.sheet_properties.tabColor = TAB_GREEN

    ra_headers = ["Risk", "Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    for c, h in enumerate(ra_headers, 1):
        ws_ra.cell(row=1, column=c, value=h)
    style_header_row(ws_ra, 1, len(ra_headers))

    severity_fills = {"Critical": RISK_CRIT, "High": RISK_HIGH, "Medium": RISK_MED, "Low": RISK_LOW}

    for idx, (risk, desc, likelihood, impact, severity, mitigation) in enumerate(RISKS):
        r = idx + 2
        vals = [risk, desc, likelihood, impact, severity, mitigation]
        for c, v in enumerate(vals, 1):
            ws_ra.cell(row=r, column=c, value=v)
        style_data_row(ws_ra, r, len(ra_headers), idx)
        sev_fill = severity_fills.get(severity)
        if sev_fill:
            ws_ra.cell(row=r, column=5).fill = sev_fill

    set_col_widths(ws_ra, [45, 60, 14, 14, 14, 60])
    ws_ra.auto_filter.ref = f"A1:{get_column_letter(len(ra_headers))}{len(RISKS) + 1}"

    # -- Test Suite Tabs -------------------------------------------------------
    tc_headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
                  "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    tc_widths = [14, 40, 35, 50, 50, 10, 16, 22, 30, 30]

    for tab_name, suite_title, cases in SUITES:
        ws_ts = wb.create_sheet(tab_name)
        ws_ts.sheet_properties.tabColor = TAB_BLUE

        back = ws_ts.cell(row=1, column=1, value="\u2190 Back to Plan")
        back.hyperlink = "#'Plan Overview'!A1"
        back.font = BACK_LINK_FONT

        ws_ts.cell(row=2, column=1, value=f"{tab_name}: {suite_title}").font = TITLE_FONT

        header_row = 4
        for c, h in enumerate(tc_headers, 1):
            ws_ts.cell(row=header_row, column=c, value=h)
        style_header_row(ws_ts, header_row, len(tc_headers))

        for idx, case in enumerate(cases):
            r = header_row + 1 + idx
            for c, v in enumerate(case, 1):
                ws_ts.cell(row=r, column=c, value=v)
            style_data_row(ws_ts, r, len(tc_headers), idx)

        set_col_widths(ws_ts, tc_widths)
        last_data = header_row + len(cases)
        ws_ts.auto_filter.ref = f"A{header_row}:{get_column_letter(len(tc_headers))}{last_data}"

    wb.save(OUTPUT)
    print(f"Generated {OUTPUT}")
    print(f"  Tabs: Plan Overview + Feature Matrix + Risk Assessment + {len(SUITES)} TS-* suites")
    print(f"  Total test cases: {total_cases}")
    print(f"  Risks: {len(RISKS)}")
    for tab, title, cases in SUITES:
        priorities = {}
        for c in cases:
            p = c[5]
            priorities[p] = priorities.get(p, 0) + 1
        pstr = "/".join(f"{priorities.get(k, 0)}" for k in ["Critical", "High", "Medium", "Low"])
        print(f"    {tab}: {len(cases)} cases (C/H/M/L: {pstr})")


if __name__ == "__main__":
    build_workbook()
